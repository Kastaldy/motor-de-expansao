"""Export do simulador de viabilidade para Excel (BLK-DIM-22).

READ-ONLY sobre o M1: nao recalcula score_priorizacao, pesos nem artefatos (DEC-001).
Sem I/O de disco: usa BytesIO exclusivamente.
"""
from __future__ import annotations

import math
from io import BytesIO
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from motor_expansao.dimensionamento.viabilidade_ponto import ViabilidadePontoResult

# --- Paleta Ultra ---
_TURQUESA_HEX = "FF00BFB3"
_BRANCO_HEX = "FFFFFFFF"
_CINZA_ESC = "FF2E3040"
_CINZA_CLR = "FFF5F5F5"
_VERDE_CLR = "FF90EE90"
_AMARELO_CLR = "FFFFFF99"
_VERMELHO_CLR = "FFFFC0C0"
_FONTE_PADRAO = "Calibri"
_FMT_BRL = "R$ #,##0.00"
_FMT_PCT = "0.0%"

# Prefixos que o Excel interpreta como formula viva ao abrir a planilha (=/+/-/@,
# tab e CR). Mesmo conjunto de `dashboard/rede_export._celula` -- ver comentario la'.
_INICIO_DE_FORMULA = ("=", "+", "-", "@", chr(9), chr(13))


def _texto_seguro(texto: str) -> str:
    """Neutraliza injecao de formula (pentest Onda B #12): um `nome_ponto` como
    `=HYPERLINK("http://evil","x")` vira formula viva no XLSX. Prefixar `'` faz o
    Excel tratar a celula como texto. Duplicado (nao importado) de proposito: o
    pacote `dimensionamento` nao deve depender de `dashboard`.
    """
    if texto[:1] in _INICIO_DE_FORMULA:
        return "'" + texto
    return texto


def _fill(hex_argb: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_argb)


def _header_font(bold: bool = True) -> Font:
    return Font(name=_FONTE_PADRAO, bold=bold, color=_BRANCO_HEX, size=12)


def _body_font(bold: bool = False) -> Font:
    return Font(name=_FONTE_PADRAO, bold=bold, color=_CINZA_ESC, size=10)


def _set_col_width(ws: openpyxl.worksheet.worksheet.Worksheet, col_idx: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_header(ws: openpyxl.worksheet.worksheet.Worksheet, titulo: str, n_cols: int = 2) -> None:
    """Escreve linha de cabecalho turquesa com titulo."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1, value=titulo)
    cell.fill = _fill(_TURQUESA_HEX)
    cell.font = _header_font()
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22


def _write_aba_resumo(
    wb: openpyxl.Workbook,
    result: ViabilidadePontoResult,
    nome_ponto: str,
) -> None:
    ws = wb.create_sheet("Resumo")
    viab = result.viabilidade

    # Linha 1: cabecalho
    _write_header(ws, "ULTRA Academia — Simulador de Viabilidade", n_cols=2)

    # Linha 2: nome_ponto
    if nome_ponto:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
        c = ws.cell(row=2, column=1, value=_texto_seguro(nome_ponto))
        c.fill = _fill(_CINZA_CLR)
        c.font = _body_font(bold=True)

    # Linha 3: vazia (separador)
    ws.row_dimensions[3].height = 6

    # Linha 4+: cabecalho da tabela
    hdr_row = 4
    for col, txt in enumerate(["Indicador", "Valor"], start=1):
        c = ws.cell(row=hdr_row, column=col, value=txt)
        c.fill = _fill(_CINZA_ESC)
        c.font = Font(name=_FONTE_PADRAO, bold=True, color=_BRANCO_HEX, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")

    # Dados
    payback_val: str | float
    if viab.payback_meses == float("inf") or (isinstance(viab.payback_meses, float) and math.isinf(viab.payback_meses)):
        payback_val = "Não atingido"
    else:
        payback_val = viab.payback_meses

    breakeven_val: str | float
    if result.alunos_breakeven == float("inf") or (isinstance(result.alunos_breakeven, float) and math.isinf(result.alunos_breakeven)):
        breakeven_val = "Inviável"
    else:
        breakeven_val = result.alunos_breakeven

    linhas = [
        ("Ponto analisado", f"{result.lat:.6f}, {result.lng:.6f}", None),
        ("Metragem (m²)", int(result.m2), "#,##0"),
        ("Aluguel pedido (R$/mês)", result.aluguel_pedido, _FMT_BRL),
        ("Demanda premissa (alunos)", int(result.demanda_premissa), "#,##0"),
        ("Faturamento bruto/mês", viab.faturamento_mensal_steady, _FMT_BRL),
        ("Receita líquida/mês", viab.receita_liquida, _FMT_BRL),
        ("EBITDA/mês", viab.ebitda_mensal, _FMT_BRL),
        ("Margem EBITDA", viab.margem_ebitda_pct, _FMT_PCT),
        ("Payback (meses)", payback_val, None if isinstance(payback_val, str) else "#,##0.0"),
        ("ROIC anual", viab.roic_anual, _FMT_PCT),
        ("Lucro líquido/mês", viab.lucro_liquido_mensal, _FMT_BRL),
        ("Alunos break-even", breakeven_val, None if isinstance(breakeven_val, str) else "#,##0.0"),
        ("Aluguel-teto", result.aluguel_teto_calculado, _FMT_BRL),
        ("Viável?", "Sim" if viab.flag_viavel else "Não", None),
        ("Fonte da demanda", result.demanda_fonte, None),
    ]

    for i, (indicador, valor, fmt) in enumerate(linhas):
        row = hdr_row + 1 + i
        bg = _CINZA_CLR if i % 2 == 0 else _BRANCO_HEX
        ca = ws.cell(row=row, column=1, value=indicador)
        ca.fill = _fill(bg)
        ca.font = _body_font()
        ca.alignment = Alignment(horizontal="left", vertical="center")

        cb = ws.cell(row=row, column=2, value=valor)
        cb.fill = _fill(bg)
        cb.font = _body_font()
        cb.alignment = Alignment(horizontal="right", vertical="center")
        if fmt:
            cb.number_format = fmt

    _set_col_width(ws, 1, 35)
    _set_col_width(ws, 2, 22)


def _write_aba_dre(
    wb: openpyxl.Workbook,
    result: ViabilidadePontoResult,
) -> None:
    ws = wb.create_sheet("DRE")
    viab = result.viabilidade

    _write_header(ws, "DRE Steady-State", n_cols=3)

    # Cabecalho da tabela
    hdr_row = 2
    for col, txt in enumerate(["Linha DRE", "R$/mês", "% Faturamento"], start=1):
        c = ws.cell(row=hdr_row, column=col, value=txt)
        c.fill = _fill(_CINZA_ESC)
        c.font = Font(name=_FONTE_PADRAO, bold=True, color=_BRANCO_HEX, size=10)
        c.alignment = Alignment(horizontal="left" if col == 1 else "right", vertical="center")

    fat = viab.faturamento_mensal_steady
    deducoes = fat - viab.receita_liquida
    impostos = viab.receita_liquida - viab.receita_pos_impostos
    custos_op_total = viab.receita_pos_impostos - viab.ebitda_mensal
    ir_csll = viab.ebitda_mensal - viab.lucro_liquido_mensal

    def _pct(val: float) -> float:
        return val / fat if fat > 0 else 0.0

    linhas_dre = [
        ("Faturamento Bruto", fat, _pct(fat), False, False),
        ("(−) Deduções", -deducoes, _pct(-deducoes), False, False),
        ("Receita Líquida", viab.receita_liquida, _pct(viab.receita_liquida), False, True),
        ("(−) PIS/COFINS/ISS", -impostos, _pct(-impostos), False, False),
        ("Receita pós-impostos", viab.receita_pos_impostos, _pct(viab.receita_pos_impostos), False, True),
        ("(−) Custos operacionais totais", -custos_op_total, _pct(-custos_op_total), False, False),
        ("EBITDA", viab.ebitda_mensal, viab.margem_ebitda_pct, True, True),
        ("(−) IR/CSLL", -ir_csll, _pct(-ir_csll), False, False),
        ("Lucro Líquido", viab.lucro_liquido_mensal, _pct(viab.lucro_liquido_mensal), False, True),
    ]

    for i, (label, valor_brl, pct_fat, negrito, _borda_topo) in enumerate(linhas_dre):
        row = hdr_row + 1 + i
        bg = _CINZA_CLR if i % 2 == 0 else _BRANCO_HEX
        c1 = ws.cell(row=row, column=1, value=label)
        c1.fill = _fill(bg)
        c1.font = Font(name=_FONTE_PADRAO, bold=negrito, color=_CINZA_ESC, size=10)
        c1.alignment = Alignment(horizontal="left", vertical="center")

        c2 = ws.cell(row=row, column=2, value=valor_brl)
        c2.fill = _fill(bg)
        c2.font = Font(name=_FONTE_PADRAO, bold=negrito, color=_CINZA_ESC, size=10)
        c2.number_format = _FMT_BRL
        c2.alignment = Alignment(horizontal="right", vertical="center")

        c3 = ws.cell(row=row, column=3, value=pct_fat)
        c3.fill = _fill(bg)
        c3.font = Font(name=_FONTE_PADRAO, bold=negrito, color=_CINZA_ESC, size=10)
        c3.number_format = _FMT_PCT
        c3.alignment = Alignment(horizontal="right", vertical="center")

    _set_col_width(ws, 1, 38)
    _set_col_width(ws, 2, 20)
    _set_col_width(ws, 3, 18)


def _write_aba_sensibilidade(
    wb: openpyxl.Workbook,
    result: ViabilidadePontoResult,
) -> None:
    ws = wb.create_sheet("Sensibilidade")
    _write_header(
        ws,
        "Grade de Sensibilidade — Margem EBITDA (alunos x aluguel)",
        n_cols=7,
    )

    grade = result.grade_sensibilidade
    if grade is None or grade.empty:
        ws.cell(row=2, column=1, value="Grade não disponível")
        return

    try:
        pivot = grade.pivot_table(
            index="alunos",
            columns="fator_aluguel",
            values="margem_liq",
            aggfunc="first",
        )
    except Exception:
        ws.cell(row=2, column=1, value="Grade não disponível")
        return

    # Cabeçalho de colunas (linha 2)
    fatores = list(pivot.columns)
    # Linha 2: cabeçalho turquesa leve
    hdr_row = 2
    c0 = ws.cell(row=hdr_row, column=1, value="Alunos")
    c0.fill = _fill(_CINZA_ESC)
    c0.font = Font(name=_FONTE_PADRAO, bold=True, color=_BRANCO_HEX, size=10)
    c0.alignment = Alignment(horizontal="center", vertical="center")

    for j, fator in enumerate(fatores):
        c = ws.cell(row=hdr_row, column=2 + j, value=f"x{fator:.1f}")
        c.fill = _fill(_CINZA_ESC)
        c.font = Font(name=_FONTE_PADRAO, bold=True, color=_BRANCO_HEX, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Dados
    for i, alunos in enumerate(pivot.index):
        row = hdr_row + 1 + i
        bg = _CINZA_CLR if i % 2 == 0 else _BRANCO_HEX
        c_al = ws.cell(row=row, column=1, value=float(alunos))
        c_al.fill = _fill(bg)
        c_al.font = _body_font()
        c_al.alignment = Alignment(horizontal="center", vertical="center")

        for j, fator in enumerate(fatores):
            margem = pivot.loc[alunos, fator]
            if margem is None or (isinstance(margem, float) and math.isnan(margem)):
                bg_cell = bg
            elif margem >= 0.10:
                bg_cell = _VERDE_CLR
            elif margem >= 0.0:
                bg_cell = _AMARELO_CLR
            else:
                bg_cell = _VERMELHO_CLR

            c_val = ws.cell(row=row, column=2 + j, value=float(margem) if margem is not None else None)
            c_val.fill = _fill(bg_cell)
            c_val.font = _body_font()
            c_val.number_format = _FMT_PCT
            c_val.alignment = Alignment(horizontal="center", vertical="center")

    # Larguras de coluna
    _set_col_width(ws, 1, 12)
    for j in range(len(fatores)):
        _set_col_width(ws, 2 + j, 12)


def _write_aba_curva(
    wb: openpyxl.Workbook,
    serie: list[dict],
) -> None:
    ws = wb.create_sheet("Curva")
    _write_header(ws, "Projeção Financeira — 60 meses", n_cols=5)

    if not serie:
        ws.cell(row=2, column=1, value="Série não disponível")
        return

    # Cabeçalho de colunas (linha 2)
    hdr_row = 2
    headers = ["Mês", "Alunos Balcão", "Faturamento (R$)", "EBITDA (R$)", "FCF Acumulado (R$)"]
    for col, txt in enumerate(headers, start=1):
        c = ws.cell(row=hdr_row, column=col, value=txt)
        c.fill = _fill(_CINZA_ESC)
        c.font = Font(name=_FONTE_PADRAO, bold=True, color=_BRANCO_HEX, size=10)
        c.alignment = Alignment(horizontal="center" if col == 1 else "right", vertical="center")

    # 60 linhas de dados
    for i, row_data in enumerate(serie):
        row = hdr_row + 1 + i
        bg = _CINZA_CLR if i % 2 == 0 else _BRANCO_HEX

        c_mes = ws.cell(row=row, column=1, value=int(row_data["mes"]))
        c_mes.fill = _fill(bg)
        c_mes.font = _body_font()
        c_mes.alignment = Alignment(horizontal="center", vertical="center")

        c_al = ws.cell(row=row, column=2, value=int(round(row_data["alunos_balcao"])))
        c_al.fill = _fill(bg)
        c_al.font = _body_font()
        c_al.alignment = Alignment(horizontal="right", vertical="center")

        c_fat = ws.cell(row=row, column=3, value=float(row_data["faturamento_mensal"]))
        c_fat.fill = _fill(bg)
        c_fat.font = _body_font()
        c_fat.number_format = _FMT_BRL
        c_fat.alignment = Alignment(horizontal="right", vertical="center")

        c_eb = ws.cell(row=row, column=4, value=float(row_data["ebitda_mensal"]))
        c_eb.fill = _fill(bg)
        c_eb.font = _body_font()
        c_eb.number_format = _FMT_BRL
        c_eb.alignment = Alignment(horizontal="right", vertical="center")

        c_fcf = ws.cell(row=row, column=5, value=float(row_data["fcf_acumulado"]))
        c_fcf.fill = _fill(bg)
        c_fcf.font = _body_font()
        c_fcf.number_format = _FMT_BRL
        c_fcf.alignment = Alignment(horizontal="right", vertical="center")

    _set_col_width(ws, 1, 8)
    _set_col_width(ws, 2, 16)
    _set_col_width(ws, 3, 20)
    _set_col_width(ws, 4, 20)
    _set_col_width(ws, 5, 22)


def gerar_excel_viabilidade(
    result: ViabilidadePontoResult,
    serie: list[dict],
    *,
    nome_ponto: str = "",
) -> bytes:
    """Gera arquivo .xlsx em memoria com 4 abas do simulador de viabilidade.

    READ-ONLY sobre o M1 (nao recalcula score_priorizacao nem artefatos).
    Sem I/O de disco — usa BytesIO exclusivamente.

    Parameters
    ----------
    result:
        Resultado de analisar_viabilidade_ponto().
    serie:
        Lista de 60 dicts de gerar_serie_mensal() (chaves: mes, alunos_balcao,
        faturamento_mensal, ebitda_mensal, fcf_acumulado).
    nome_ponto:
        Nome opcional do ponto/imovel para o cabecalho do Excel.

    Returns
    -------
    bytes
        Conteudo binario do arquivo .xlsx. NUNCA escreve em disco.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove aba padrao vazia

    _write_aba_resumo(wb, result, nome_ponto)
    _write_aba_dre(wb, result)
    _write_aba_sensibilidade(wb, result)
    _write_aba_curva(wb, serie)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["gerar_excel_viabilidade"]
