"""Simulador financeiro completo em XLSX com FORMULAS VIVAS (FIN-VIAB-01).

Diferenca em relacao a `excel_export.gerar_excel_viabilidade()`: aquele exporta
VALORES estaticos (4 abas, foto do resultado). Este exporta o MODELO — toda
celula de DRE, fluxo de caixa e KPI e uma FORMULA de Excel a partir da aba
`Premissas` e da aba `Folha`. Mudar a demanda, o aluguel ou um salario dentro do
Excel recalcula os 64 meses (M-4..M60) na hora, sem voltar ao sistema.

Pedido do dono do produto (Felipe, 2026-07-24): "precisa ter toda DRE, folha de
pagamento, fluxo de caixa e etc. Para que possamos abrir a planilha na frente do
investidor e defender os numeros."

DUAS DECISOES TOMADAS POR ELE (nao rediscutir aqui):
  (a) FOLHA: quadro de pessoal EDITAVEL cujo total alimenta a DRE. Os salarios
      default sao ESTIMATIVA da equipe de expansao (esse dado nao existe no repo)
      e estao marcados como tal na aba e em nota de celula.
  (b) FORMATO: XLSX com formulas vivas (nao valores).

INTERRUPTOR DE MODO DA FOLHA: `Folha!$B$5` tem validacao de lista
["percentual do faturamento maduro", "quadro de pessoal"] e DEFAULT
"percentual do faturamento maduro", para o arquivo abrir reproduzindo o motor ao
centavo. O investidor pode abrir o detalhe do quadro sem que o arquivo passe a
divergir do sistema por default.

FIN-VIAB-01, 3a rodada (decisoes de Felipe, 2026-07-24) — DUAS MUDANCAS DE PRODUTO
que quebravam formula por formula e foram re-espelhadas aqui:

  (1) FOLHA FIXA DESDE O MES 1. Antes a DRE calculava a folha como
      `folha_pct x faturamento DO MES`, entao ela encolhia junto com a rampa —
      equivalia a supor que se contrata gente na medida em que o aluno entra. Agora
      a folha e dimensionada UMA vez pelo faturamento MADURO (aba Premissas expoe a
      base e o R$ resultante em celulas proprias) e a DRE referencia essa UNICA
      celula em todos os 64 meses, so multiplicada pelo reajuste anual de custos.
      Consequencias nas formulas: o `k` (receita -> EBITDA) NAO subtrai mais a folha
      (`=(1-deducoes)*(1-impostos-cvar)`), e o break-even da aba Resumo passa a somar
      a folha no custo fixo. Os dois modos do interruptor sao agora FIXOS no tempo:
      "percentual do faturamento maduro" e "quadro de pessoal".

  (2) TAXA DE FRANQUIA PARCELADA (default 4x sem juros). O fluxo de caixa mostra a
      PARCELA nos meses de contrato 1..N (M-4..M-1 com N=4), nao a taxa inteira no
      M-4; a aba Investimento ganhou "Parcelas da franquia" e "Valor da parcela".

E um pedido cosmetico do mesmo dia: na aba Resumo, todo valor que e SAIDA DE
DINHEIRO aparece com a FONTE em VERMELHO. Onde o numero pode alternar de sinal
(EBITDA, FCF, VPL, TIR, retorno) a cor vem de formatacao CONDICIONAL, para a
planilha nao mentir quando o valor virar.

FIN-VIAB-01, 4a rodada (decisoes de Felipe, 2026-07-25) — a reconciliacao das TAXAS
e das DUAS OTICAS de retorno, espelhada formula por formula:

  (1) TAXA. A antiga "taxa de desconto" (12% a.a.) sumiu. Entra a TAXA MINIMA DO
      NEGOCIO (25% a.a., `SIM_TAXA_MINIMA_NEGOCIO_AA`) como UNICA taxa editavel; a
      TAXA MINIMA DO SOCIO e DERIVADA por formula na propria aba Premissas
      (`= negocio + (negocio - divida) * divida/aporte`), junto do CUSTO DA DIVIDA
      a.a. (do juros a.m.) e da ALAVANCAGEM. Nao existe celula onde alguem possa
      digitar uma taxa de socio menor que o custo da divida.

  (2) DUAS OTICAS, nunca no mesmo numero. O fluxo de caixa ganhou duas linhas novas:
      FLUXO DO NEGOCIO (EBITDA - IR - investimento do mes - o equipamento financiado
      desembolsado na vespera da abertura) e FLUXO DA DIVIDA (+captacao no mes da
      vespera, -PMT depois). A aba Resumo traz TIR e VPL em PAR rotulado, cada um com
      IRR/NPV sobre a SUA linha e a SUA taxa.

  (3) CHEQUE TOTAL. O pior ponto do caixa acumulado, com o mes, contra o APORTE
      INICIAL (obra + franquia) — que e o novo nome do que se chamava "equity
      aportado". E o numero que decide se o negocio e FINANCIAVEL.

  (4) IDENTIDADES. A aba Afericao ganhou as duas que fecham EXATAS:
      (a) VPL(fluxo da divida @ custo da divida) == 0  -> prova a tabela Price;
      (b) VPL socio @taxa do negocio == VPL negocio + VPL divida, ambos @taxa do
          negocio -> prova que as duas oticas sao a MESMA economia repartida.
      NAO se testa "os dois VPLs coincidem": isso e FALSO com uma taxa de socio unica
      calculada na alavancagem INICIAL, porque o saldo devedor cai a zero ao longo do
      contrato. O residuo dessa comparacao vive na planilha como DIAGNOSTICO rotulado,
      nao como tolerancia escondida.

A aba `Afericao` e a defesa do arquivo: lado a lado, o valor que o MOTOR Python
calculou (estatico, escrito por nos) e o valor que a FORMULA da planilha produz.
Quem abre ve na hora se alguma formula quebrou.

VOCABULARIO DE USUARIO (obrigatorio nos rotulos): "taxa minima do negocio" e "taxa
minima do socio"; "do negocio" no lugar de "desalavancado"; "cheque total" para o pior
ponto do caixa e "aporte inicial" para obra+franquia. Ku/Ke/FCFF/FCFE ficam SO em
docstring e comentario.

READ-ONLY sobre o M1: nao recalcula score_priorizacao, pesos nem artefatos
(DEC-001/DEC-008/DEC-009). Sem I/O de disco: usa BytesIO exclusivamente.
Demanda NUNCA e derivada de lat/lng (DEC-009) — entra como parametro.
"""

from __future__ import annotations

import math
import re
from collections.abc import Mapping
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from motor_expansao.dimensionamento.config import (
    SIM_ALUGUEL_TETO_EXCECAO,
    SIM_ALUGUEL_TETO_IDEAL,
    SIM_ALUGUEL_TETO_TETO,
    SIM_MARGEM_VIAVEL_MIN,
    SIM_PARCELAS_FRANQUIA_DEFAULT,
    SIM_PARCELAS_OBRA_DEFAULT,
    SIM_PAYBACK_VIAVEL_MAX,
    SIM_TAXA_FRANQUIA,
)

# Reuso dos helpers de estilo do export estatico — NAO duplicar paleta nem fontes.
from motor_expansao.dimensionamento.excel_export import (
    _AMARELO_CLR,
    _BRANCO_HEX,
    _CINZA_CLR,
    _CINZA_ESC,
    _FONTE_PADRAO,
    _TURQUESA_HEX,
    _VERDE_CLR,
    _VERMELHO_CLR,
    _body_font,
    _fill,
    _set_col_width,
    _write_header,
)
from motor_expansao.dimensionamento.simulador import (
    IR_MODO_FAIXA,
    Premissas,
    ViabilidadeResult,
    simular,
)

# --- Nomes das abas (identificadores: SEM acento, §2 do CLAUDE.md) -----------
ABA_PREMISSAS = "Premissas"
ABA_FOLHA = "Folha"
ABA_DRE = "DRE mensal"
ABA_FLUXO = "Fluxo de caixa"
ABA_INVESTIMENTO = "Investimento"
ABA_RESUMO = "Resumo"
ABA_SENSIBILIDADE = "Sensibilidade"
ABA_AFERICAO = "Afericao"

ABAS_ESPERADAS = (
    ABA_PREMISSAS,
    ABA_FOLHA,
    ABA_DRE,
    ABA_FLUXO,
    ABA_INVESTIMENTO,
    ABA_RESUMO,
    ABA_SENSIBILIDADE,
    ABA_AFERICAO,
)

# --- Formatos ---------------------------------------------------------------
# Contabil: negativo entre parenteses e em vermelho (regra do comite).
_FMT_CONTABIL = '"R$" #,##0.00;[Red]("R$" #,##0.00)'
_FMT_PCT2 = "0.00%"
_FMT_PCT3 = "0.000%"
_FMT_ALUNOS = "#,##0.0"
_FMT_INT = "#,##0"

_CINZA_TRAVADO = "FFE2E2E8"
# Vermelho de FONTE (o `_VERMELHO_CLR` do export estatico e cor de PREENCHIMENTO).
# Pedido de Felipe (2026-07-24): na aba Resumo, saida de dinheiro em texto vermelho.
_VERMELHO_FONTE = "FFC00000"

_BORDA_FINA = Border(
    left=Side(style="thin", color="FFB0B0BC"),
    right=Side(style="thin", color="FFB0B0BC"),
    top=Side(style="thin", color="FFB0B0BC"),
    bottom=Side(style="thin", color="FFB0B0BC"),
)

# Natureza da celula de premissa.
_EDIT = "editavel"
_DERIV = "derivada"
_FIXO = "estrutural"

_QUEM_OPERADOR = "Operador"
_QUEM_CONTROLADORIA = "Controladoria"
_QUEM_FRANQUEADORA = "Franqueadora"
_QUEM_DERIVADO = "Ninguem (fórmula)"
_QUEM_ESTRUTURAL = "Ninguem (estrutura da planilha)"

# Decomposicao dos outros fixos (comentario do config.py: SIM_OUTROS_FIXOS_MES).
#
# ACHADO E RESOLUCAO: o comentario do `config.py` listava SETE componentes que somam
# R$ 40.150,00, contra os R$ 38.150,00 da constante. A conta mostra onde esta o furo:
#   2.000 + 17.000 + 500 + 14.000 + 2.150 + 2.500          = 38.150  <- a constante
#                                                  + 2.000  = 40.150  <- o comentario
# As SEIS primeiras fecham EXATAMENTE no valor que o motor usa; o "Outros (2.000)" do
# comentario e espurio (entrou no texto e nunca no numero). Entao a constante esta
# certa e o comentario estava errado — corrigido no `config.py` no mesmo commit.
#
# Consequencia pratica: aqui ficam as seis componentes com os valores REDONDOS, sem
# rateio. A alternativa (ratear as sete por 0,950187 para fechar no total) dava
# "IPTU R$ 1.900,37" e "Telefone R$ 475,09" — numeros que ninguem defende linha a
# linha na frente de um investidor, que e exatamente para o que esta planilha existe.
_OUTROS_FIXOS_DECOMP: tuple[tuple[str, str, float], ...] = (
    ("of_iptu", "IPTU", 2_000.0),
    ("of_agua_luz", "Água e luz", 17_000.0),
    ("of_telefone", "Telefone", 500.0),
    ("of_limpeza", "Limpeza", 14_000.0),
    ("of_tecnologia", "Tecnologia", 2_150.0),
    ("of_assessorias", "Assessorias", 2_500.0),
)
_OUTROS_FIXOS_SOMA_DECOMP = sum(v for _k, _l, v in _OUTROS_FIXOS_DECOMP)
# Pontas do range das sublinhas. DERIVADAS da tupla acima de proposito: quando o
# "Outros" espurio saiu da decomposicao, as duas referencias literais a `of_outros`
# ficaram penduradas e o gerador estourou KeyError. Agora acompanham sozinhas.
_OF_KEY_INI = _OUTROS_FIXOS_DECOMP[0][0]
_OF_KEY_FIM = _OUTROS_FIXOS_DECOMP[-1][0]

# Quadro de pessoal default de unidade low-cost. SALARIOS = ESTIMATIVA (o repo nao
# tem folha real). Calibrado para o total ficar proximo de 17% da receita bruta do
# caso de referencia (R$ 49.003,79): 12 pessoas, base R$ 29.200, encargos 68% ->
# R$ 49.056,00 (delta +R$ 52,21 contra os 17%).
_ENCARGOS_DEFAULT = 0.68
_QUADRO_PESSOAL_DEFAULT: tuple[tuple[str, int, float], ...] = (
    ("Gerente de unidade", 1, 4_700.00),
    ("Recepção / atendimento", 3, 2_000.00),
    ("Consultores de vendas", 2, 2_300.00),
    ("Professores / instrutores", 4, 2_600.00),
    ("Limpeza e manutenção", 2, 1_750.00),
)

_NOTA_ESTIMATIVA = (
    "ESTIMATIVA da equipe de expansão — este dado não existe nas bases do "
    "sistema. Substitua pelo valor real da unidade antes de defender o número."
)

# Os DOIS modos produzem uma folha FIXA no tempo (decisao de Felipe, 2026-07-24). O
# rotulo antigo do primeiro modo era "percentual da receita", o que passou a mentir
# quando a folha deixou de escalar com a receita do mes: o percentual dimensiona a
# folha pelo faturamento MADURO e o valor resultante vale desde o mes 1.
_MODOS_FOLHA = ("percentual do faturamento maduro", "quadro de pessoal")
_MODOS_RAMPA = ("demanda total", "apenas balcão")
_SIM_NAO = ("Sim", "Não")

# Fatores da grade de sensibilidade.
_SENS_FATORES_ALUNOS = (0.60, 0.70, 0.80, 0.90, 1.00, 1.10, 1.20, 1.30, 1.40)
_SENS_FATORES_ALUGUEL = (0.60, 0.80, 1.00, 1.20, 1.40)
# Linha/coluna do centro da grade (fatores 1,0 x 1,0) — aferida contra o motor.
_SENS_HDR_FATOR_ROW = 5
_SENS_HDR_ALUGUEL_ROW = 6
_SENS_GRID_ROW_INI = 7
_SENS_GRID_COL_INI = 3

# --- Layout FIXO das abas (as formulas cruzadas dependem destes numeros) -----
# Aba Folha
_FOLHA_MODO_ROW = 5
_FOLHA_HDR_ROW = 8
_FOLHA_CARGO_ROW_INI = 9
_FOLHA_TOTAL_ROW = _FOLHA_CARGO_ROW_INI + len(_QUADRO_PESSOAL_DEFAULT)
_FOLHA_AFER_ROW = _FOLHA_TOTAL_ROW + 2  # 3 linhas de afericao a partir daqui

# Aba Investimento (Item | Valor | Unidade | Fonte), dados a partir da linha 5
_INVEST_ROW_INI = 5
_INVEST_ORDEM = (
    "obra", "parcelas_obra", "obra_parcela", "equip", "prazo_equip", "juros_am",
    "taxa_franquia", "parcelas_franquia", "franquia_parcela",
    "capex_total", "investimento_total", "aporte_inicial", "equip_financiado",
    "pmt", "total_pago", "juros_totais", "saldo_m60",
)
_INVEST_ROW = {k: _INVEST_ROW_INI + i for i, k in enumerate(_INVEST_ORDEM)}

# Primeira coluna de mes nas abas de linha do tempo (A = rotulo da linha).
_MES_COL_INI = 2

# Cabecalho das abas de linha do tempo.
_TL_TITULO_ROW = 1
_TL_NOTA_ROW = 2
_TL_HDR_ROW = 4
_TL_DADOS_ROW_INI = 5

# --- Linhas da aba "DRE mensal" (key, rotulo, tipo, sublinha, negrito) -------
# tipo: int | txt | fator | alunos | brl | pct
_DRE_ORDEM: tuple[tuple[str, str, str, bool, bool], ...] = (
    ("mes", "Mês", "int", False, True),
    ("mes_contrato", "Mês de contrato (desde a entrega, M-4)", "int", True, False),
    ("t", "Mês de operação (t)", "int", True, False),
    ("fase", "Fase", "txt", True, False),
    ("f_ticket", "Fator de reajuste do ticket", "fator", True, False),
    ("f_aluguel", "Fator de reajuste do aluguel", "fator", True, False),
    ("f_custos", "Fator de reajuste dos custos", "fator", True, False),
    ("alunos_total", "Alunos totais", "alunos", False, True),
    ("alunos_balcao", "Alunos de balcão", "alunos", True, False),
    ("alunos_agregadores", "Alunos de agregadores", "alunos", True, False),
    ("rec_balcao", "Receita de mensalidades — balcão", "brl", True, False),
    ("rec_agregadores", "Receita de mensalidades — agregadores", "brl", True, False),
    ("rec_personal", "Receita de personal (fixa)", "brl", True, False),
    ("rec_anuidade", "Receita de anuidade", "brl", True, False),
    ("faturamento", "(=) FATURAMENTO BRUTO", "brl", False, True),
    ("deducoes", "(-) Deduções (devoluções)", "brl", False, False),
    ("receita_liquida", "(=) Receita líquida", "brl", False, True),
    ("impostos", "(-) Impostos sobre receita (PIS/COFINS/ISS)", "brl", False, False),
    ("receita_pos_impostos", "(=) Receita pós-impostos", "brl", False, True),
    ("cvar_total", "(-) Custo variável", "brl", False, False),
    ("cvar_royalties", "Royalties", "brl", True, False),
    ("cvar_marketing", "Marketing / FPP", "brl", True, False),
    ("cvar_manutencao", "Manutenção", "brl", True, False),
    ("cvar_cartoes", "Taxas de cartão", "brl", True, False),
    ("folha", "(-) Folha (FIXA desde o mês 1)", "brl", False, False),
    ("outros_total", "(-) Outros custos fixos", "brl", False, False),
    ("of_iptu", "IPTU", "brl", True, False),
    ("of_agua_luz", "Água e luz", "brl", True, False),
    ("of_telefone", "Telefone", "brl", True, False),
    ("of_limpeza", "Limpeza", "brl", True, False),
    ("of_tecnologia", "Tecnologia", "brl", True, False),
    ("of_assessorias", "Assessorias", "brl", True, False),
    ("aluguel", "(-) Aluguel", "brl", False, False),
    ("custo_pre_op", "(-) Custo pré-operacional", "brl", False, False),
    ("custos_op", "(=) Custos operacionais totais", "brl", False, True),
    ("ebitda", "(=) EBITDA", "brl", False, True),
    ("margem_ebitda", "Margem EBITDA", "pct", True, False),
    ("ir_base", "Base presumida (32% da receita bruta)", "brl", True, False),
    ("ir_irpj", "IRPJ 15% sobre a base", "brl", True, False),
    ("ir_adicional", "Adicional de IRPJ 10% (acima do limite mensal)", "brl", True, False),
    ("ir_csll_al", "CSLL 9% sobre a base", "brl", True, False),
    ("ir_total", "(-) IR/CSLL", "brl", False, False),
    ("juros", "(-) Despesa financeira (juros do financiamento)", "brl", False, False),
    # "do negócio" no lugar de "desalavancado" (vocabulário obrigatório, 4a rodada).
    ("resultado_desalav", "(=) Resultado DO NEGÓCIO após IR (EBITDA - IR/CSLL)",
     "brl", False, True),
    ("resultado_apos_juros", "(=) Resultado após IR e despesa financeira", "brl", False, True),
)
_DRE_ROW = {k: _TL_DADOS_ROW_INI + i for i, (k, *_r) in enumerate(_DRE_ORDEM)}

# --- Linhas da aba "Fluxo de caixa" -----------------------------------------
_FLX_ORDEM: tuple[tuple[str, str, str, bool, bool], ...] = (
    ("mes", "Mês", "int", False, True),
    ("mes_contrato", "Mês de contrato (desde a entrega, M-4)", "int", True, False),
    ("t", "Mês de operação (t)", "int", True, False),
    ("ebitda", "EBITDA", "brl", False, True),
    ("ir_csll", "(-) IR/CSLL", "brl", False, False),
    ("saldo_inicial", "Saldo devedor no início do mês", "brl", True, False),
    ("juros", "Juros do mês (saldo anterior x taxa)", "brl", True, False),
    ("pmt", "(-) PMT do financiamento (Price)", "brl", False, False),
    ("amortizacao", "Amortização do principal", "brl", True, False),
    ("saldo_final", "Saldo devedor no fim do mês", "brl", True, False),
    ("inv_obra", "Obra (parcela do mês)", "brl", True, False),
    ("inv_franquia", "Taxa de franquia (parcela do mês)", "brl", True, False),
    ("inv_equip_vista", "Equipamentos à vista (quando não financiados)", "brl", True, False),
    ("inv_capex_renov", "CAPEX de renovação", "brl", True, False),
    ("investimento", "(-) Investimento do mês", "brl", False, False),
    ("valor_residual", "(+) Valor residual", "brl", True, False),
    # O FCF da serie do motor E o fluxo DO SÓCIO: a PMT inteira já saiu e o
    # equipamento financiado nunca passou por este caixa (o banco pagou).
    ("fcf", "(=) FLUXO DO SÓCIO no mês (a PMT inteira já saiu)", "brl", False, True),
    ("fcf_acumulado", "(=) FLUXO DO SÓCIO ACUMULADO (caixa acumulado)", "brl", False, True),
    ("payback", "Marcador de payback", "txt", True, False),
    # --- As duas linhas novas da 4a rodada (2026-07-25) ----------------------
    # Ótica DO NEGÓCIO: sem financiamento nenhum, o CAPEX inteiro desembolsado.
    # Mede o ATIVO. É a linha que a TIR e o VPL "do negócio" descontam.
    ("neg_equip", "Equipamentos desembolsados na véspera (ótica do negócio)",
     "brl", True, False),
    ("fluxo_negocio", "(=) FLUXO DO NEGÓCIO (sem financiamento, CAPEX inteiro)",
     "brl", False, True),
    # Ótica DA DÍVIDA: entra a captação no mês da véspera, saem as PMT. Serve às
    # duas identidades da aba Aferição (e o VPL dela à taxa do negócio é a
    # ARBITRAGEM: a única forma pela qual a alavancagem cria valor sem escudo fiscal).
    ("div_captacao", "(+) Captação do financiamento (no mês da véspera)",
     "brl", True, False),
    ("fluxo_divida", "(=) FLUXO DA DÍVIDA (captação - PMT)", "brl", False, True),
)
_FLX_ROW = {k: _TL_DADOS_ROW_INI + i for i, (k, *_r) in enumerate(_FLX_ORDEM)}


# ---------------------------------------------------------------------------
# Helpers de escrita
# ---------------------------------------------------------------------------


def _titulo_bloco(ws: Worksheet, row: int, texto: str, n_cols: int) -> None:
    """Faixa turquesa de titulo de bloco dentro de uma aba."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=texto)
    c.fill = _fill(_TURQUESA_HEX)
    c.font = Font(name=_FONTE_PADRAO, bold=True, color=_BRANCO_HEX, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")


def _cabecalho_tabela(ws: Worksheet, row: int, labels: list[str], *, centro: bool = False) -> None:
    for col, txt in enumerate(labels, start=1):
        c = ws.cell(row=row, column=col, value=txt)
        c.fill = _fill(_CINZA_ESC)
        c.font = Font(name=_FONTE_PADRAO, bold=True, color=_BRANCO_HEX, size=9)
        c.alignment = Alignment(
            horizontal="center" if (centro or col > 1) else "left",
            vertical="center",
            wrap_text=True,
        )


def _nota(ws: Worksheet, row: int, texto: str, n_cols: int, *, alerta: bool = False) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=texto)
    c.fill = _fill(_AMARELO_CLR if alerta else _CINZA_CLR)
    c.font = Font(name=_FONTE_PADRAO, bold=alerta, italic=not alerta, color=_CINZA_ESC, size=9)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _estilo_valor(cell: Any, kind: str, fmt: str | None) -> None:
    """Amarelo+borda = editavel; cinza claro+italico = derivada; cinza = estrutural."""
    if kind == _EDIT:
        cell.fill = _fill(_AMARELO_CLR)
        cell.font = Font(name=_FONTE_PADRAO, bold=True, color=_CINZA_ESC, size=10)
    elif kind == _DERIV:
        cell.fill = _fill(_CINZA_CLR)
        cell.font = Font(name=_FONTE_PADRAO, italic=True, color=_CINZA_ESC, size=10)
    else:
        cell.fill = _fill(_CINZA_TRAVADO)
        cell.font = Font(name=_FONTE_PADRAO, italic=True, color="FF6B6B7B", size=10)
    cell.border = _BORDA_FINA
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if fmt:
        cell.number_format = fmt


def _linha_label(ws: Worksheet, row: int, texto: str, *, negrito: bool = False,
                 sub: bool = False, faixa: str | None = None) -> None:
    c = ws.cell(row=row, column=1, value=("    " + texto) if sub else texto)
    c.font = Font(
        name=_FONTE_PADRAO,
        bold=negrito,
        italic=sub,
        color=_CINZA_ESC if not sub else "FF55555F",
        size=9,
    )
    c.alignment = Alignment(horizontal="left", vertical="center")
    if faixa:
        c.fill = _fill(faixa)


def _sanitizar(valor: float) -> float:
    """Nenhum inf/NaN pode entrar na planilha (guardrail do projeto)."""
    if valor is None or not isinstance(valor, (int, float)):
        return 0.0
    v = float(valor)
    if math.isnan(v) or math.isinf(v):
        return 0.0
    return v


def _finito(valor: Any) -> bool:
    return isinstance(valor, (int, float)) and math.isfinite(float(valor))


def _num_br(valor: float, decimais: int = 2) -> str:
    """Numero com separadores pt-BR para TEXTO de usuario (1200 -> '1.200,00').

    As CELULAS numericas nao precisam disto: o Excel aplica o separador do locale
    a partir do `number_format`. Isto e so para numero embutido em frase — titulo
    de aba, nota de celula — onde o `f"{v:,.2f}"` do Python entregaria o separador
    ingles ("1,200.00") no meio de uma sentenca em portugues.
    """
    return f"{valor:,.{decimais}f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _aplicar(template: str, refs: dict[str, str]) -> str:
    """Resolve `{chave}` para a referencia absoluta da celula correspondente.

    Deixa as formulas legiveis no codigo (`={ticket_cheio}*{tag_fator}`) e
    absolutas no arquivo (`=Premissas!$B$12*Premissas!$B$11`).
    """
    out = template

    def _range(m: re.Match[str]) -> str:
        # Excel exige `Aba!$B$5:$B$9` — repetir a aba nos dois lados (`Aba!$B$5:Aba!$B$9`)
        # e sintaxe invalida e a formula quebra na abertura.
        ini, fim = refs[m.group(1)], refs[m.group(2)]
        return f"{ini}:{fim.split('!', 1)[-1]}"

    out = re.sub(r"\{RANGE:([a-z_0-9]+):([a-z_0-9]+)\}", _range, out)
    for k, ref in refs.items():
        out = out.replace("{" + k + "}", ref)
    if "{" in out:
        faltando = out[out.index("{") : out.index("{") + 40]
        raise KeyError(f"referencia nao resolvida na formula: {faltando}")
    return out


def _bool_sim_nao(v: bool) -> str:
    return "Sim" if v else "Não"


def _e(
    key: str,
    label: str,
    *,
    valor: Any = None,
    formula: str | None = None,
    unidade: str = "",
    fonte: str = "",
    quem: str = _QUEM_OPERADOR,
    kind: str = _EDIT,
    fmt: str | None = None,
    nota: str | None = None,
) -> dict[str, Any]:
    """Uma linha da aba Premissas."""
    return {
        "key": key, "label": label, "valor": valor, "formula": formula,
        "unidade": unidade, "fonte": fonte, "quem": quem, "kind": kind,
        "fmt": fmt, "nota": nota,
    }


def _bloco_demanda(demanda_total: float, p: Premissas, modo_rampa: str) -> list[dict[str, Any]]:
    return [
        _e("demanda", "Demanda total (alunos na maturidade)", valor=float(demanda_total),
           unidade="alunos", fonte="Entrada do operador (DEC-009: nunca de lat/lng)",
           fmt=_FMT_ALUNOS),
        _e("alunos_inicial", "Alunos no mês 1 (início da rampa)", valor=float(p.alunos_inicial),
           unidade="alunos", fonte="config.py SIM_ALUNOS_INICIAL", fmt=_FMT_ALUNOS),
        _e("alunos_ini_ef", "Alunos iniciais efetivos (piso da rampa)",
           formula='=MIN({alunos_inicial},IF({modo_rampa}="apenas balcão",'
                   "{demanda}*{share},{demanda}))",
           unidade="alunos", fonte="Derivado: a rampa nunca decresce",
           quem=_QUEM_DERIVADO, kind=_DERIV, fmt=_FMT_ALUNOS),
        _e("maturacao", "Maturação da rampa", valor=int(p.maturacao_meses), unidade="meses",
           fonte="config.py SIM_MATURACAO_MESES", fmt=_FMT_INT),
        _e("modo_rampa", "Modo da rampa", valor=modo_rampa, unidade="lista",
           fonte="simulador.Premissas.rampa_apenas_balcao"),
        _e("share", "Share de balcão na demanda", valor=float(p.share_balcao), unidade="%",
           fonte="config.py SIM_SHARE_BALCAO", fmt=_FMT_PCT2),
        _e("ticket_cheio", "Ticket cheio (balcão)", valor=float(p.ticket_cheio),
           unidade="R$/aluno/mês", fonte="Entrada do operador", fmt=_FMT_CONTABIL),
        _e("tag_fator", "Ticket do agregador (fração do cheio)",
           valor=float(p.ticket_agregador_fator), unidade="%",
           fonte="config.py SIM_TICKET_AGREGADOR_FATOR", fmt=_FMT_PCT2),
        _e("ticket_agregador", "Ticket do agregador", formula="={ticket_cheio}*{tag_fator}",
           unidade="R$/aluno/mês", fonte="Derivado (acoplado ao ticket cheio)",
           quem=_QUEM_DERIVADO, kind=_DERIV, fmt=_FMT_CONTABIL),
        _e("churn", "Churn mensal", valor=float(p.churn), unidade="%",
           fonte="config.py SIM_CHURN", fmt=_FMT_PCT2, quem=_QUEM_CONTROLADORIA),
        _e("inadimplencia", "Inadimplência", valor=float(p.inadimplencia), unidade="%",
           fonte="config.py SIM_INADIMPLENCIA", fmt=_FMT_PCT2, quem=_QUEM_CONTROLADORIA),
        _e("personal_mes", "Receita fixa de personal", valor=float(p.personal_mes),
           unidade="R$/mês", fonte="config.py SIM_PERSONAL_MES_RECEITA", fmt=_FMT_CONTABIL,
           nota="Receita fixa: NÃO escala com alunos e NÃO sofre reajuste anual, "
                "igual ao motor."),
        _e("ticket_blended", "Ticket blended por aluno total",
           formula="={share}*(1-{churn})*{ticket_cheio}*(1-{inadimplencia})"
                   "+(1-{share})*{ticket_agregador}*(1-{inadimplencia})",
           unidade="R$/aluno/mês", fonte="Derivado (líquido de churn e inadimplência)",
           quem=_QUEM_DERIVADO, kind=_DERIV, fmt=_FMT_CONTABIL),
        _e("receita_por_aluno", "Receita por aluno total (com anuidade)",
           formula="={ticket_blended}+{anuidade_por_aluno}"
                   '*IF({anuidade_apenas_balcao}="Sim",{share},1)',
           unidade="R$/aluno/mês", fonte="Derivado (base do break-even)",
           quem=_QUEM_DERIVADO, kind=_DERIV, fmt=_FMT_CONTABIL),
    ]


def _bloco_impostos(p: Premissas) -> list[dict[str, Any]]:
    ctr = _QUEM_CONTROLADORIA
    return [
        _e("devolucoes", "Devoluções / deduções da receita bruta", valor=float(p.devolucoes_pct),
           unidade="% da bruta", fonte="config.py SIM_DEVOLUCOES_PCT (DRE F30)",
           fmt=_FMT_PCT3, quem=ctr),
        _e("pis", "PIS", valor=float(p.pis), unidade="% da líquida",
           fonte="config.py SIM_PIS (Tributos E38)", fmt=_FMT_PCT3, quem=ctr),
        _e("cofins", "COFINS", valor=float(p.cofins), unidade="% da líquida",
           fonte="config.py SIM_COFINS (Tributos E40)", fmt=_FMT_PCT3, quem=ctr),
        _e("iss", "ISS", valor=float(p.iss), unidade="% da líquida",
           fonte="config.py SIM_ISS (Tributos E42)", fmt=_FMT_PCT3, quem=ctr),
        _e("impostos_total", "Impostos sobre receita (total)", formula="={pis}+{cofins}+{iss}",
           unidade="% da líquida", fonte="Derivado", quem=_QUEM_DERIVADO, kind=_DERIV,
           fmt=_FMT_PCT3),
        _e("ir_modo", "Regime de IR/CSLL", valor=str(p.ir_modo), unidade="lista",
           fonte="simulador.IR_MODO_FAIXA (Presumido com a faixa do adicional)", quem=ctr),
        _e("base_presumida", "Base presumida (Lucro Presumido, serviços)",
           valor=float(p.base_presumida_pct), unidade="% da bruta",
           fonte="config.py SIM_BASE_PRESUMIDA_PCT", fmt=_FMT_PCT2, quem=ctr),
        _e("irpj_al", "IRPJ", valor=float(p.irpj_aliquota), unidade="% da base",
           fonte="config.py SIM_IRPJ_ALIQUOTA", fmt=_FMT_PCT2, quem=ctr),
        _e("irpj_adic_al", "Adicional de IRPJ", valor=float(p.irpj_adicional_aliquota),
           unidade="% do excedente", fonte="config.py SIM_IRPJ_ADICIONAL_ALIQUOTA",
           fmt=_FMT_PCT2, quem=ctr),
        _e("irpj_adic_limite", "Limite mensal do adicional (pró-rata de R$ 60 mil/trimestre)",
           valor=float(p.irpj_adicional_limite_mes), unidade="R$/mês",
           fonte="config.py SIM_IRPJ_ADICIONAL_LIMITE_MES", fmt=_FMT_CONTABIL, quem=ctr),
        _e("csll_al", "CSLL", valor=float(p.csll_aliquota), unidade="% da base",
           fonte="config.py SIM_CSLL_ALIQUOTA", fmt=_FMT_PCT2, quem=ctr),
        _e("ir_efetivo", "IR efetivo (usado só no regime legado)", valor=float(p.ir_efetivo),
           unidade="% da líquida", fonte="config.py SIM_IR_EFETIVO", fmt=_FMT_PCT2, quem=ctr),
        _e("csll_efetivo", "CSLL efetivo (usado só no regime legado)",
           valor=float(p.csll_efetivo), unidade="% da líquida",
           fonte="config.py SIM_CSLL_EFETIVO", fmt=_FMT_PCT2, quem=ctr),
    ]


def _bloco_custo_variavel(p: Premissas) -> list[dict[str, Any]]:
    return [
        _e("royalties", "Royalties", valor=float(p.royalties_pct), unidade="% da líquida",
           fonte="config.py SIM_ROYALTIES_PCT (Simulador N11)", fmt=_FMT_PCT2,
           quem=_QUEM_FRANQUEADORA),
        _e("marketing", "Marketing / FPP", valor=float(p.marketing_pct), unidade="% da líquida",
           fonte="config.py SIM_MARKETING_PCT (DRE F63)", fmt=_FMT_PCT2,
           quem=_QUEM_FRANQUEADORA),
        _e("manutencao", "Manutenção", valor=float(p.manutencao_pct), unidade="% da líquida",
           fonte="config.py SIM_MANUTENCAO_PCT (DRE F67)", fmt=_FMT_PCT2,
           quem=_QUEM_CONTROLADORIA),
        _e("cartoes", "Taxas de cartão", valor=float(p.cartoes_pct), unidade="% da líquida",
           fonte="config.py SIM_CARTOES_PCT (DRE F79)", fmt=_FMT_PCT2,
           quem=_QUEM_CONTROLADORIA),
        _e("cvar_total", "Custo variável total",
           formula="={royalties}+{marketing}+{manutencao}+{cartoes}", unidade="% da líquida",
           fonte="Derivado", quem=_QUEM_DERIVADO, kind=_DERIV, fmt=_FMT_PCT2),
    ]


def _bloco_folha(p: Premissas) -> list[dict[str, Any]]:
    """Folha FIXA no tempo, dimensionada pelo faturamento MADURO.

    As duas celulas novas (`fat_maduro` e `folha_fixa`) existem para a DRE referenciar
    UMA celula em todos os 64 meses, em vez de recalcular a folha coluna por coluna a
    partir do faturamento daquele mes — que era exatamente o defeito reportado ("a
    folha esta escalando junto com a unidade").
    """
    return [
        _e("folha_pct", "Folha como % do faturamento MADURO", valor=float(p.folha_pct),
           unidade="% do maduro", fonte="config.py SIM_FOLHA_PCT", fmt=_FMT_PCT2,
           quem=_QUEM_CONTROLADORIA,
           nota="O percentual DIMENSIONA a folha pelo faturamento de regime pleno; o R$ "
                "resultante é FIXO desde o mês 1 (decisão de Felipe, 2026-07-24). "
                "Conflito aberto (BLK-VIAB-11): 6 DREs gerenciais reais apuraram 0,25-0,26. "
                "O 0,17 mantém o nível do status quo; a calibração segue pendente de gate."),
        _e("modo_folha", "Modo da folha (interruptor)", formula=f"={ABA_FOLHA}!$B${_FOLHA_MODO_ROW}",
           unidade="lista", fonte=f"Aba {ABA_FOLHA}, célula B{_FOLHA_MODO_ROW} (validação de lista)",
           quem=_QUEM_DERIVADO, kind=_DERIV),
        _e("folha_quadro_total", "Total do quadro de pessoal (aba Folha)",
           formula=f"={ABA_FOLHA}!$E${_FOLHA_TOTAL_ROW}", unidade="R$/mês",
           fonte=f"Aba {ABA_FOLHA} (soma do quadro editável)", quem=_QUEM_DERIVADO,
           kind=_DERIV, fmt=_FMT_CONTABIL),
        _e("fat_maduro", "Faturamento MADURO (base de dimensionamento da folha)",
           formula="={demanda}*{receita_por_aluno}+{personal_mes}", unidade="R$/mês",
           fonte="Derivado: regime pleno a preços do ano 1 (casa cheia, anuidade em cobrança)",
           quem=_QUEM_DERIVADO, kind=_DERIV, fmt=_FMT_CONTABIL,
           nota="A PREÇOS DO ANO 1, de propósito: o reajuste anual de custos entra uma única "
                "vez, na linha da DRE, e não duas vezes (aqui e lá)."),
        _e("folha_fixa", "Folha mensal FIXA (vale desde o mês 1)",
           formula='=IF({modo_folha}="quadro de pessoal",{folha_quadro_total},'
                   "{folha_pct}*{fat_maduro})",
           unidade="R$/mês",
           fonte="Derivado: os DOIS modos do interruptor são fixos no tempo",
           quem=_QUEM_DERIVADO, kind=_DERIV, fmt=_FMT_CONTABIL,
           nota="A equipe existe ANTES dos alunos: a folha é paga integralmente desde o mês 1 "
                "e a DRE lê esta única célula nos 64 meses, só multiplicada pelo reajuste "
                "anual de custos. Antes ela era % do faturamento DO MÊS e encolhia com a "
                "rampa, o que subestimava o EBITDA do mês 1 e o break-even."),
    ]


def _rateio_outros_fixos(total_alvo: float) -> list[tuple[str, str, float, bool]]:
    """Ajusta as componentes para somarem EXATAMENTE `total_alvo`.

    Com o default do motor (SIM_OUTROS_FIXOS_MES = 38.150) a escala e 1,0 e os valores
    saem REDONDOS, sem rateio — as seis componentes ja fecham no total. O rateio abaixo
    e rede de seguranca: se alguem mudar `SIM_OUTROS_FIXOS_MES` sem atualizar esta
    decomposicao, a planilha continua reproduzindo o motor ao centavo (com a nota de
    celula avisando que houve rateio) em vez de divergir em silencio.

    O residuo de arredondamento vai na ULTIMA linha, para o total fechar ao centavo.
    """
    if _OUTROS_FIXOS_SOMA_DECOMP <= 0:
        return [(k, lbl, 0.0, False) for k, lbl, _v in _OUTROS_FIXOS_DECOMP]
    escala = float(total_alvo) / _OUTROS_FIXOS_SOMA_DECOMP
    rateado = abs(escala - 1.0) > 1e-12
    itens = [
        (k, lbl, round(v * escala, 2), rateado)
        for k, lbl, v in _OUTROS_FIXOS_DECOMP[:-1]
    ]
    k_ult, lbl_ult, _v_ult = _OUTROS_FIXOS_DECOMP[-1]
    residuo = round(float(total_alvo) - sum(x[2] for x in itens), 2)
    itens.append((k_ult, lbl_ult, residuo, rateado))
    return itens


def _bloco_custos_fixos(p: Premissas) -> list[dict[str, Any]]:
    itens = _rateio_outros_fixos(p.outros_fixos_mes)
    nota_of = (
        "RATEADO. A decomposição desta planilha soma R$ "
        f"{_num_br(_OUTROS_FIXOS_SOMA_DECOMP)}, mas o total que o motor usa é R$ "
        f"{_num_br(p.outros_fixos_mes)} (SIM_OUTROS_FIXOS_MES) — alguém mudou a constante "
        "sem atualizar a decomposição. Para a planilha continuar reproduzindo o motor ao "
        "centavo, as componentes foram rateadas proporcionalmente (fator "
        f"{_num_br(p.outros_fixos_mes / _OUTROS_FIXOS_SOMA_DECOMP, 6)}), o que deixa os "
        "valores quebrados. Substitua pelos valores reais da unidade."
    )
    of = [
        _e(k, lbl, valor=valor, unidade="R$/mês",
           fonte="config.py SIM_OUTROS_FIXOS_MES (decomposição do comentário)",
           fmt=_FMT_CONTABIL, nota=(nota_of if rateado else None))
        for k, lbl, valor, rateado in itens
    ]
    return [
        *of,
        _e("outros_total", "Outros fixos (total)", formula=f"=SUM({{RANGE:{_OF_KEY_INI}:{_OF_KEY_FIM}}})",
           unidade="R$/mês",
           fonte=f"Derivado (soma das {len(_OUTROS_FIXOS_DECOMP)} linhas acima)",
           quem=_QUEM_DERIVADO, kind=_DERIV, fmt=_FMT_CONTABIL),
        _e("aluguel", "Aluguel", valor=float(p.aluguel_mes), unidade="R$/mês",
           fonte="Entrada do operador (imóvel negociado)", fmt=_FMT_CONTABIL),
        _e("carencia", "Carência de aluguel (contada de M-4)",
           valor=int(p.carencia_aluguel_meses), unidade="meses",
           fonte="config.py SIM_CARENCIA_ALUGUEL_MESES", fmt=_FMT_INT),
        _e("custo_pre_op", "Custo pré-operacional (meses de obra)",
           valor=float(p.custo_pre_operacional_mes), unidade="R$/mês",
           fonte="config.py SIM_CUSTO_PRE_OPERACIONAL_MES", fmt=_FMT_CONTABIL,
           nota="Default ZERO de propósito: torna explícito que hoje o modelo assume "
                "ausência total de custo pré-operacional."),
        _e("custo_fixo_total", "Custo fixo total, sem aluguel (outros fixos + folha)",
           formula="={outros_total}+{folha_fixa}",
           unidade="R$/mês", fonte="Derivado (base do break-even)", quem=_QUEM_DERIVADO,
           kind=_DERIV, fmt=_FMT_CONTABIL,
           nota="A folha entra AQUI, no custo fixo, porque deixou de ser percentual da "
                "receita do mês (decisão de Felipe, 2026-07-24). É por isso que o "
                "break-even subiu: a folha não é mais diluída pelo volume."),
        _e("k_ebitda", "Fator receita -> EBITDA (k)",
           formula="=(1-{devolucoes})*(1-{impostos_total}-{cvar_total})",
           unidade="R$ por R$ de bruta",
           fonte="Derivado: quanto de cada R$ 1 de bruta sobra antes do custo fixo",
           quem=_QUEM_DERIVADO, kind=_DERIV, fmt="0.000000",
           nota="A FOLHA NÃO ENTRA MAIS AQUI (era subtraída como folha_pct). Ela virou "
                "custo FIXO, dimensionado pelo faturamento maduro — ver 'Folha mensal FIXA'. "
                "Por isso k passou de 0,628985 para 0,798985."),
    ]


def _bloco_investimento(
    *,
    obra: float,
    parcelas_obra: int,
    equipamentos: float,
    prazo_equipamentos: int,
    juros_equipamentos_am: float,
    taxa_franquia: float,
    parcelas_franquia: int,
) -> list[dict[str, Any]]:
    return [
        _e("obra", "Obra (aporte do franqueado, parcelada sem juros)",
           valor=float(obra), unidade="R$",
           fonte="Entrada do operador", fmt=_FMT_CONTABIL),
        _e("parcelas_obra", "Parcelas da obra", valor=int(parcelas_obra), unidade="meses",
           fonte="config.py SIM_PARCELAS_OBRA_DEFAULT", fmt=_FMT_INT),
        _e("equip", "Equipamentos (financiados)", valor=float(equipamentos), unidade="R$",
           fonte="Entrada do operador", fmt=_FMT_CONTABIL),
        _e("prazo_equip", "Prazo do financiamento", valor=int(prazo_equipamentos),
           unidade="meses", fonte="Entrada do operador", fmt=_FMT_INT),
        _e("juros_am", "Juros do financiamento", valor=float(juros_equipamentos_am),
           unidade="% a.m.", fonte="Entrada do operador (proposta do banco)", fmt=_FMT_PCT3),
        _e("taxa_franquia", "Taxa de franquia (parcelada sem juros)", valor=float(taxa_franquia),
           unidade="R$", fonte="config.py SIM_TAXA_FRANQUIA", fmt=_FMT_CONTABIL,
           quem=_QUEM_FRANQUEADORA,
           nota="Divergência de fonte registrada: a planilha original (célula R10) e o doc "
                "do modelo dizem R$ 140.000. R$ 160.000 é o valor em produção e o que o "
                "comitê já viu (decisão de Felipe, 2026-07-24)."),
        _e("parcelas_franquia", "Parcelas da taxa de franquia", valor=int(parcelas_franquia),
           unidade="meses", fonte="config.py SIM_PARCELAS_FRANQUIA_DEFAULT", fmt=_FMT_INT,
           quem=_QUEM_FRANQUEADORA,
           nota="Parcelada SEM JUROS, junto da obra: as parcelas caem nos meses de CONTRATO "
                "1..N (M-4..M-1 com N=4), não a taxa inteira no M-4 (decisão de Felipe, "
                "2026-07-24). É só TIMING de caixa — não muda EBITDA nem break-even."),
        _e("franquia_parcela", "Valor da parcela da franquia",
           formula="=IF({taxa_franquia}>0,{taxa_franquia}/MAX({parcelas_franquia},1),0)",
           unidade="R$/mês", fonte="Derivado (sem juros: divisão simples)",
           quem=_QUEM_DERIVADO, kind=_DERIV, fmt=_FMT_CONTABIL),
        _e("equip_financiado", "Principal financiado",
           formula="=IF(AND({equip}>0,{prazo_equip}>0),{equip},0)", unidade="R$",
           fonte="Derivado (sem prazo, o equipamento é pago à vista no M-1)",
           quem=_QUEM_DERIVADO, kind=_DERIV, fmt=_FMT_CONTABIL),
        _e("capex_total", "CAPEX total (obra + equipamentos)", formula="={obra}+{equip}",
           unidade="R$", fonte="Derivado", quem=_QUEM_DERIVADO, kind=_DERIV, fmt=_FMT_CONTABIL),
        _e("investimento_total", "Investimento total (CAPEX + franquia)",
           formula="={capex_total}+{taxa_franquia}", unidade="R$", fonte="Derivado",
           quem=_QUEM_DERIVADO, kind=_DERIV, fmt=_FMT_CONTABIL),
        # VOCABULARIO (4a rodada): "aporte inicial" no lugar de "equity aportado". Os
        # R$ 760 mil daqui NAO sao o cheque que o investidor precisa ter — esse e o
        # "cheque total" da aba Resumo (o pior ponto do caixa acumulado). Este segue
        # sendo o DENOMINADOR do retorno do socio e a base da alavancagem.
        _e("aporte_inicial", "Aporte inicial (obra + franquia)",
           formula="={obra}+{taxa_franquia}",
           unidade="R$", fonte="Derivado (denominador do retorno do sócio)",
           quem=_QUEM_DERIVADO, kind=_DERIV, fmt=_FMT_CONTABIL,
           nota="NÃO confundir com o CHEQUE TOTAL da aba Resumo: aquele é o pior ponto do "
                "caixa acumulado (o dinheiro que precisa estar disponível) e é bem maior. "
                "Este é o aporte contratado de obra + taxa de franquia."),
        _e("pmt", "PMT do financiamento (Price)",
           formula=f"={ABA_INVESTIMENTO}!$B${_INVEST_ROW['pmt']}", unidade="R$/mês",
           fonte=f"Aba {ABA_INVESTIMENTO} (=PGTO do Excel)", quem=_QUEM_DERIVADO,
           kind=_DERIV, fmt=_FMT_CONTABIL),
    ]


def _bloco_anuidade(p: Premissas) -> list[dict[str, Any]]:
    if p.anuidade_elegivel_pct is not None:
        elegivel = _e("anuidade_elegivel", "Fração de alunos que chega ao mês de cobrança",
                      valor=float(p.anuidade_elegivel_pct), unidade="%",
                      fonte="config.py SIM_ANUIDADE_ELEGIVEL_PCT (override explícito)",
                      fmt=_FMT_PCT2, quem=_QUEM_CONTROLADORIA)
    else:
        elegivel = _e("anuidade_elegivel", "Fração de alunos que chega ao mês de cobrança",
                      formula="=(1-{churn})^{anuidade_mes_inicio}", unidade="%",
                      fonte="Derivado do churn ((1-churn)^mês de início)",
                      quem=_QUEM_DERIVADO, kind=_DERIV, fmt=_FMT_PCT2)
    return [
        _e("anuidade_valor", "Valor da anuidade", valor=float(p.anuidade_valor),
           unidade="R$/aluno/ano", fonte="config.py SIM_ANUIDADE_VALOR (Simulador J10)",
           fmt=_FMT_CONTABIL, quem=_QUEM_FRANQUEADORA),
        _e("anuidade_mes_inicio", "Mês de operação em que a cobrança começa",
           valor=int(p.anuidade_mes_inicio), unidade="mês",
           fonte="config.py SIM_ANUIDADE_MES_INICIO (Simulador J12)", fmt=_FMT_INT),
        _e("anuidade_apenas_balcao", "Somente o balcão paga?",
           valor=_bool_sim_nao(bool(p.anuidade_apenas_balcao)), unidade="Sim/Não",
           fonte="config.py SIM_ANUIDADE_APENAS_BALCAO"),
        _e("anuidade_pro_rata", "Reconhecimento pró-rata mensal?",
           valor=_bool_sim_nao(bool(p.anuidade_pro_rata)), unidade="Sim/Não",
           fonte="config.py SIM_ANUIDADE_PRO_RATA"),
        elegivel,
        _e("anuidade_por_aluno", "Anuidade por aluno de balcão / mês",
           formula='=IF({anuidade_valor}<=0,0,IF({anuidade_pro_rata}="Sim",'
                   "{anuidade_valor}*{anuidade_elegivel}/12,"
                   "{anuidade_valor}*{anuidade_elegivel}))",
           unidade="R$/aluno/mês", fonte="Derivado", quem=_QUEM_DERIVADO, kind=_DERIV,
           fmt=_FMT_CONTABIL),
    ]


def _bloco_reajuste(p: Premissas) -> list[dict[str, Any]]:
    return [
        _e("reajuste_ticket", "Reajuste do ticket", valor=float(p.reajuste_ticket_aa),
           unidade="% a.a.", fonte="config.py SIM_REAJUSTE_TICKET_AA", fmt=_FMT_PCT2),
        _e("reajuste_aluguel", "Reajuste do aluguel", valor=float(p.reajuste_aluguel_aa),
           unidade="% a.a.", fonte="config.py SIM_REAJUSTE_ALUGUEL_AA", fmt=_FMT_PCT2),
        _e("reajuste_custos", "Reajuste dos custos fixos", valor=float(p.reajuste_custos_aa),
           unidade="% a.a.", fonte="config.py SIM_REAJUSTE_CUSTOS_AA", fmt=_FMT_PCT2),
    ]


def _bloco_taxas(p: Premissas) -> list[dict[str, Any]]:
    """As taxas mínimas de retorno: UMA editável, o resto DERIVADO por fórmula.

    Substitui o bloco "Desconto (VPL / TIR)", que tinha uma unica "taxa de desconto"
    de 12% a.a. aplicada a um fluxo DE SOCIO — incoerente, porque o socio e subordinado
    ao banco e nao pode exigir menos que o credor (custo da divida de 23,87% a.a. no
    caso de referencia).

    Aqui a taxa minima do NEGOCIO (Ku na literatura) e a unica entrada; a do SOCIO (Ke)
    e DERIVADA por `Ke = Ku + (Ku - Kd) * D/E`, exatamente como `simulador.simular()`.
    Isso torna a incoerencia impossivel por construcao: nao existe celula onde alguem
    possa digitar uma taxa de socio abaixo do custo da divida. E como o Lucro Presumido
    nao tem escudo fiscal da divida, o WACC E a taxa minima do negocio — nao ha media
    ponderada a fazer, e por isso ela nao aparece aqui.
    """
    return [
        _e("taxa_negocio_aa", "Taxa mínima do negócio (a.a.)",
           valor=float(p.taxa_minima_negocio_aa), unidade="% a.a.",
           fonte="config.py SIM_TAXA_MINIMA_NEGOCIO_AA", fmt=_FMT_PCT2,
           quem=_QUEM_CONTROLADORIA,
           nota="ÚNICA taxa editável do modelo (decisão de Felipe, 2026-07-25). 25% a.a. "
                "NOMINAL: piso implícito da própria decisão de financiar (o custo da dívida "
                "de 1,8% a.m. = 23,87% a.a.) mais build-up sobre a Selic de 14,25%. Sem "
                "escudo fiscal no Lucro Presumido, esta é também a taxa do ativo (o WACC)."),
        _e("taxa_negocio_am", "Taxa mínima do negócio (a.m. equivalente)",
           formula="=(1+{taxa_negocio_aa})^(1/12)-1", unidade="% a.m.",
           fonte="Derivado (equivalência composta)", quem=_QUEM_DERIVADO, kind=_DERIV,
           fmt=_FMT_PCT3),
        _e("custo_divida_am", "Custo da dívida (a.m.)",
           formula="=IF(AND({equip}>0,{prazo_equip}>0),{juros_am},0)", unidade="% a.m.",
           fonte="Derivado dos juros do financiamento (zero sem financiamento)",
           quem=_QUEM_DERIVADO, kind=_DERIV, fmt=_FMT_PCT3),
        _e("custo_divida_aa", "Custo da dívida (a.a.)",
           formula="=(1+{custo_divida_am})^12-1", unidade="% a.a.",
           fonte="Derivado: anualização composta do juros a.m.", quem=_QUEM_DERIVADO,
           kind=_DERIV, fmt=_FMT_PCT2),
        _e("alavancagem", "Alavancagem (dívida / aporte inicial)",
           formula="=IF(AND({equip_financiado}>0,{aporte_inicial}>0),"
                   "{equip_financiado}/{aporte_inicial},0)",
           unidade="x", fonte="Derivado: principal financiado sobre obra + franquia",
           quem=_QUEM_DERIVADO, kind=_DERIV, fmt="0.0000"),
        _e("taxa_socio_aa", "Taxa mínima do sócio (a.a.)",
           formula="={taxa_negocio_aa}+({taxa_negocio_aa}-{custo_divida_aa})*{alavancagem}",
           unidade="% a.a.",
           fonte="Derivado: taxa do negócio + prêmio da alavancagem (NÃO é editável)",
           quem=_QUEM_DERIVADO, kind=_DERIV, fmt=_FMT_PCT2,
           nota="FÓRMULA, nunca um número cravado: o sócio é subordinado ao banco, então a "
                "taxa dele é a do negócio MAIS o prêmio de correr o risco depois do credor. "
                "Se a dívida custasse mais que o negócio, este prêmio ficaria NEGATIVO — e é "
                "isso que a linha de alerta logo abaixo denuncia."),
        _e("taxa_socio_am", "Taxa mínima do sócio (a.m. equivalente)",
           formula="=(1+{taxa_socio_aa})^(1/12)-1", unidade="% a.m.",
           fonte="Derivado (equivalência composta)", quem=_QUEM_DERIVADO, kind=_DERIV,
           fmt=_FMT_PCT3),
        _e("alerta_divida", "A dívida custa mais que a taxa mínima do negócio?",
           formula='=IF(AND({equip_financiado}>0,{custo_divida_aa}>{taxa_negocio_aa}),'
                   '"Sim - a alavancagem destroi valor","Nao")',
           unidade="Sim/Não",
           fonte="Derivado: guarda da 4a rodada", quem=_QUEM_DERIVADO, kind=_DERIV,
           nota="Sem escudo fiscal no Lucro Presumido, a alavancagem só cria valor por "
                "ARBITRAGEM: tomar dinheiro a uma taxa menor do que o ativo rende. Se a "
                "dívida custar mais que a taxa mínima do negócio, ela DESTRÓI valor e o "
                "prêmio da taxa do sócio fica negativo."),
    ]


def _bloco_teto() -> list[dict[str, Any]]:
    return [
        _e("teto_ideal_pct", "Faixa ideal", valor=float(SIM_ALUGUEL_TETO_IDEAL),
           unidade="% da bruta", fonte="config.py SIM_ALUGUEL_TETO_IDEAL", fmt=_FMT_PCT2,
           quem=_QUEM_FRANQUEADORA),
        _e("teto_pct", "TETO (canônico)", valor=float(SIM_ALUGUEL_TETO_TETO),
           unidade="% da bruta", fonte="config.py SIM_ALUGUEL_TETO_TETO", fmt=_FMT_PCT2,
           quem=_QUEM_FRANQUEADORA,
           nota="Este é o número canônico do card grande (decisão de Felipe, 2026-07-24). "
                "A faixa de exceção (30%) é exceção, não referência."),
        _e("teto_excecao_pct", "Exceção", valor=float(SIM_ALUGUEL_TETO_EXCECAO),
           unidade="% da bruta", fonte="config.py SIM_ALUGUEL_TETO_EXCECAO", fmt=_FMT_PCT2,
           quem=_QUEM_FRANQUEADORA),
    ]


def _bloco_horizonte(p: Premissas) -> list[dict[str, Any]]:
    nota_estrutural = (
        "ESTRUTURAL: define quantas colunas a planilha tem. Mudar aqui NÃO cria colunas "
        "novas — gere o arquivo de novo pelo sistema."
    )
    return [
        _e("meses_pre", "Meses de pré-abertura (obra)", valor=int(p.meses_pre_abertura),
           unidade="meses", fonte="config.py SIM_MESES_PRE_ABERTURA", kind=_FIXO,
           quem=_QUEM_ESTRUTURAL, fmt=_FMT_INT, nota=nota_estrutural),
        _e("horizonte", "Horizonte de operação", valor=int(p.horizonte_meses), unidade="meses",
           fonte="config.py SIM_HORIZONTE_MESES", kind=_FIXO, quem=_QUEM_ESTRUTURAL,
           fmt=_FMT_INT, nota=nota_estrutural),
        _e("valor_residual", "Valor residual no M60", valor=float(p.valor_residual_mes_60),
           unidade="R$", fonte="config.py SIM_VALOR_RESIDUAL_MES_60", fmt=_FMT_CONTABIL,
           nota="Default ZERO de propósito: o corte em 60 meses hoje ignora o valor residual."),
        _e("capex_renov", "CAPEX de renovação no M60", valor=float(p.capex_renovacao),
           unidade="R$", fonte="config.py SIM_CAPEX_RENOVACAO", fmt=_FMT_CONTABIL,
           nota="Default ZERO de propósito: o corte em 60 meses hoje ignora a renovação."),
        _e("mes_steady", "Mês de referência do regime pleno",
           formula="=MIN({horizonte},MAX(1,{maturacao},"
                   "IF({anuidade_valor}>0,{anuidade_mes_inicio},0)))",
           unidade="mês", fonte="Derivado: alunos maduros E anuidade já em cobrança",
           quem=_QUEM_DERIVADO, kind=_DERIV, fmt=_FMT_INT),
        _e("margem_viavel_min", "Margem EBITDA mínima para viável",
           valor=float(SIM_MARGEM_VIAVEL_MIN), unidade="%",
           fonte="config.py SIM_MARGEM_VIAVEL_MIN", fmt=_FMT_PCT2, quem=_QUEM_FRANQUEADORA),
        _e("payback_viavel_max", "Payback máximo para viável",
           valor=int(SIM_PAYBACK_VIAVEL_MAX), unidade="meses",
           fonte="config.py SIM_PAYBACK_VIAVEL_MAX", fmt=_FMT_INT, quem=_QUEM_FRANQUEADORA),
    ]


def _blocos_premissas(
    demanda_total: float,
    p: Premissas,
    *,
    obra: float,
    parcelas_obra: int,
    equipamentos: float,
    prazo_equipamentos: int,
    juros_equipamentos_am: float,
    taxa_franquia: float,
    parcelas_franquia: int,
) -> list[tuple[str, list[dict[str, Any]]]]:
    """Especificacao declarativa da aba Premissas, bloco a bloco.

    Todo default vem do objeto `Premissas` recebido — nenhuma constante financeira
    e redigitada aqui. A excecao sao as 7 sublinhas dos outros fixos, que o
    dataclass so tem agregadas e por isso sao RATEADAS proporcionalmente.
    """
    modo_rampa = _MODOS_RAMPA[1] if p.rampa_apenas_balcao else _MODOS_RAMPA[0]
    return [
        ("Demanda e ticket", _bloco_demanda(demanda_total, p, modo_rampa)),
        ("Deduções e impostos", _bloco_impostos(p)),
        ("Custo variável (% da receita líquida)", _bloco_custo_variavel(p)),
        ("Folha", _bloco_folha(p)),
        ("Custos fixos", _bloco_custos_fixos(p)),
        (
            "Investimento e financiamento",
            _bloco_investimento(
                obra=obra, parcelas_obra=parcelas_obra, equipamentos=equipamentos,
                prazo_equipamentos=prazo_equipamentos,
                juros_equipamentos_am=juros_equipamentos_am, taxa_franquia=taxa_franquia,
                parcelas_franquia=parcelas_franquia,
            ),
        ),
        ("Anuidade (taxa de manutenção)", _bloco_anuidade(p)),
        ("Reajuste anual (degrau a partir do mês 13)", _bloco_reajuste(p)),
        ("Taxas mínimas de retorno (negócio e sócio)", _bloco_taxas(p)),
        ("Aluguel-teto (régua de decisão)", _bloco_teto()),
        ("Rampa e horizonte", _bloco_horizonte(p)),
    ]


# ---------------------------------------------------------------------------
# Aba Premissas
# ---------------------------------------------------------------------------

_PREM_COLS = ["Parâmetro", "Valor", "Unidade", "Fonte", "Quem pode alterar"]


def _write_aba_premissas(
    wb: openpyxl.Workbook,
    blocos: list[tuple[str, list[dict[str, Any]]]],
    nome_ponto: str,
) -> dict[str, str]:
    """Escreve a aba Premissas e devolve o mapa `key -> "Premissas!$B$n"`.

    Passo 1 aloca as linhas (para as formulas derivadas poderem se referir umas as
    outras), passo 2 escreve. Sem isso uma derivada nao conseguiria apontar para
    uma celula definida depois dela.
    """
    ws = wb.create_sheet(ABA_PREMISSAS)
    titulo = "ULTRA Academia — Simulador financeiro: PREMISSAS (toda célula de entrada)"
    if nome_ponto:
        titulo = f"{titulo} — {nome_ponto}"
    _write_header(ws, titulo, n_cols=len(_PREM_COLS))
    _nota(
        ws, 2,
        "Amarelo = você pode editar; cinza claro = fórmula derivada (não digite); "
        "cinza escuro = estrutural (mudar aqui não recria as colunas da planilha).",
        len(_PREM_COLS),
    )

    # Passo 1: alocar linhas.
    row = 4
    plano: list[tuple[str, int, int, list[tuple[dict[str, Any], int]]]] = []
    for bloco, entradas in blocos:
        bloco_row = row
        hdr_row = row + 1
        row += 2
        alocadas: list[tuple[dict[str, Any], int]] = []
        for ent in entradas:
            alocadas.append((ent, row))
            row += 1
        plano.append((bloco, bloco_row, hdr_row, alocadas))
        row += 1  # respiro entre blocos

    refs: dict[str, str] = {}
    for _bloco, _brow, _hrow, itens in plano:
        for ent, r in itens:
            refs[ent["key"]] = f"{ABA_PREMISSAS}!$B${r}"

    # Passo 2: escrever.
    for bloco, bloco_row, hdr_row, itens in plano:
        _titulo_bloco(ws, bloco_row, bloco, len(_PREM_COLS))
        _cabecalho_tabela(ws, hdr_row, _PREM_COLS)
        for ent, r in itens:
            _linha_label(ws, r, ent["label"])
            cel = ws.cell(row=r, column=2)
            if ent["formula"] is not None:
                cel.value = _aplicar(ent["formula"], refs)
            else:
                cel.value = ent["valor"]
            _estilo_valor(cel, ent["kind"], ent["fmt"])
            if ent["nota"]:
                cel.comment = Comment(ent["nota"], "Motor de Expansão")
            for col, txt in ((3, ent["unidade"]), (4, ent["fonte"]), (5, ent["quem"])):
                c = ws.cell(row=r, column=col, value=txt)
                c.font = Font(name=_FONTE_PADRAO, color="FF55555F", size=8)
                c.alignment = Alignment(horizontal="left", vertical="center")

    # Validacoes de lista nas premissas de texto.
    _validacao_lista(ws, _MODOS_RAMPA, [refs["modo_rampa"]])
    _validacao_lista(ws, (IR_MODO_FAIXA, "efetivo_legado"), [refs["ir_modo"]])
    _validacao_lista(
        ws, _SIM_NAO, [refs["anuidade_apenas_balcao"], refs["anuidade_pro_rata"]]
    )

    _set_col_width(ws, 1, 52)
    _set_col_width(ws, 2, 18)
    _set_col_width(ws, 3, 16)
    _set_col_width(ws, 4, 58)
    _set_col_width(ws, 5, 26)
    ws.freeze_panes = "B4"
    return refs


def _validacao_lista(ws: Worksheet, valores: tuple[str, ...], celulas: list[str]) -> None:
    """Validacao de lista aplicada a celulas dadas como referencia absoluta."""
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(valores) + '"',
        allow_blank=False,
        showDropDown=False,
    )
    dv.error = "Escolha um dos valores da lista."
    dv.errorTitle = "Valor inválido"
    ws.add_data_validation(dv)
    for ref in celulas:
        dv.add(ref.split("!", 1)[-1].replace("$", ""))


# ---------------------------------------------------------------------------
# Aba Folha
# ---------------------------------------------------------------------------

_FOLHA_COLS = ["Cargo", "Qtd", "Salário base", "Encargos %", "Custo total"]


def _write_aba_folha(wb: openpyxl.Workbook, p: Premissas, refs: dict[str, str],
                     col_steady: str) -> None:
    """Quadro de pessoal EDITAVEL + interruptor de modo da folha.

    O total do quadro alimenta a DRE SOMENTE quando o interruptor esta em
    "quadro de pessoal". O default e "percentual do faturamento maduro" para o arquivo
    abrir reproduzindo o motor ao centavo (decisao de Felipe, 2026-07-24).

    Nos DOIS modos a folha e FIXA no tempo: no modo percentual ela e
    `folha_pct x faturamento MADURO` (nao o faturamento do mes), no modo quadro e o
    total desta aba. Quem monta o valor unico e a celula "Folha mensal FIXA" da aba
    Premissas; esta aba so alimenta um dos dois lados do interruptor.
    """
    ws = wb.create_sheet(ABA_FOLHA)
    _write_header(ws, "FOLHA DE PAGAMENTO — quadro de pessoal editável", n_cols=len(_FOLHA_COLS))
    _nota(
        ws, 2,
        "ATENÇÃO: os salários abaixo são ESTIMATIVA da equipe de expansão. Este dado "
        "NÃO existe nas bases do sistema.",
        len(_FOLHA_COLS), alerta=True,
    )
    _nota(
        ws, 3,
        "Substitua cargo por cargo pela folha REAL da unidade antes de levar o número ao "
        "investidor. O total desta aba só entra na DRE se o interruptor abaixo estiver em "
        '"quadro de pessoal".',
        len(_FOLHA_COLS),
    )

    # Interruptor de modo.
    _linha_label(ws, _FOLHA_MODO_ROW, "Modo da folha", negrito=True)
    modo_default = (
        _MODOS_FOLHA[1] if p.pessoal_mes_override is not None else _MODOS_FOLHA[0]
    )
    cel_modo = ws.cell(row=_FOLHA_MODO_ROW, column=2, value=modo_default)
    _estilo_valor(cel_modo, _EDIT, None)
    cel_modo.alignment = Alignment(horizontal="left", vertical="center")
    cel_modo.comment = Comment(
        'DEFAULT "percentual do faturamento maduro": nesse modo a folha é '
        "folha_pct x faturamento de REGIME PLENO — um valor único, pago integralmente "
        "desde o mês 1 — e o arquivo reproduz o motor do sistema ao centavo. Troque para "
        '"quadro de pessoal" para a folha passar a ser o TOTAL desta aba (também fixo no '
        "tempo) — aí o arquivo reflete a folha real e pode divergir do sistema de propósito. "
        "Nos dois modos a folha NÃO acompanha a rampa de alunos.",
        "Motor de Expansão",
    )
    _validacao_lista(ws, _MODOS_FOLHA, [f"{ABA_FOLHA}!$B${_FOLHA_MODO_ROW}"])
    for col, txt in ((3, "lista"), (4, "Interruptor desta planilha"), (5, "Operador")):
        c = ws.cell(row=_FOLHA_MODO_ROW, column=col, value=txt)
        c.font = Font(name=_FONTE_PADRAO, color="FF55555F", size=8)
    _nota(
        ws, _FOLHA_MODO_ROW + 1,
        'A DRE lê UMA célula (Premissas: "Folha mensal FIXA") = SE(modo="quadro de pessoal"; '
        "Folha!TOTAL; folha_% x faturamento MADURO), repetida nos 64 meses e só corrigida "
        "pelo reajuste anual de custos. A folha NÃO escala com a rampa de alunos.",
        len(_FOLHA_COLS),
    )

    # Quadro de pessoal.
    _cabecalho_tabela(ws, _FOLHA_HDR_ROW, _FOLHA_COLS)
    quadro = _quadro_pessoal(p)
    for i, (cargo, qtd, salario) in enumerate(quadro):
        r = _FOLHA_CARGO_ROW_INI + i
        c_cargo = ws.cell(row=r, column=1, value=cargo)
        c_cargo.font = _body_font()
        c_cargo.alignment = Alignment(horizontal="left", vertical="center")
        c_cargo.fill = _fill(_AMARELO_CLR)
        c_cargo.border = _BORDA_FINA

        c_qtd = ws.cell(row=r, column=2, value=int(qtd))
        _estilo_valor(c_qtd, _EDIT, _FMT_INT)
        c_sal = ws.cell(row=r, column=3, value=float(salario))
        _estilo_valor(c_sal, _EDIT, _FMT_CONTABIL)
        c_sal.comment = Comment(_NOTA_ESTIMATIVA, "Motor de Expansão")
        c_enc = ws.cell(row=r, column=4, value=_ENCARGOS_DEFAULT)
        _estilo_valor(c_enc, _EDIT, _FMT_PCT2)
        c_enc.comment = Comment(
            "ESTIMATIVA de encargos (INSS patronal, FGTS, férias, 13o, provisões). "
            "Substitua pelo percentual apurado pela contabilidade da unidade.",
            "Motor de Expansão",
        )
        c_tot = ws.cell(row=r, column=5, value=f"=B{r}*C{r}*(1+D{r})")
        _estilo_valor(c_tot, _DERIV, _FMT_CONTABIL)

    # TOTAL.
    r_tot = _FOLHA_TOTAL_ROW
    ini, fim = _FOLHA_CARGO_ROW_INI, _FOLHA_TOTAL_ROW - 1
    _linha_label(ws, r_tot, "TOTAL DA FOLHA (com encargos)", negrito=True)
    c_hc = ws.cell(row=r_tot, column=2, value=f"=SUM(B{ini}:B{fim})")
    _estilo_valor(c_hc, _DERIV, _FMT_INT)
    for col in (3, 4):
        cc = ws.cell(row=r_tot, column=col, value=None)
        cc.fill = _fill(_CINZA_TRAVADO)
        cc.border = _BORDA_FINA
    c_tt = ws.cell(row=r_tot, column=5, value=f"=SUM(E{ini}:E{fim})")
    _estilo_valor(c_tt, _DERIV, _FMT_CONTABIL)
    c_tt.font = Font(name=_FONTE_PADRAO, bold=True, color=_CINZA_ESC, size=11)

    # Afericao do quadro contra a regua do % do faturamento MADURO.
    folha_dre = f"'{ABA_DRE}'!{col_steady}{_DRE_ROW['folha']}"
    afer = [
        (
            "Folha efetivamente usada pela DRE (mês de referência)",
            f"={folha_dre}",
            "Depende do interruptor acima",
        ),
        (
            "Folha FIXA da régua do motor: % x faturamento MADURO",
            f"={refs['folha_fixa']}",
            "Vale desde o mês 1",
        ),
        (
            "Aferido vs a régua (TOTAL do quadro - % x faturamento maduro)",
            f"=E{r_tot}-{refs['folha_pct']}*{refs['fat_maduro']}",
            "Quanto o quadro estimado se afasta da régua",
        ),
    ]
    for i, (label, formula, obs) in enumerate(afer):
        r = _FOLHA_AFER_ROW + i
        _linha_label(ws, r, label, negrito=(i == 2))
        cel = ws.cell(row=r, column=5, value=formula)
        _estilo_valor(cel, _DERIV, _FMT_CONTABIL)
        c_obs = ws.cell(row=r, column=4, value=obs)
        c_obs.font = Font(name=_FONTE_PADRAO, color="FF55555F", size=8)
        c_obs.alignment = Alignment(horizontal="right", vertical="center")

    _set_col_width(ws, 1, 52)
    _set_col_width(ws, 2, 10)
    _set_col_width(ws, 3, 16)
    _set_col_width(ws, 4, 32)
    _set_col_width(ws, 5, 18)
    ws.freeze_panes = f"A{_FOLHA_HDR_ROW + 1}"


def _quadro_pessoal(p: Premissas) -> list[tuple[str, int, float]]:
    """Quadro default; reescalado quando o cenario traz folha ABSOLUTA (override).

    No caminho de override o total do quadro TEM de fechar exatamente com
    `pessoal_mes_override`, senao a DRE (que nesse caso ja abre no modo "quadro de
    pessoal") divergiria do motor por centavos de arredondamento. O ultimo cargo
    absorve o residuo com precisao cheia — o formato contabil da celula exibe
    arredondado, mas o valor armazenado fecha a conta.
    """
    base = list(_QUADRO_PESSOAL_DEFAULT)
    if p.pessoal_mes_override is None:
        return [(c, q, s) for c, q, s in base]
    total_default = sum(q * s for _c, q, s in base) * (1.0 + _ENCARGOS_DEFAULT)
    if total_default <= 0 or not base:
        return [(c, q, s) for c, q, s in base]
    escala = float(p.pessoal_mes_override) / total_default
    ajustado = [(c, q, round(s * escala, 2)) for c, q, s in base[:-1]]
    c_ult, q_ult, _s_ult = base[-1]
    alvo_sem_encargos = float(p.pessoal_mes_override) / (1.0 + _ENCARGOS_DEFAULT)
    resto = alvo_sem_encargos - sum(q * s for _c, q, s in ajustado)
    ajustado.append((c_ult, q_ult, (resto / q_ult) if q_ult else 0.0))
    return ajustado


# ---------------------------------------------------------------------------
# Linha do tempo (compartilhada por DRE mensal e Fluxo de caixa)
# ---------------------------------------------------------------------------


def _meses_da_serie(p: Premissas) -> list[int]:
    """M-pre..M-1 e M1..M-horizonte, na MESMA ordem da serie do motor."""
    pre = max(int(p.meses_pre_abertura), 0)
    horizonte = max(int(p.horizonte_meses), 1)
    return [i - pre for i in range(pre)] + list(range(1, horizonte + 1))


def _fmt_por_tipo(tipo: str) -> str | None:
    return {
        "int": _FMT_INT,
        "txt": None,
        "fator": "0.0000",
        "alunos": _FMT_ALUNOS,
        "brl": _FMT_CONTABIL,
        "pct": _FMT_PCT2,
    }[tipo]


def _write_hdr_linha_do_tempo(
    ws: Worksheet, meses: list[int], titulo: str, nota: str,
    ordem: tuple[tuple[str, str, str, bool, bool], ...],
    *, col_total: int | None, col_steady: int | None, label_steady: str,
) -> None:
    n_cols = (col_steady or col_total or (_MES_COL_INI + len(meses) - 1))
    _write_header(ws, titulo, n_cols=n_cols)
    _nota(ws, _TL_NOTA_ROW, nota, n_cols)

    c0 = ws.cell(row=_TL_HDR_ROW, column=1, value="Linha")
    c0.fill = _fill(_CINZA_ESC)
    c0.font = Font(name=_FONTE_PADRAO, bold=True, color=_BRANCO_HEX, size=9)
    c0.alignment = Alignment(horizontal="left", vertical="center")
    for j, mes in enumerate(meses):
        c = ws.cell(row=_TL_HDR_ROW, column=_MES_COL_INI + j, value=f"M{mes}")
        c.fill = _fill(_CINZA_ESC if mes >= 1 else "FF4A4C60")
        c.font = Font(name=_FONTE_PADRAO, bold=True, color=_BRANCO_HEX, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
    if col_total:
        c = ws.cell(row=_TL_HDR_ROW, column=col_total, value="Total do horizonte")
        c.fill = _fill(_TURQUESA_HEX)
        c.font = Font(name=_FONTE_PADRAO, bold=True, color=_BRANCO_HEX, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if col_steady:
        c = ws.cell(row=_TL_HDR_ROW, column=col_steady, value=label_steady)
        c.fill = _fill(_TURQUESA_HEX)
        c.font = Font(name=_FONTE_PADRAO, bold=True, color=_BRANCO_HEX, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (_key, label, _tipo, sub, negrito) in enumerate(ordem):
        _linha_label(
            ws, _TL_DADOS_ROW_INI + i, label, negrito=negrito, sub=sub,
            faixa=_CINZA_CLR if negrito else None,
        )

    _set_col_width(ws, 1, 52)
    for j in range(len(meses)):
        _set_col_width(ws, _MES_COL_INI + j, 15)
    if col_total:
        _set_col_width(ws, col_total, 18)
    if col_steady:
        _set_col_width(ws, col_steady, 18)
    ws.freeze_panes = ws.cell(row=_TL_DADOS_ROW_INI, column=_MES_COL_INI).coordinate


def _dre_formulas(j: int, meses: list[int], refs: dict[str, str]) -> dict[str, Any]:
    """Formulas de UMA coluna de mes da DRE. `j` e o indice na linha do tempo."""
    L = get_column_letter(_MES_COL_INI + j)
    R = _DRE_ROW
    P = refs
    mes = meses[j]

    def c(key: str) -> str:
        return f"{L}{R[key]}"

    t = c("t")
    frac = f"MIN({t}/{P['maturacao']},1)"
    rampa_total = f"{P['alunos_ini_ef']}+({P['demanda']}-{P['alunos_ini_ef']})*{frac}"
    rampa_bal = f"{P['alunos_ini_ef']}+({P['demanda']}*{P['share']}-{P['alunos_ini_ef']})*{frac}"

    f: dict[str, Any] = {
        # A linha do tempo e ESTATICA: e a estrutura do arquivo, nao um resultado.
        "mes": mes,
        "mes_contrato": j + 1,
        "t": f"=MAX(0,{c('mes')})",
        "fase": f'=IF({t}>0,"Operação","Pré-operacional")',
        "f_ticket": f"=IF({t}<1,1,(1+{P['reajuste_ticket']})^INT(({t}-1)/12))",
        "f_aluguel": f"=IF({t}<1,1,(1+{P['reajuste_aluguel']})^INT(({t}-1)/12))",
        "f_custos": f"=IF({t}<1,1,(1+{P['reajuste_custos']})^INT(({t}-1)/12))",
        "alunos_total": (
            f"=IF({t}=0,0,IF({P['modo_rampa']}=\"apenas balcão\","
            f"({rampa_bal})+{P['demanda']}*(1-{P['share']}),{rampa_total}))"
        ),
        "alunos_balcao": (
            f"=IF({t}=0,0,IF({P['modo_rampa']}=\"apenas balcão\","
            f"{rampa_bal},{c('alunos_total')}*{P['share']}))"
        ),
        "alunos_agregadores": f"={c('alunos_total')}-{c('alunos_balcao')}",
        "rec_balcao": (
            f"={c('alunos_balcao')}*(1-{P['churn']})*{P['ticket_cheio']}"
            f"*{c('f_ticket')}*(1-{P['inadimplencia']})"
        ),
        "rec_agregadores": (
            f"={c('alunos_agregadores')}*{P['ticket_agregador']}"
            f"*{c('f_ticket')}*(1-{P['inadimplencia']})"
        ),
        "rec_personal": f"=IF({t}>0,{P['personal_mes']},0)",
        "rec_anuidade": (
            f"=IF(AND({t}>0,{P['anuidade_valor']}>0,{t}>={P['anuidade_mes_inicio']}),"
            f"IF({P['anuidade_apenas_balcao']}=\"Sim\",{c('alunos_balcao')},"
            f"{c('alunos_total')})*{P['anuidade_por_aluno']},0)"
        ),
        "faturamento": (
            f"={c('rec_balcao')}+{c('rec_agregadores')}+{c('rec_personal')}"
            f"+{c('rec_anuidade')}"
        ),
        "deducoes": f"={c('faturamento')}*{P['devolucoes']}",
        "receita_liquida": f"={c('faturamento')}-{c('deducoes')}",
        "impostos": f"={c('receita_liquida')}*{P['impostos_total']}",
        "receita_pos_impostos": f"={c('receita_liquida')}-{c('impostos')}",
        "cvar_royalties": f"={c('receita_liquida')}*{P['royalties']}",
        "cvar_marketing": f"={c('receita_liquida')}*{P['marketing']}",
        "cvar_manutencao": f"={c('receita_liquida')}*{P['manutencao']}",
        "cvar_cartoes": f"={c('receita_liquida')}*{P['cartoes']}",
        "cvar_total": (
            f"={c('cvar_royalties')}+{c('cvar_marketing')}+{c('cvar_manutencao')}"
            f"+{c('cvar_cartoes')}"
        ),
        # FOLHA FIXA: uma UNICA celula de Premissas nos 64 meses, so corrigida pelo
        # reajuste anual de custos. Antes era `folha_pct * faturamento DO MES`, que
        # encolhia com a rampa — o defeito reportado.
        "folha": f"=IF({t}=0,0,{P['folha_fixa']}*{c('f_custos')})",
        "outros_total": f"=SUM({c(_OF_KEY_INI)}:{c(_OF_KEY_FIM)})",
        "aluguel": (
            f"=IF({c('mes_contrato')}<={P['carencia']},0,{P['aluguel']}"
            f"*IF({t}>0,{c('f_aluguel')},1))"
        ),
        "custo_pre_op": f"=IF({t}=0,{P['custo_pre_op']},0)",
        "custos_op": (
            f"={c('cvar_total')}+{c('folha')}+{c('outros_total')}+{c('aluguel')}"
            f"+{c('custo_pre_op')}"
        ),
        "ebitda": f"={c('receita_pos_impostos')}-{c('custos_op')}",
        "margem_ebitda": (
            f"=IF({c('faturamento')}>0,{c('ebitda')}/{c('faturamento')},0)"
        ),
        "ir_base": f"={c('faturamento')}*{P['base_presumida']}",
        "ir_irpj": f"={c('ir_base')}*{P['irpj_al']}",
        "ir_adicional": (
            f"=MAX(0,{c('ir_base')}-{P['irpj_adic_limite']})*{P['irpj_adic_al']}"
        ),
        "ir_csll_al": f"={c('ir_base')}*{P['csll_al']}",
        "ir_total": (
            f'=IF({P["ir_modo"]}="{IR_MODO_FAIXA}",'
            f"{c('ir_irpj')}+{c('ir_adicional')}+{c('ir_csll_al')},"
            f"{c('receita_liquida')}*({P['ir_efetivo']}+{P['csll_efetivo']}))"
        ),
        "juros": f"='{ABA_FLUXO}'!{L}{_FLX_ROW['juros']}",
        "resultado_desalav": f"={c('ebitda')}-{c('ir_total')}",
        "resultado_apos_juros": f"={c('resultado_desalav')}-{c('juros')}",
    }
    for key, _lbl, _valor in _OUTROS_FIXOS_DECOMP:
        f[key] = f"=IF({t}=0,0,{P[key]}*{c('f_custos')})"
    return f


def _write_aba_dre(
    wb: openpyxl.Workbook, meses: list[int], refs: dict[str, str],
) -> tuple[int, str]:
    """Uma COLUNA por mes, uma LINHA por linha do DRE, tudo em formula.

    Devolve (indice da coluna de steady, letra dessa coluna).
    """
    ws = wb.create_sheet(ABA_DRE)
    col_steady = _MES_COL_INI + len(meses)
    letra_steady = get_column_letter(col_steady)
    _write_hdr_linha_do_tempo(
        ws, meses, "DRE MENSAL — M-4 a M60, tudo em fórmula sobre Premissas e Folha",
        "Reajuste anual por degrau a partir do mês 13; carência de aluguel contada de M-4; "
        "custo operacional INTEGRAL desde o mês 1 — inclusive a FOLHA, que é fixa e não "
        "acompanha a rampa (é por isso que o EBITDA do mês 1 é bem negativo).",
        _DRE_ORDEM, col_total=None, col_steady=col_steady,
        label_steady="Steady (regime pleno)",
    )

    for j in range(len(meses)):
        f = _dre_formulas(j, meses, refs)
        col = _MES_COL_INI + j
        for key, _label, tipo, _sub, negrito in _DRE_ORDEM:
            cel = ws.cell(row=_DRE_ROW[key], column=col, value=f[key])
            cel.number_format = _fmt_por_tipo(tipo) or "General"
            cel.font = Font(name=_FONTE_PADRAO, bold=negrito, color=_CINZA_ESC, size=9)
            cel.alignment = Alignment(
                horizontal="center" if tipo in ("int", "txt") else "right",
                vertical="center",
            )
            if negrito:
                cel.fill = _fill(_CINZA_CLR)

    # Coluna de steady: INDEX/MATCH pelo mes de referencia (nao um numero cravado).
    ini = get_column_letter(_MES_COL_INI)
    fim = get_column_letter(_MES_COL_INI + len(meses) - 1)
    r_mes = _DRE_ROW["mes"]
    for key, _label, tipo, _sub, negrito in _DRE_ORDEM:
        r = _DRE_ROW[key]
        cel = ws.cell(
            row=r, column=col_steady,
            value=(
                f"=INDEX(${ini}{r}:${fim}{r},"
                f"MATCH({refs['mes_steady']},${ini}${r_mes}:${fim}${r_mes},0))"
            ),
        )
        cel.number_format = _fmt_por_tipo(tipo) or "General"
        cel.font = Font(name=_FONTE_PADRAO, bold=True, color=_CINZA_ESC, size=9)
        cel.alignment = Alignment(horizontal="right", vertical="center")
        cel.fill = _fill(_VERDE_CLR if negrito else _CINZA_CLR)
        cel.border = _BORDA_FINA
    return col_steady, letra_steady


# ---------------------------------------------------------------------------
# Aba Fluxo de caixa
# ---------------------------------------------------------------------------

_FLX_SEM_TOTAL = frozenset(
    {"mes", "mes_contrato", "t", "saldo_inicial", "saldo_final", "fcf_acumulado",
     "payback", "margem_ebitda"}
)


def _flx_formulas(j: int, refs: dict[str, str]) -> dict[str, str]:
    """Formulas de UMA coluna de mes do fluxo de caixa (Price + payback)."""
    L = get_column_letter(_MES_COL_INI + j)
    ant = get_column_letter(_MES_COL_INI + j - 1) if j > 0 else None
    F, D, P = _FLX_ROW, _DRE_ROW, refs

    def c(key: str) -> str:
        return f"{L}{F[key]}"

    def a(key: str) -> str:
        return f"{ant}{F[key]}"

    t = c("t")
    dre = f"'{ABA_DRE}'!{L}"
    # Mes em que o equipamento financiado seria desembolsado se NAO houvesse banco:
    # a vespera da abertura (ultimo mes de pre-abertura). Sem pre-abertura o motor cai
    # no primeiro mes de operacao — as duas pernas do IF reproduzem `mes_equip`.
    mes_da_vespera = (
        f"IF({P['meses_pre']}>0,{c('mes_contrato')}={P['meses_pre']},"
        f"{c('mes_contrato')}=1)"
    )
    saldo_ini = (
        f"=IF({t}=1,{P['equip_financiado']},0)"
        if ant is None
        else f"=IF({t}=1,{P['equip_financiado']},{a('saldo_final')})"
    )
    dentro_prazo = f"AND({t}>=1,{t}<={P['prazo_equip']})"
    return {
        "mes": f"={dre}{D['mes']}",
        "mes_contrato": f"={dre}{D['mes_contrato']}",
        "t": f"={dre}{D['t']}",
        "ebitda": f"={dre}{D['ebitda']}",
        "ir_csll": f"={dre}{D['ir_total']}",
        "saldo_inicial": saldo_ini,
        "juros": f"=IF({dentro_prazo},{c('saldo_inicial')}*{P['juros_am']},0)",
        "pmt": f"=IF({dentro_prazo},{P['pmt']},0)",
        "amortizacao": f"=MAX(0,{c('pmt')}-{c('juros')})",
        "saldo_final": f"=MAX(0,{c('saldo_inicial')}-{c('amortizacao')})",
        "inv_obra": (
            f"=IF(AND({P['obra']}>0,{c('mes_contrato')}<=MAX({P['parcelas_obra']},1)),"
            f"{P['obra']}/MAX({P['parcelas_obra']},1),0)"
        ),
        # Taxa de franquia PARCELADA sem juros: a parcela cai nos meses de CONTRATO
        # 1..N (M-4..M-1 com N=4), junto da obra — nao a taxa inteira no M-4. O motor
        # trata `parcelas <= 0` como 1 parcela, e o MAX(...,1) reproduz isso.
        "inv_franquia": (
            f"=IF(AND({P['taxa_franquia']}>0,"
            f"{c('mes_contrato')}<=MAX({P['parcelas_franquia']},1)),"
            f"{P['franquia_parcela']},0)"
        ),
        "inv_equip_vista": (
            f"=IF(AND({P['equip_financiado']}=0,{P['meses_pre']}>0,"
            f"{c('mes_contrato')}={P['meses_pre']}),{P['equip']},0)"
        ),
        "inv_capex_renov": f"=IF({t}={P['horizonte']},{P['capex_renov']},0)",
        "investimento": (
            f"={c('inv_obra')}+{c('inv_franquia')}+{c('inv_equip_vista')}"
            f"+{c('inv_capex_renov')}"
        ),
        "valor_residual": f"=IF({t}={P['horizonte']},{P['valor_residual']},0)",
        "fcf": (
            f"={c('ebitda')}-{c('ir_csll')}-{c('pmt')}-{c('investimento')}"
            f"+{c('valor_residual')}"
        ),
        "fcf_acumulado": (
            f"={c('fcf')}" if ant is None else f"={a('fcf_acumulado')}+{c('fcf')}"
        ),
        "payback": (
            f'=IF({c("fcf_acumulado")}>=0,"PAYBACK","")'
            if ant is None
            else f'=IF(AND({c("fcf_acumulado")}>=0,{a("fcf_acumulado")}<0),"PAYBACK","")'
        ),
        # --- Ótica DO NEGÓCIO ------------------------------------------------
        # Quando o equipamento é financiado ele NÃO passa pelo caixa da série (o banco
        # paga), então aqui ele volta como saída real na véspera da abertura — senão
        # estaríamos medindo o ativo com o dinheiro de outra pessoa.
        "neg_equip": f"=IF(AND({P['equip_financiado']}>0,{mes_da_vespera}),"
                     f"{P['equip_financiado']},0)",
        # Sem a PMT e SEM o valor residual — exatamente o fluxo que o motor desconta
        # em `vpl_negocio` / `tir_negocio_*`. Se o valor residual deixar de ser zero,
        # esta linha continua reproduzindo o motor (que também o ignora nesta ótica).
        "fluxo_negocio": (
            f"={c('ebitda')}-{c('ir_csll')}-{c('investimento')}-{c('neg_equip')}"
        ),
        # --- Ótica DA DÍVIDA -------------------------------------------------
        # A captação é o MESMO principal do desembolso acima, com o sinal invertido:
        # referenciar a célula garante que as duas nunca se descolem.
        "div_captacao": f"={c('neg_equip')}",
        "fluxo_divida": f"={c('div_captacao')}-{c('pmt')}",
    }


def _write_aba_fluxo(
    wb: openpyxl.Workbook, meses: list[int], refs: dict[str, str],
) -> tuple[int, int, str]:
    """Devolve (col_total, col_steady, letra_steady) do fluxo de caixa."""
    ws = wb.create_sheet(ABA_FLUXO)
    col_total = _MES_COL_INI + len(meses)
    col_steady = col_total + 1
    _write_hdr_linha_do_tempo(
        ws, meses, "FLUXO DE CAIXA — mesma linha do tempo da DRE",
        "A PMT é NOMINAL (não reajusta). Price: juros do mês = saldo devedor anterior x taxa. "
        "O CAPEX aparece INTEIRO aqui, então o payback do gráfico e o do KPI são o mesmo número. "
        "TRÊS fluxos, nunca somados entre si: DO SÓCIO (a PMT inteira sai), DO NEGÓCIO (sem "
        "financiamento, CAPEX inteiro desembolsado) e DA DÍVIDA (captação menos as PMT).",
        _FLX_ORDEM, col_total=col_total, col_steady=col_steady,
        label_steady="Steady (regime pleno)",
    )

    for j in range(len(meses)):
        f = _flx_formulas(j, refs)
        col = _MES_COL_INI + j
        for key, _label, tipo, _sub, negrito in _FLX_ORDEM:
            cel = ws.cell(row=_FLX_ROW[key], column=col, value=f[key])
            cel.number_format = _fmt_por_tipo(tipo) or "General"
            cel.font = Font(name=_FONTE_PADRAO, bold=negrito, color=_CINZA_ESC, size=9)
            cel.alignment = Alignment(
                horizontal="center" if tipo in ("int", "txt") else "right",
                vertical="center",
            )
            if negrito:
                cel.fill = _fill(_CINZA_CLR)

    ini = get_column_letter(_MES_COL_INI)
    fim = get_column_letter(_MES_COL_INI + len(meses) - 1)
    r_mes = _FLX_ROW["mes"]
    for key, _label, tipo, _sub, negrito in _FLX_ORDEM:
        r = _FLX_ROW[key]
        if key not in _FLX_SEM_TOTAL:
            ct = ws.cell(row=r, column=col_total, value=f"=SUM({ini}{r}:{fim}{r})")
            ct.number_format = _fmt_por_tipo(tipo) or "General"
            ct.font = Font(name=_FONTE_PADRAO, bold=True, color=_CINZA_ESC, size=9)
            ct.alignment = Alignment(horizontal="right", vertical="center")
            ct.fill = _fill(_CINZA_CLR)
            ct.border = _BORDA_FINA
        cs = ws.cell(
            row=r, column=col_steady,
            value=(
                f"=INDEX(${ini}{r}:${fim}{r},"
                f"MATCH({refs['mes_steady']},${ini}${r_mes}:${fim}${r_mes},0))"
            ),
        )
        cs.number_format = _fmt_por_tipo(tipo) or "General"
        cs.font = Font(name=_FONTE_PADRAO, bold=True, color=_CINZA_ESC, size=9)
        cs.alignment = Alignment(horizontal="right", vertical="center")
        cs.fill = _fill(_VERDE_CLR if negrito else _CINZA_CLR)
        cs.border = _BORDA_FINA

    return col_total, col_steady, get_column_letter(col_steady)


# ---------------------------------------------------------------------------
# Aba Investimento
# ---------------------------------------------------------------------------

_INVEST_COLS = ["Item", "Valor", "Unidade", "Fonte / fórmula"]


def _write_aba_investimento(
    wb: openpyxl.Workbook, meses: list[int], refs: dict[str, str], col_total_fluxo: int,
) -> None:
    ws = wb.create_sheet(ABA_INVESTIMENTO)
    _write_header(ws, "INVESTIMENTO E FINANCIAMENTO", n_cols=len(_INVEST_COLS))
    _nota(
        ws, 2,
        "Obra = aporte do franqueado, parcelada sem juros. Equipamentos = financiados (Price). "
        "Taxa de franquia PARCELADA sem juros, junto da obra: as parcelas caem nos meses de "
        "contrato 1 a N, contados da entrega da unidade. Edite os valores na aba Premissas.",
        len(_INVEST_COLS),
    )
    _cabecalho_tabela(ws, 4, _INVEST_COLS)

    P = refs
    ult = get_column_letter(_MES_COL_INI + len(meses) - 1)
    letra_total = get_column_letter(col_total_fluxo)
    r_juros = _FLX_ROW["juros"]
    r_saldo = _FLX_ROW["saldo_final"]

    linhas: list[tuple[str, str, str, str, str | None]] = [
        ("obra", "Obra (aporte)", f"={P['obra']}", "R$", _FMT_CONTABIL),
        ("parcelas_obra", "Parcelas da obra", f"={P['parcelas_obra']}", "meses", _FMT_INT),
        ("obra_parcela", "Parcela mensal da obra",
         f"=IF({P['parcelas_obra']}>0,{P['obra']}/{P['parcelas_obra']},0)", "R$/mês",
         _FMT_CONTABIL),
        ("equip", "Equipamentos", f"={P['equip']}", "R$", _FMT_CONTABIL),
        ("prazo_equip", "Prazo do financiamento", f"={P['prazo_equip']}", "meses", _FMT_INT),
        ("juros_am", "Juros do financiamento", f"={P['juros_am']}", "% a.m.", _FMT_PCT3),
        ("taxa_franquia", "Taxa de franquia", f"={P['taxa_franquia']}", "R$", _FMT_CONTABIL),
        ("parcelas_franquia", "Parcelas da franquia", f"={P['parcelas_franquia']}", "meses",
         _FMT_INT),
        ("franquia_parcela", "Valor da parcela da franquia",
         f"=IF({P['taxa_franquia']}>0,{P['taxa_franquia']}/MAX({P['parcelas_franquia']},1),0)",
         "R$/mês", _FMT_CONTABIL),
        ("capex_total", "CAPEX total (obra + equipamentos)",
         f"={P['obra']}+{P['equip']}", "R$", _FMT_CONTABIL),
        ("investimento_total", "Investimento total (CAPEX + franquia)",
         f"=B{_INVEST_ROW['capex_total']}+{P['taxa_franquia']}", "R$", _FMT_CONTABIL),
        ("aporte_inicial", "Aporte inicial (obra + franquia)",
         f"={P['obra']}+{P['taxa_franquia']}", "R$", _FMT_CONTABIL),
        ("equip_financiado", "Principal financiado",
         f"=IF(AND({P['equip']}>0,{P['prazo_equip']}>0),{P['equip']},0)", "R$", _FMT_CONTABIL),
        ("pmt", "PMT do financiamento (Price)",
         f"=IF(AND({P['equip']}>0,{P['prazo_equip']}>0),"
         f"-PMT(MAX({P['juros_am']},0),{P['prazo_equip']},{P['equip']}),0)",
         "R$/mês", _FMT_CONTABIL),
        ("total_pago", "Total pago no financiamento (PMT x prazo)",
         f"=B{_INVEST_ROW['pmt']}*{P['prazo_equip']}", "R$", _FMT_CONTABIL),
        ("juros_totais", "Juros totais do financiamento",
         f"='{ABA_FLUXO}'!{letra_total}{r_juros}", "R$", _FMT_CONTABIL),
        ("saldo_m60", "Saldo devedor no fim do horizonte",
         f"='{ABA_FLUXO}'!{ult}{r_saldo}", "R$", _FMT_CONTABIL),
    ]
    for key, label, formula, unidade, fmt in linhas:
        r = _INVEST_ROW[key]
        _linha_label(ws, r, label, negrito=key in ("capex_total", "investimento_total", "pmt"))
        cel = ws.cell(row=r, column=2, value=formula)
        _estilo_valor(cel, _DERIV, fmt)
        c_un = ws.cell(row=r, column=3, value=unidade)
        c_un.font = Font(name=_FONTE_PADRAO, color="FF55555F", size=8)
        c_fo = ws.cell(row=r, column=4, value=formula.lstrip("="))
        c_fo.font = Font(name=_FONTE_PADRAO, color="FF8A8A99", size=7)

    _set_col_width(ws, 1, 44)
    _set_col_width(ws, 2, 20)
    _set_col_width(ws, 3, 12)
    _set_col_width(ws, 4, 72)
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# Aba Resumo (KPIs do comite, TODOS como formula)
# ---------------------------------------------------------------------------

_RESUMO_ROW_INI = 5

# --- Cor da fonte na aba Resumo (pedido de Felipe, 2026-07-24) ---------------
# SAIDA DE DINHEIRO -> texto VERMELHO fixo. Sao valores que so existem com um sinal:
# custo, imposto, despesa financeira, PMT, juros totais, CAPEX/investimento/equity.
_RESUMO_SAIDA = frozenset(
    {
        "custos_op", "folha", "ir_csll", "pmt", "juros_totais",
        "capex_total", "investimento_total", "aporte_inicial",
    }
)
# Pode ALTERNAR de sinal -> formatacao CONDICIONAL (vermelho so quando negativo).
# Pintar de vermelho fixo aqui seria mentir quando o numero virar positivo, e pintar
# de verde fixo seria mentir quando virar negativo.
_RESUMO_ALTERNA = frozenset(
    {
        "ebitda", "margem", "resultado_desalav", "tir_mensal", "tir_anual", "vpl",
        "acumulado_m60", "retorno_desalav", "retorno_equity", "ebitda_m1",
    }
)


def _linhas_resumo(
    meses: list[int], refs: dict[str, str], st_dre: str, st_flx: str,
) -> list[tuple[str, str, str | None, str | None, str]]:
    """(key, rotulo, formula, formato, observacao). key vazia = separador."""
    P = refs
    D, F = _DRE_ROW, _FLX_ROW
    ini = get_column_letter(_MES_COL_INI)
    fim = get_column_letter(_MES_COL_INI + len(meses) - 1)
    col_m1 = get_column_letter(_MES_COL_INI + meses.index(1)) if 1 in meses else ini
    fcf = f"'{ABA_FLUXO}'!${ini}${F['fcf']}:${fim}${F['fcf']}"
    dre_st = f"'{ABA_DRE}'!{st_dre}"
    flx_st = f"'{ABA_FLUXO}'!{st_flx}"

    # O custo fixo do break-even agora INCLUI a folha (ela deixou de ser percentual da
    # receita do mes), e o `k` deixou de subtrai-la. Trocar so um dos dois lados daria
    # um break-even sem defesa: 840,6 alunos no caso de referencia contra 1.152 reais.
    fat_be = f"({P['custo_fixo_total']}+{P['aluguel']})/{P['k_ebitda']}"
    fat_be_caixa = f"({P['custo_fixo_total']}+{P['aluguel']}+{P['pmt']})/{P['k_ebitda']}"

    def be(fat: str) -> str:
        return (
            f"=IF(AND({P['k_ebitda']}>0,{P['receita_por_aluno']}>0),"
            f"MAX(0,(({fat})-{P['personal_mes']})/{P['receita_por_aluno']}),\"Inviável\")"
        )

    return [
        ("mes_ref", "Mês de referência (regime pleno)", f"={P['mes_steady']}", _FMT_INT,
         "Alunos maduros E anuidade já em cobrança"),
        ("alunos_ref", "Alunos totais no mês de referência", f"={dre_st}{D['alunos_total']}",
         _FMT_ALUNOS, "Demanda na maturidade"),
        ("faturamento", "Faturamento bruto / mês", f"={dre_st}{D['faturamento']}",
         _FMT_CONTABIL, "Mensalidades + personal + anuidade"),
        ("receita_anuidade", "  dos quais anuidade", f"={dre_st}{D['rec_anuidade']}",
         _FMT_CONTABIL, "Pró-rata mensal, só balcão"),
        ("receita_liquida", "Receita líquida / mês", f"={dre_st}{D['receita_liquida']}",
         _FMT_CONTABIL, "Bruta menos deduções"),
        ("receita_pos_impostos", "Receita pós-impostos / mês",
         f"={dre_st}{D['receita_pos_impostos']}", _FMT_CONTABIL, "PIS + COFINS + ISS"),
        ("custos_op", "Custos operacionais / mês", f"={dre_st}{D['custos_op']}",
         _FMT_CONTABIL, "Variável + folha + outros fixos + aluguel"),
        ("folha", "  dos quais folha (FIXA desde o mês 1)", f"={dre_st}{D['folha']}",
         _FMT_CONTABIL, "% do faturamento MADURO, não do faturamento do mês"),
        ("ebitda", "EBITDA / mês", f"={dre_st}{D['ebitda']}", _FMT_CONTABIL, ""),
        ("margem", "Margem EBITDA", f"={dre_st}{D['margem_ebitda']}", _FMT_PCT2,
         "Sobre o faturamento bruto"),
        ("ir_csll", "IR/CSLL / mês", f"={dre_st}{D['ir_total']}", _FMT_CONTABIL,
         "Lucro Presumido com a faixa do adicional"),
        ("resultado_desalav", "Resultado desalavancado após IR / mês",
         f"={dre_st}{D['resultado_desalav']}", _FMT_CONTABIL, "EBITDA menos IR/CSLL"),
        ("", "Break-even (em alunos TOTAIS, o mix escala junto)", None, None, ""),
        ("break_even_ebitda", "Break-even de EBITDA", be(fat_be), _FMT_ALUNOS,
         "Alunos totais para EBITDA zero, com a folha dimensionada para a demanda assumida"),
        ("break_even_caixa", "Break-even de caixa (cobre a PMT)", be(fat_be_caixa),
         _FMT_ALUNOS, "Alunos totais para o caixa fechar com o financiamento"),
        ("", "Investimento", None, None, ""),
        ("capex_total", "CAPEX total", f"={P['capex_total']}", _FMT_CONTABIL,
         "Obra + equipamentos"),
        ("investimento_total", "Investimento total", f"={P['investimento_total']}",
         _FMT_CONTABIL, "CAPEX + taxa de franquia"),
        ("aporte_inicial", "Aporte inicial", f"={P['aporte_inicial']}", _FMT_CONTABIL,
         "Obra + franquia — NAO e o cheque total; ver o cheque total abaixo"),
        ("pmt", "PMT do financiamento", f"={P['pmt']}", _FMT_CONTABIL, "Price, nominal"),
        ("juros_totais", "Juros totais do financiamento",
         f"=Investimento!$B${_INVEST_ROW['juros_totais']}", _FMT_CONTABIL,
         "Soma da linha de juros do fluxo"),
        ("", "Retorno", None, None, ""),
        ("payback", "Payback (meses desde M-4)",
         f"=IFERROR(INDEX('{ABA_FLUXO}'!${ini}${F['mes']}:${fim}${F['mes']},"
         f"MATCH(\"PAYBACK\",'{ABA_FLUXO}'!${ini}${F['payback']}:${fim}${F['payback']},0)),"
         '"Não atingido")',
         _FMT_INT, "Primeiro mês com FCF acumulado >= 0"),
        ("tir_mensal", "TIR mensal do sócio", f'=IFERROR(IRR({fcf}),"n/d")', _FMT_PCT3,
         "Sobre a linha de FCF (fluxo DO SÓCIO: PMT inteira sai)"),
        ("tir_anual", "TIR anual do sócio", None, _FMT_PCT2, "(1 + TIR mensal)^12 - 1"),
        ("vpl", "VPL do sócio", f"=NPV({P['taxa_socio_am']},{fcf})*(1+{P['taxa_socio_am']})",
         _FMT_CONTABIL,
         "Fluxo DO SÓCIO descontado à taxa mínima DO SÓCIO (derivada). "
         "Primeiro fluxo NÃO descontado (período 0)"),
        ("acumulado_m60", "FCF acumulado no fim do horizonte",
         f"='{ABA_FLUXO}'!{fim}{F['fcf_acumulado']}", _FMT_CONTABIL, ""),
        ("retorno_desalav", "Retorno anual do negócio",
         f"=IF({P['investimento_total']}>0,{dre_st}{D['resultado_desalav']}*12"
         f"/{P['investimento_total']},0)", _FMT_PCT2,
         "Resultado ANTES da PMT sobre o investimento cheio"),
        ("retorno_equity", "Retorno anual do sócio",
         f"=IF({P['aporte_inicial']}>0,({dre_st}{D['resultado_desalav']}-{flx_st}{F['pmt']})*12"
         f"/{P['aporte_inicial']},0)", _FMT_PCT2,
         "Resultado DEPOIS da PMT sobre o equity aportado"),
        ("", "Ticket", None, None, ""),
        ("ticket_blended", "Ticket blended por aluno total", f"={P['ticket_blended']}",
         _FMT_CONTABIL, "Líquido de churn e inadimplência"),
        ("ticket_agregador", "Ticket do agregador", f"={P['ticket_agregador']}",
         _FMT_CONTABIL, "Fração do ticket cheio"),
        ("receita_por_aluno", "Receita por aluno total (com anuidade)",
         f"={P['receita_por_aluno']}", _FMT_CONTABIL, "Base do break-even"),
        ("", "Aluguel-teto (% do faturamento bruto do mês de referência)", None, None, ""),
        ("teto_ideal", "Aluguel-teto — faixa ideal",
         f"={dre_st}{D['faturamento']}*{P['teto_ideal_pct']}", _FMT_CONTABIL, "Faixa ideal"),
        ("teto_teto", "Aluguel-teto — TETO (CANÔNICO)",
         f"={dre_st}{D['faturamento']}*{P['teto_pct']}", _FMT_CONTABIL,
         "Este é o número do card grande"),
        ("teto_excecao", "Aluguel-teto — exceção",
         f"={dre_st}{D['faturamento']}*{P['teto_excecao_pct']}", _FMT_CONTABIL,
         "Exceção, não referência"),
        ("", "Leitura do comitê", None, None, ""),
        ("ebitda_m1", "EBITDA do mês 1", f"='{ABA_DRE}'!{col_m1}{D['ebitda']}",
         _FMT_CONTABIL, "Negativo por construção: o custo é integral desde o mês 1"),
        ("flag_viavel", "Viável pelos critérios da franqueadora?", None, None,
         "Margem mínima E payback máximo"),
    ]


def _write_aba_resumo(
    wb: openpyxl.Workbook, meses: list[int], refs: dict[str, str], st_dre: str, st_flx: str,
    nome_ponto: str,
) -> dict[str, str]:
    ws = wb.create_sheet(ABA_RESUMO)
    titulo = "RESUMO — KPIs do comitê (todos em fórmula sobre as abas do modelo)"
    if nome_ponto:
        titulo = f"{titulo} — {nome_ponto}"
    _write_header(ws, titulo, n_cols=3)
    _nota(
        ws, 2,
        "Nenhum número desta aba é digitado: tudo aponta para DRE mensal, Fluxo de caixa e "
        "Premissas. Mude uma premissa e todos estes KPIs se movem.",
        3,
    )
    _nota(
        ws, 3,
        "Números em VERMELHO são SAÍDA DE DINHEIRO (custo, imposto, despesa financeira, PMT, "
        "juros, CAPEX e aporte). Onde o valor pode virar de sinal (EBITDA, resultado, FCF "
        "acumulado, VPL, TIR e retorno), o vermelho aparece SÓ quando o número fica negativo.",
        3,
    )
    _cabecalho_tabela(ws, 4, ["Indicador", "Valor", "Observação"])

    linhas = _linhas_resumo(meses, refs, st_dre, st_flx)
    rrefs: dict[str, str] = {}
    r = _RESUMO_ROW_INI
    for key, _label, _formula, _fmt, _obs in linhas:
        if key:
            rrefs[key] = f"{ABA_RESUMO}!$B${r}"
        r += 1

    # Formulas que dependem de outras celulas do proprio Resumo.
    dinamicas = {
        "tir_anual": f"=IFERROR((1+{rrefs['tir_mensal']})^12-1,\"n/d\")",
        "flag_viavel": (
            f"=IF(AND({rrefs['margem']}>={refs['margem_viavel_min']},"
            f"ISNUMBER({rrefs['payback']}),{rrefs['payback']}<={refs['payback_viavel_max']}),"
            '"Sim","Não")'
        ),
    }

    r = _RESUMO_ROW_INI
    for key, label, formula, fmt, obs in linhas:
        if not key:
            _titulo_bloco(ws, r, label, 3)
            r += 1
            continue
        _linha_label(ws, r, label, negrito=key in ("ebitda", "margem", "payback", "teto_teto"))
        cel = ws.cell(row=r, column=2, value=dinamicas.get(key, formula))
        _estilo_valor(cel, _DERIV, fmt)
        if key in ("ebitda", "margem", "teto_teto"):
            cel.fill = _fill(_VERDE_CLR)
        if key in _RESUMO_SAIDA:
            # Vermelho FIXO: nao ha cenario em que estas linhas sejam entrada de caixa.
            cel.font = Font(
                name=_FONTE_PADRAO, italic=True, bold=False,
                color=_VERMELHO_FONTE, size=10,
            )
        elif key in _RESUMO_ALTERNA:
            ws.conditional_formatting.add(
                f"B{r}",
                CellIsRule(
                    operator="lessThan", formula=["0"],
                    font=Font(name=_FONTE_PADRAO, italic=True, color=_VERMELHO_FONTE),
                ),
            )
        c_obs = ws.cell(row=r, column=3, value=obs)
        c_obs.font = Font(name=_FONTE_PADRAO, color="FF55555F", size=8)
        c_obs.alignment = Alignment(horizontal="left", vertical="center")
        r += 1

    _set_col_width(ws, 1, 48)
    _set_col_width(ws, 2, 20)
    _set_col_width(ws, 3, 54)
    ws.freeze_panes = "A5"
    return rrefs


# ---------------------------------------------------------------------------
# Aba Sensibilidade
# ---------------------------------------------------------------------------


def _write_aba_sensibilidade(wb: openpyxl.Workbook, refs: dict[str, str]) -> str:
    """Grade alunos x aluguel em forma fechada. Devolve a referencia do centro."""
    ws = wb.create_sheet(ABA_SENSIBILIDADE)
    n_cols = _SENS_GRID_COL_INI + len(_SENS_FATORES_ALUGUEL) - 1
    _write_header(ws, "SENSIBILIDADE — margem EBITDA no regime pleno", n_cols=n_cols)
    _nota(
        ws, 2,
        "Forma fechada do motor: faturamento = alunos x receita por aluno + personal; "
        "EBITDA = faturamento x k - (custo fixo total + aluguel), onde o custo fixo já inclui "
        "a FOLHA (que não varia com os alunos desta grade). Verde >= 10%, amarelo >= 0%, "
        "vermelho < 0%.",
        n_cols,
    )
    P = refs
    _linha_label(ws, 4, "Linhas = alunos totais | Colunas = aluguel mensal", negrito=True)

    _linha_label(ws, _SENS_HDR_FATOR_ROW, "Fator do aluguel", sub=True)
    _linha_label(ws, _SENS_HDR_ALUGUEL_ROW, "Aluguel (R$/mês)", negrito=True)
    for i, fa in enumerate(_SENS_FATORES_ALUGUEL):
        col = _SENS_GRID_COL_INI + i
        cf = ws.cell(row=_SENS_HDR_FATOR_ROW, column=col, value=fa)
        _estilo_valor(cf, _EDIT, _FMT_PCT2)
        cf.alignment = Alignment(horizontal="center", vertical="center")
        letra = get_column_letter(col)
        ca = ws.cell(
            row=_SENS_HDR_ALUGUEL_ROW, column=col,
            value=f"={P['aluguel']}*{letra}${_SENS_HDR_FATOR_ROW}",
        )
        _estilo_valor(ca, _DERIV, _FMT_CONTABIL)
        ca.alignment = Alignment(horizontal="center", vertical="center")

    _linha_label(ws, _SENS_HDR_ALUGUEL_ROW - 1, "Fator de alunos / Alunos totais", sub=True)
    ws.cell(row=_SENS_HDR_ALUGUEL_ROW, column=2, value="Alunos totais").font = Font(
        name=_FONTE_PADRAO, bold=True, color=_CINZA_ESC, size=9
    )

    for i, fal in enumerate(_SENS_FATORES_ALUNOS):
        r = _SENS_GRID_ROW_INI + i
        cf = ws.cell(row=r, column=1, value=fal)
        _estilo_valor(cf, _EDIT, _FMT_PCT2)
        cf.alignment = Alignment(horizontal="center", vertical="center")
        cal = ws.cell(row=r, column=2, value=f"={P['demanda']}*$A{r}")
        _estilo_valor(cal, _DERIV, _FMT_ALUNOS)
        for jj in range(len(_SENS_FATORES_ALUGUEL)):
            col = _SENS_GRID_COL_INI + jj
            letra = get_column_letter(col)
            fat = f"($B{r}*{P['receita_por_aluno']}+{P['personal_mes']})"
            cel = ws.cell(
                row=r, column=col,
                value=(
                    f"=IF({fat}>0,({fat}*{P['k_ebitda']}"
                    f"-({P['custo_fixo_total']}+{letra}${_SENS_HDR_ALUGUEL_ROW}))/{fat},0)"
                ),
            )
            cel.number_format = _FMT_PCT2
            cel.font = _body_font()
            cel.alignment = Alignment(horizontal="center", vertical="center")
            cel.border = _BORDA_FINA

    ini_c = get_column_letter(_SENS_GRID_COL_INI)
    fim_c = get_column_letter(_SENS_GRID_COL_INI + len(_SENS_FATORES_ALUGUEL) - 1)
    faixa = (
        f"{ini_c}{_SENS_GRID_ROW_INI}:{fim_c}"
        f"{_SENS_GRID_ROW_INI + len(_SENS_FATORES_ALUNOS) - 1}"
    )
    ws.conditional_formatting.add(
        faixa,
        CellIsRule(operator="greaterThanOrEqual", formula=["0.1"], fill=_fill(_VERDE_CLR)),
    )
    ws.conditional_formatting.add(
        faixa, CellIsRule(operator="lessThan", formula=["0"], fill=_fill(_VERMELHO_CLR))
    )
    ws.conditional_formatting.add(
        faixa,
        CellIsRule(operator="between", formula=["0", "0.0999999"], fill=_fill(_AMARELO_CLR)),
    )

    _set_col_width(ws, 1, 16)
    _set_col_width(ws, 2, 16)
    for jj in range(len(_SENS_FATORES_ALUGUEL)):
        _set_col_width(ws, _SENS_GRID_COL_INI + jj, 16)
    ws.freeze_panes = ws.cell(row=_SENS_GRID_ROW_INI, column=_SENS_GRID_COL_INI).coordinate

    centro_r = _SENS_GRID_ROW_INI + _SENS_FATORES_ALUNOS.index(1.00)
    centro_c = get_column_letter(_SENS_GRID_COL_INI + _SENS_FATORES_ALUGUEL.index(1.00))
    return f"{ABA_SENSIBILIDADE}!${centro_c}${centro_r}"


# ---------------------------------------------------------------------------
# Aba Afericao — a aba que defende o arquivo
# ---------------------------------------------------------------------------

_AFER_ROW_INI = 6
_AFER_COLS = [
    "Indicador", "Valor do MOTOR (Python)", "Valor da FÓRMULA (planilha)", "Delta", "Status",
]


def _pares_afericao(
    r: ViabilidadeResult, refs: dict[str, str], rrefs: dict[str, str], st_dre: str,
    centro_sens: str, meses: list[int], p: Premissas, demanda_total: float,
) -> list[tuple[str, Any, str, str | None]]:
    """(rotulo, valor do motor, referencia da formula, formato).

    Os pares por MES (folha do mes 1, EBITDA dos meses 1/4/8, investimento do M-4) sao
    o que pega as duas mudancas de produto desta rodada: uma folha que voltasse a
    escalar com a rampa, ou uma franquia que voltasse a sair inteira no M-4, batem aqui
    mesmo que o mes de steady continue igual — no steady as duas mudancas sao invisiveis.
    """
    serie = {int(row["mes"]): row for row in r.serie_mensal}
    st = serie.get(int(r.mes_referencia_steady), {})
    m1 = serie.get(1, {})

    def d(key: str) -> str:
        return f"'{ABA_DRE}'!{st_dre}{_DRE_ROW[key]}"

    def col_mes(mes: int) -> str:
        return get_column_letter(_MES_COL_INI + meses.index(mes)) if mes in meses else "B"

    def dre_mes(key: str, mes: int) -> str:
        return f"'{ABA_DRE}'!{col_mes(mes)}{_DRE_ROW[key]}"

    def flx_mes(key: str, mes: int) -> str:
        return f"'{ABA_FLUXO}'!{col_mes(mes)}{_FLX_ROW[key]}"

    brl, pct, al, num = _FMT_CONTABIL, _FMT_PCT3, _FMT_ALUNOS, _FMT_INT
    pares: list[tuple[str, Any, str, str | None]] = [
        ("Mês de referência do steady", int(r.mes_referencia_steady), rrefs["mes_ref"], num),
        ("Alunos totais no mês de referência", st.get("alunos_total", 0.0),
         rrefs["alunos_ref"], al),
        ("Faturamento bruto (steady)", r.faturamento_mensal_steady, rrefs["faturamento"], brl),
        ("  dos quais receita de anuidade", r.receita_anuidade_mensal,
         rrefs["receita_anuidade"], brl),
        ("Deduções (devoluções)", st.get("deducoes", 0.0), d("deducoes"), brl),
        ("Receita líquida", r.receita_liquida, rrefs["receita_liquida"], brl),
        ("Impostos sobre receita", st.get("impostos", 0.0), d("impostos"), brl),
        ("Receita pós-impostos", r.receita_pos_impostos, rrefs["receita_pos_impostos"], brl),
        ("Custo variável", r.custos_variaveis_mensal, d("cvar_total"), brl),
        ("Folha", r.folha_mensal, d("folha"), brl),
        ("Faturamento MADURO (base de dimensionamento da folha)",
         p.faturamento_maduro(float(demanda_total)), refs["fat_maduro"], brl),
        ("Folha FIXA dimensionada pelo faturamento maduro",
         p.folha_fixa_mes(float(demanda_total)), refs["folha_fixa"], brl),
        ("Folha no mês 1 (a folha NÃO acompanha a rampa)", m1.get("folha", 0.0),
         dre_mes("folha", 1), brl),
        ("Outros custos fixos", st.get("outros_fixos", 0.0), d("outros_total"), brl),
        ("Aluguel", st.get("aluguel", 0.0), d("aluguel"), brl),
        ("Custo fixo total sem aluguel (outros fixos + folha)",
         p.custo_fixo_total_mes(float(demanda_total)), refs["custo_fixo_total"], brl),
        ("Fator receita -> EBITDA (k), sem a folha", p.fator_receita_para_ebitda,
         refs["k_ebitda"], "0.000000"),
        ("Custos operacionais totais", r.custos_op_mensal, rrefs["custos_op"], brl),
        ("EBITDA", r.ebitda_mensal, rrefs["ebitda"], brl),
        ("Margem EBITDA", r.margem_ebitda_pct, rrefs["margem"], pct),
        ("Margem EBITDA pelo centro da grade de sensibilidade", r.margem_ebitda_pct,
         centro_sens, pct),
        ("IR/CSLL", r.ir_csll_mensal, rrefs["ir_csll"], brl),
        ("Despesa financeira (juros do mês de referência)", r.despesa_financeira_mensal,
         d("juros"), brl),
        ("Resultado desalavancado após IR", r.resultado_apos_ir_mensal,
         rrefs["resultado_desalav"], brl),
        ("PMT do financiamento", r.pmt_mensal, rrefs["pmt"], brl),
        ("Juros totais do financiamento", r.juros_totais, rrefs["juros_totais"], brl),
        ("CAPEX total", r.capex_total, rrefs["capex_total"], brl),
        ("Investimento total", r.investimento_total, rrefs["investimento_total"], brl),
        ("Break-even de EBITDA (alunos totais)", r.alunos_break_even_total,
         rrefs["break_even_ebitda"], al),
        ("Break-even de caixa (alunos totais)", r.alunos_break_even_caixa_total,
         rrefs["break_even_caixa"], al),
        ("Payback (meses)", r.payback_meses, rrefs["payback"], num),
        ("TIR mensal", r.tir_mensal, rrefs["tir_mensal"], pct),
        ("TIR anual", r.tir_anual, rrefs["tir_anual"], pct),
        ("VPL na taxa de desconto", r.vpl, rrefs["vpl"], brl),
        ("FCF acumulado no fim do horizonte", r.acumulado_mes_final,
         rrefs["acumulado_m60"], brl),
        ("Retorno anual desalavancado", r.retorno_anual_desalavancado,
         rrefs["retorno_desalav"], pct),
        ("Retorno anual do equity", r.retorno_anual_equity, rrefs["retorno_equity"], pct),
        ("Ticket blended", r.ticket_blended, rrefs["ticket_blended"], brl),
        ("Aluguel-teto — faixa ideal", r.aluguel_teto.get("ideal", 0.0),
         rrefs["teto_ideal"], brl),
        ("Aluguel-teto — TETO (canônico)", r.aluguel_teto.get("teto", 0.0),
         rrefs["teto_teto"], brl),
        ("Aluguel-teto — exceção", r.aluguel_teto.get("excecao", 0.0),
         rrefs["teto_excecao"], brl),
        ("EBITDA do mês 1 (negativo por construção)", m1.get("ebitda_mensal", 0.0),
         rrefs["ebitda_m1"], brl),
        ("EBITDA do mês 4 (meio da rampa)", serie.get(4, {}).get("ebitda_mensal", 0.0),
         dre_mes("ebitda", 4), brl),
        ("EBITDA do mês 8 (fim da rampa)", serie.get(8, {}).get("ebitda_mensal", 0.0),
         dre_mes("ebitda", 8), brl),
        # Franquia PARCELADA: o motor nao expoe a parcela isolada na serie (e nada aqui
        # pode re-derivar formula financeira fora do simulador), entao a afericao usa o
        # investimento TOTAL de cada mes, que a serie expoe. Isso ja e suficiente: com a
        # taxa saindo inteira no M-4 o par do M-4 daria R$ 310 mil e o do M-1 R$ 150 mil.
        ("Investimento do M-4 (obra + parcela da franquia)",
         serie.get(-4, {}).get("investimento", 0.0), flx_mes("investimento", -4), brl),
        ("Investimento do M-1 (4a parcela da franquia ainda dentro da obra)",
         serie.get(-1, {}).get("investimento", 0.0), flx_mes("investimento", -1), brl),
        ("Investimento no mês 1 (as parcelas não vazam da pré-abertura)",
         serie.get(1, {}).get("investimento", 0.0), flx_mes("investimento", 1), brl),
    ]
    return pares


def _write_aba_afericao(
    wb: openpyxl.Workbook, r: ViabilidadeResult, refs: dict[str, str],
    rrefs: dict[str, str], st_dre: str, centro_sens: str, meses: list[int],
    p: Premissas, demanda_total: float,
) -> None:
    ws = wb.create_sheet(ABA_AFERICAO)
    _write_header(ws, "AFERIÇÃO — a planilha confere com o motor do sistema?",
                  n_cols=len(_AFER_COLS))
    _nota(
        ws, 2,
        "Para que serve: a coluna B é o número que o MOTOR em Python calculou no momento em "
        "que este arquivo foi gerado (valor ESTÁTICO, escrito por nós). A coluna C é o que as "
        "FÓRMULAS desta planilha produzem agora.",
        len(_AFER_COLS),
    )
    _nota(
        ws, 3,
        "Se todos os status estão OK, a planilha reproduz o sistema ao centavo e você pode "
        "defendê-la. Um DIVERGENTE significa que alguma fórmula quebrou OU que alguma "
        "premissa foi editada (aí a divergência é esperada e proposital).",
        len(_AFER_COLS), alerta=True,
    )
    _cabecalho_tabela(ws, 5, _AFER_COLS)

    pares = _pares_afericao(
        r, refs, rrefs, st_dre, centro_sens, meses, p, demanda_total
    )
    for i, (label, motor, ref_formula, fmt) in enumerate(pares):
        row = _AFER_ROW_INI + i
        _linha_label(ws, row, label)

        c_motor = ws.cell(row=row, column=2)
        if _finito(motor):
            c_motor.value = _sanitizar(motor)
        else:
            # payback/TIR podem ser inf/None no motor — nunca escrever inf num arquivo.
            c_motor.value = "Não atingido" if motor is not None else "n/d"
        c_motor.number_format = fmt or "General"
        c_motor.fill = _fill(_CINZA_TRAVADO)
        c_motor.font = Font(name=_FONTE_PADRAO, color=_CINZA_ESC, size=9)
        c_motor.alignment = Alignment(horizontal="right", vertical="center")
        c_motor.border = _BORDA_FINA

        c_form = ws.cell(row=row, column=3, value=f"={ref_formula}")
        c_form.number_format = fmt or "General"
        c_form.fill = _fill(_CINZA_CLR)
        c_form.font = Font(name=_FONTE_PADRAO, italic=True, color=_CINZA_ESC, size=9)
        c_form.alignment = Alignment(horizontal="right", vertical="center")
        c_form.border = _BORDA_FINA

        c_delta = ws.cell(row=row, column=4, value=f'=IFERROR(C{row}-B{row},"n/a")')
        c_delta.number_format = "0.0000"
        c_delta.font = Font(name=_FONTE_PADRAO, color=_CINZA_ESC, size=9)
        c_delta.alignment = Alignment(horizontal="right", vertical="center")
        c_delta.border = _BORDA_FINA

        c_st = ws.cell(
            row=row, column=5,
            value=f'=IF(ISNUMBER(D{row}),IF(ABS(D{row})<0.01,"OK","DIVERGENTE"),"n/a")',
        )
        c_st.font = Font(name=_FONTE_PADRAO, bold=True, color=_CINZA_ESC, size=9)
        c_st.alignment = Alignment(horizontal="center", vertical="center")
        c_st.border = _BORDA_FINA

    faixa = f"E{_AFER_ROW_INI}:E{_AFER_ROW_INI + len(pares) - 1}"
    ws.conditional_formatting.add(
        faixa, CellIsRule(operator="equal", formula=['"OK"'], fill=_fill(_VERDE_CLR))
    )
    ws.conditional_formatting.add(
        faixa, CellIsRule(operator="equal", formula=['"DIVERGENTE"'], fill=_fill(_VERMELHO_CLR))
    )

    _set_col_width(ws, 1, 52)
    _set_col_width(ws, 2, 24)
    _set_col_width(ws, 3, 24)
    _set_col_width(ws, 4, 16)
    _set_col_width(ws, 5, 16)
    ws.freeze_panes = f"A{_AFER_ROW_INI}"


# ---------------------------------------------------------------------------
# Nomes definidos (para as premissas serem chamaveis por nome no Excel)
# ---------------------------------------------------------------------------

_NOMES_DEFINIDOS = (
    "demanda", "ticket_cheio", "ticket_agregador", "ticket_blended", "share", "churn",
    "inadimplencia", "aluguel", "folha_pct", "fat_maduro", "folha_fixa", "outros_total",
    "k_ebitda", "custo_fixo_total",
    "receita_por_aluno", "personal_mes", "obra", "equip", "taxa_franquia",
    "parcelas_franquia", "franquia_parcela", "pmt",
    "investimento_total", "aporte_inicial", "mes_steady",
    "maturacao", "horizonte", "meses_pre",
)


def _registrar_nomes(wb: openpyxl.Workbook, refs: dict[str, str]) -> None:
    """Nomes ASCII (identificadores — §2: nunca acentuar identificador)."""
    for nome in _NOMES_DEFINIDOS:
        ref = refs.get(nome)
        if ref:
            wb.defined_names.add(DefinedName(f"prem_{nome}", attr_text=ref))


# ---------------------------------------------------------------------------
# API publica
# ---------------------------------------------------------------------------


_INVEST_KEYS = (
    "obra", "parcelas_obra", "equipamentos", "prazo_equipamentos",
    "juros_equipamentos_am", "taxa_franquia", "parcelas_franquia",
)


def gerar_simulador_xlsx(
    demanda_total: float,
    premissas: Premissas,
    investimento: Mapping[str, Any] | None = None,
    *,
    obra: float = 0.0,
    parcelas_obra: int = SIM_PARCELAS_OBRA_DEFAULT,
    equipamentos: float = 0.0,
    prazo_equipamentos: int = 0,
    juros_equipamentos_am: float = 0.0,
    taxa_franquia: float = SIM_TAXA_FRANQUIA,
    parcelas_franquia: int = SIM_PARCELAS_FRANQUIA_DEFAULT,
    nome_ponto: str = "",
    rotulo: str | None = None,
    m2: float | None = None,
    resultado: ViabilidadeResult | None = None,
) -> bytes:
    """Gera o simulador financeiro completo em .xlsx com FORMULAS VIVAS.

    8 abas, na ordem: Premissas, Folha, DRE mensal, Fluxo de caixa, Investimento,
    Resumo, Sensibilidade, Afericao.

    Parameters
    ----------
    demanda_total:
        Alunos TOTAIS na maturidade (nunca derivado de lat/lng — DEC-009).
    premissas:
        Objeto `Premissas` do motor. TODO default da aba Premissas sai daqui.
    investimento:
        Dict de investimento (as MESMAS chaves de `simulador.simular()`), aceito
        POSICIONAL porque e assim que a rota do piloto web chama. O que vier aqui
        sobrescreve os kwargs equivalentes; chaves desconhecidas sao ignoradas.
    obra, parcelas_obra, equipamentos, prazo_equipamentos, juros_equipamentos_am,
    taxa_franquia, parcelas_franquia:
        Mesma semantica de `simulador.simular()`, para uso direto sem o dict. A taxa
        de franquia e PARCELADA sem juros (default 4x): as parcelas caem nos meses de
        contrato 1..N (M-4..M-1 com N=4), junto da obra.
    nome_ponto, rotulo:
        Nome do imovel/ponto exibido nos cabecalhos. `rotulo` e o nome que a rota
        do piloto web usa e tem precedencia quando os dois vem.
    m2:
        Metragem, so para o cabecalho (nao entra em nenhuma conta).
    resultado:
        Resultado de `simular()` ja calculado. Quando None, e calculado aqui — e a
        FONTE dos valores estaticos da aba Afericao.

    Returns
    -------
    bytes
        Conteudo binario do .xlsx. NUNCA escreve em disco.
    """
    inv: dict[str, Any] = {
        "obra": obra,
        "parcelas_obra": parcelas_obra,
        "equipamentos": equipamentos,
        "prazo_equipamentos": prazo_equipamentos,
        "juros_equipamentos_am": juros_equipamentos_am,
        "taxa_franquia": taxa_franquia,
        "parcelas_franquia": parcelas_franquia,
    }
    if investimento:
        inv.update({k: v for k, v in investimento.items() if k in _INVEST_KEYS})
    obra = float(inv["obra"])
    parcelas_obra = int(inv["parcelas_obra"])
    equipamentos = float(inv["equipamentos"])
    prazo_equipamentos = int(inv["prazo_equipamentos"])
    juros_equipamentos_am = float(inv["juros_equipamentos_am"])
    taxa_franquia = float(inv["taxa_franquia"])
    parcelas_franquia = int(inv["parcelas_franquia"])

    titulo_ponto = (rotulo or nome_ponto or "").strip()
    if m2:
        area = f"{_num_br(float(m2), 0)} m²"
        titulo_ponto = f"{titulo_ponto} ({area})".strip() if titulo_ponto else area

    r = resultado or simular(
        demanda_total, premissas, obra=obra, parcelas_obra=parcelas_obra,
        equipamentos=equipamentos, prazo_equipamentos=prazo_equipamentos,
        juros_equipamentos_am=juros_equipamentos_am, taxa_franquia=taxa_franquia,
        parcelas_franquia=parcelas_franquia,
    )
    meses = _meses_da_serie(premissas)

    # Colunas especiais: calculadas ANTES de criar as abas, porque Folha e Resumo
    # referenciam a coluna de steady da DRE e do Fluxo.
    letra_st_dre = get_column_letter(_MES_COL_INI + len(meses))
    col_total_flx = _MES_COL_INI + len(meses)
    letra_st_flx = get_column_letter(col_total_flx + 1)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove a aba padrao vazia

    blocos = _blocos_premissas(
        demanda_total, premissas, obra=obra, parcelas_obra=parcelas_obra,
        equipamentos=equipamentos, prazo_equipamentos=prazo_equipamentos,
        juros_equipamentos_am=juros_equipamentos_am, taxa_franquia=taxa_franquia,
        parcelas_franquia=parcelas_franquia,
    )
    refs = _write_aba_premissas(wb, blocos, titulo_ponto)
    _write_aba_folha(wb, premissas, refs, letra_st_dre)
    _write_aba_dre(wb, meses, refs)
    _write_aba_fluxo(wb, meses, refs)
    _write_aba_investimento(wb, meses, refs, col_total_flx)
    rrefs = _write_aba_resumo(wb, meses, refs, letra_st_dre, letra_st_flx, titulo_ponto)
    centro_sens = _write_aba_sensibilidade(wb, refs)
    _write_aba_afericao(
        wb, r, refs, rrefs, letra_st_dre, centro_sens, meses, premissas, demanda_total
    )

    _registrar_nomes(wb, refs)
    wb.calculation.fullCalcOnLoad = True

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["ABAS_ESPERADAS", "gerar_simulador_xlsx"]
