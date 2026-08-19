"""Faturamento oficial da rede, lido da planilha do Financeiro (base dos royalties).

Por que este módulo existe
--------------------------
A base Growth (`growth_api_historico.parquet`) subdimensiona o faturamento da rede. Medido
contra a planilha do Financeiro em 2026-08-12, competência 2026-07 fechada:

* a rede fecha **R$ 23,05 mi** na planilha contra **R$ 18,68 mi** na Growth — 19% a menos;
* a **receita de agregador morreu na Growth em maio/2025**: de 2025-05 em diante,
  `faturamento - faturamento_sem_agregador` é exatamente 0 em **100% das linhas**, embora
  passagens e alunos Gympass continuem chegando (3,65 mi de passagens em 2026-07). Ou
  seja, `faturamento_agregador` e `pct_agregador_receita` exibem zero há 15 meses. O valor
  real de 2026-07 é R$ 5,11 mi (Gympass 3,60 + Totalpass 1,67 - Tem Saúde 0,16), 23% da
  receita;
* a coluna `faturamento` da Growth mistura DOIS conceitos entre linhas: em algumas
  unidades ela bate com `VENDAS UX` da planilha (sem agregador) e em outras com o TOTAL.
  Praia Grande bate com `VENDAS UX` até 2026-05 e salta para o total em 2026-06.

A planilha, ao contrário, é internamente consistente: `TOTAL = VENDAS UX + GYMPASS +
TOTALPASS - (-) TEM SAÚDE` fecha em **100% das 6.432 células** com diferença máxima de
R$ 0,00. É a base sobre a qual os royalties são cobrados.

O que este módulo NÃO resolve
-----------------------------
A planilha é MENSAL. O quarteto do topo da Visão Executiva compara períodos parciais dia a
dia (01-10/08 contra 01-10/07) e continua vindo da Growth. Só a camada de meses FECHADOS
troca de fonte — ver `rede_metricas.aplicar_faturamento_financeiro`.

Três unidades existem na planilha e **não existem na Growth** (medido em 2026-07): São
Carlos-SP (R$ 183 mil), Jardim das Américas-MT (R$ 182 mil) e Vila Izabel-PR (R$ 158 mil).
A sobreposição é um OVERRIDE, nunca um INSERT: sem série de alunos, churn e NPS, uma linha
nova na carteira seria uma unidade pela metade. Elas saem em `unidades_sem_par` para o
operador cobrar a inclusão na Growth.

Estabilidade da fonte
---------------------
Meses fechados não são reescritos. A própria planilha carrega um snapshot antigo na aba
`Faturamento & Alunos2` (até 2025-09): das 1.083 células comparáveis, **98,25% são
idênticas**, e as 19 divergências estão TODAS na última competência daquele snapshot — que
ainda estava sendo preenchida (0,00 -> valor real). Daí a regra do `validar`: a última
coluna só vale depois que o mês virou.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd

#: Aba com a série mensal por unidade. A planilha carrega snapshots antigos em abas de
#: mesmo prefixo (`... 2`), que NÃO devem ser lidos — daí o nome exato.
ABA_FATURAMENTO = "Faturamento & Alunos"

#: De-para oficial `COD_UNIDADE -> UNIDADE_UX`, mantido pelo próprio Financeiro. É o que
#: torna o join confiável: casar por nome falha em massa nesta rede.
ABA_DEPARA = "Unidades_UX"

#: Rótulos das sub-linhas de cada bloco de unidade, já normalizados (a planilha tem
#: "GYMPASS " com espaço à direita em um dos layouts).
_COMPONENTES: dict[str, str] = {
    "VENDAS UX": "vendas_ux",
    "GYMPASS": "gympass",
    "TOTALPASS": "totalpass",
    "(-) TEM SAUDE": "tem_saude",
    "ALUNOS": "alunos",
}

#: Prefixo do cabeçalho de bloco. O layout ATUAL usa o código da unidade ("01 - AUGUSTA");
#: o snapshot antigo usava a UF ("SP - AUGUSTA"). Capturamos os dois e só tratamos como
#: código o que existir no de-para — os códigos vão de "01" a "B0" e nenhuma UF começa
#: com dígito, então não há colisão.
_CABECALHO_RE = re.compile(r"^\s*(?:(?P<prefixo>[0-9A-Z]{2})\s*-\s*)?(?P<nome>.+?)\s*$")

#: Célula "não existia neste mês". Vira NA, e não zero: zero é faturamento nulo.
_VAZIOS = {"", "-", "--"}

#: Tolerância da checagem de fechamento. Um centavo — a planilha fecha em R$ 0,00 exatos,
#: e a folga existe só para não quebrar em ruído de ponto flutuante do próprio Excel.
TOLERANCIA_CENTAVOS = 0.01

COLUNAS: tuple[str, ...] = (
    "cod_unidade",
    "unidade_planilha",
    "unidade_ux",
    "tem_depara",
    "competencia",
    "faturamento",
    "vendas_ux",
    "gympass",
    "totalpass",
    "tem_saude",
)


@dataclass(frozen=True)
class Achado:
    """Resultado de uma checagem. `erro` barra a ingestão; `aviso` só informa."""

    nivel: str
    codigo: str
    mensagem: str

    @property
    def eh_erro(self) -> bool:
        return self.nivel == "erro"

    def __str__(self) -> str:
        return f"[{self.nivel.upper()}] {self.codigo}: {self.mensagem}"


def _texto(valor: object) -> str:
    """MAIÚSCULAS, sem acento, espaços colapsados."""
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    texto = unicodedata.normalize("NFKD", str(valor))
    texto = "".join(c for c in texto if not unicodedata.combining(c)).upper().strip()
    return " ".join(texto.split())


def _numero(valor: object) -> float | None:
    """Célula -> float. Texto de preenchimento e vazio viram None."""
    if isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)) and not pd.isna(valor):
        return float(valor)
    if isinstance(valor, str) and valor.strip() in _VAZIOS:
        return None
    return None


def _linha_de_meses(linhas: list[tuple[Any, ...]], ate: int = 12) -> tuple[int, dict[int, str]]:
    """Acha a linha de cabeçalho pela DENSIDADE de datas, não por número fixo.

    O layout já mudou uma vez dentro do mesmo arquivo (cabeçalho na linha 4 numa aba e na
    6 na outra). Procurar a linha com mais células de data sobrevive à próxima mudança.
    """
    melhor_i: int = -1
    melhor_meses: dict[int, str] = {}
    for i, linha in enumerate(linhas[:ate]):
        meses = {
            j: pd.Timestamp(v).strftime("%Y-%m")
            for j, v in enumerate(linha)
            if hasattr(v, "year") and hasattr(v, "month") and not isinstance(v, str)
        }
        if len(meses) > len(melhor_meses):
            melhor_i, melhor_meses = i, meses
    return melhor_i, melhor_meses


def _ler_depara(wb: Any) -> dict[str, str]:
    """`COD_UNIDADE -> UNIDADE_UX`. Vazio se a aba não existir (a leitura degrada)."""
    if ABA_DEPARA not in wb.sheetnames:
        return {}
    depara: dict[str, str] = {}
    for linha in wb[ABA_DEPARA].iter_rows(min_row=2, values_only=True):
        if len(linha) < 3 or linha[1] is None or linha[2] is None:
            continue
        depara[_cod(linha[1])] = str(linha[2]).strip()
    return depara


def _cod(valor: object) -> str:
    """Código de unidade normalizado.

    O Excel entrega "01" como texto numa aba e pode entregar 1 como número noutra; sem
    aparar o zero à esquerda, o de-para erra silenciosamente e a unidade fica sem par.
    Códigos alfanuméricos ("A0", "B0") passam intactos.
    """
    texto = _texto(valor)
    return texto.lstrip("0") or "0" if texto.isdigit() else texto


def ler_planilha(caminho: Path | str, aba: str = ABA_FATURAMENTO) -> pd.DataFrame:
    """Lê a planilha do Financeiro para formato longo (uma linha por unidade-competência).

    Abre em `read_only` e `data_only`: nunca escreve no arquivo e lê o valor CALCULADO das
    fórmulas (o Excel grava o resultado em cache; sem `data_only` viria "=SOMA(...)").
    """
    import openpyxl

    caminho = Path(caminho)
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    try:
        if aba not in wb.sheetnames:
            raise ValueError(f"aba {aba!r} não existe em {caminho.name} (abas: {wb.sheetnames})")
        depara = _ler_depara(wb)
        linhas = list(wb[aba].iter_rows(values_only=True))
    finally:
        wb.close()

    _, meses = _linha_de_meses(linhas)
    if not meses:
        raise ValueError(f"não achei a linha de competências na aba {aba!r} de {caminho.name}")

    registros: list[dict[str, Any]] = []
    i = 0
    while i < len(linhas):
        linha = linhas[i]
        indice = linha[0] if linha else None
        rotulo = linha[1] if len(linha) > 1 else None
        # Cabeçalho de bloco = índice numérico na coluna A + nome na B. O bloco
        # "TOTAL GERAL" não tem índice e por isso fica de fora sozinho.
        if not (indice is not None and str(indice).strip().isdigit() and isinstance(rotulo, str) and rotulo.strip()):
            i += 1
            continue

        # `_CABECALHO_RE` casa qualquer texto não vazio (o `.+?` do grupo `nome`), e o
        # `rotulo.strip()` acima já é não vazio — mas o tipo diz `Match | None`, e um dia
        # alguém mexe no padrão. Sem par, o bloco é pulado em vez de derrubar a leitura.
        casado = _CABECALHO_RE.match(rotulo.strip())
        if casado is None:
            i += 1
            continue
        prefixo, nome = casado.group("prefixo"), casado.group("nome").strip()
        # Três casos, nesta ordem: código conhecido -> usa o de-para; UF (layout antigo)
        # -> descarta o prefixo; qualquer outro -> é código novo, ainda sem linha no
        # de-para, e o nome do bloco vira o nome de join.
        if prefixo and _cod(prefixo) in depara:
            cod = prefixo
        elif prefixo and _eh_uf(prefixo):
            cod = None
        else:
            cod = prefixo or None

        total = {j: _numero(linha[j]) if j < len(linha) else None for j in meses}
        componentes: dict[str, dict[int, float | None]] = {}
        j = i + 1
        while j < len(linhas):
            sub = linhas[j][1] if len(linhas[j]) > 1 else None
            chave = _COMPONENTES.get(_texto(sub))
            if chave is None:
                break
            componentes[chave] = {
                k: _numero(linhas[j][k]) if k < len(linhas[j]) else None for k in meses
            }
            j += 1

        tem_depara = bool(cod) and _cod(cod) in depara
        unidade_ux = depara[_cod(cod)] if tem_depara else nome
        for k, mes in meses.items():
            registros.append(
                {
                    "cod_unidade": cod,
                    "unidade_planilha": nome,
                    "unidade_ux": unidade_ux,
                    "tem_depara": tem_depara,
                    "competencia": mes,
                    "faturamento": total.get(k),
                    "vendas_ux": componentes.get("vendas_ux", {}).get(k),
                    "gympass": componentes.get("gympass", {}).get(k),
                    "totalpass": componentes.get("totalpass", {}).get(k),
                    "tem_saude": componentes.get("tem_saude", {}).get(k),
                }
            )
        i = j

    if not registros:
        return _vazio()
    return pd.DataFrame(registros)[list(COLUNAS)]


_TEXTOS = frozenset({"cod_unidade", "unidade_planilha", "unidade_ux", "competencia"})


def _vazio() -> pd.DataFrame:
    tipos = {c: ("object" if c in _TEXTOS else "bool" if c == "tem_depara" else "float64") for c in COLUNAS}
    return pd.DataFrame({c: pd.Series(dtype=t) for c, t in tipos.items()})


_UFS = frozenset(
    "AC AL AP AM BA CE DF ES GO MA MT MS MG PA PB PR PE PI RJ RN RS RO RR SC SP SE TO".split()
)


def _eh_uf(prefixo: str) -> bool:
    return _texto(prefixo) in _UFS


def competencia_fechada(hoje: date | None = None) -> str:
    """Última competência que já virou. É sempre o mês ANTERIOR ao corrente."""
    hoje = hoje or date.today()
    return str(pd.Period(pd.Timestamp(hoje), freq="M") - 1)


def validar(
    fat: pd.DataFrame,
    hoje: date | None = None,
    anterior: pd.DataFrame | None = None,
) -> list[Achado]:
    """Portões de qualidade. Devolve os achados; quem decide barrar é o chamador."""
    achados: list[Achado] = []
    if not len(fat):
        return [Achado("erro", "vazia", "a planilha não produziu nenhuma linha")]

    # --- E1: a aritmética interna fecha? -----------------------------------------------
    com_total = fat[fat["faturamento"].notna()]

    # `to_numeric` antes do `fillna`: uma componente inteiramente vazia chega como coluna
    # de objeto, e somar objeto com float dispara downcast silencioso no pandas.
    def _zerado(coluna: str) -> pd.Series:
        return pd.to_numeric(com_total[coluna], errors="coerce").fillna(0.0)

    partes = _zerado("vendas_ux") + _zerado("gympass") + _zerado("totalpass") - _zerado("tem_saude")
    desvio = (pd.to_numeric(com_total["faturamento"], errors="coerce") - partes).abs()
    fora = desvio > TOLERANCIA_CENTAVOS
    if fora.any():
        pior = desvio.max()
        achados.append(
            Achado(
                "erro",
                "aritmetica",
                f"{int(fora.sum())} de {len(com_total)} células não fecham "
                f"TOTAL = VENDAS UX + GYMPASS + TOTALPASS - TEM SAÚDE "
                f"(maior diferença R$ {pior:,.2f})",
            )
        )

    meses = sorted(fat["competencia"].dropna().unique())
    ultima = meses[-1]

    # --- E2: a última competência já fechou? -------------------------------------------
    # O snapshot antigo embutido na planilha prova o risco: a última coluna dele tinha 19
    # unidades zeradas que depois viraram valor real. Ler cedo demais grava um mês pela
    # metade como se fosse definitivo.
    esperada = competencia_fechada(hoje)
    if ultima > esperada:
        achados.append(
            Achado(
                "erro",
                "mes_aberto",
                f"a última competência da planilha é {ultima}, que ainda NÃO fechou "
                f"(o último mês fechado é {esperada})",
            )
        )
    elif ultima < esperada:
        achados.append(
            Achado(
                "aviso",
                "planilha_atrasada",
                f"a planilha vai até {ultima}, mas {esperada} já fechou — cópia velha?",
            )
        )

    # --- E3: a série é contígua? -------------------------------------------------------
    faixa = pd.period_range(meses[0], ultima, freq="M")
    buracos = sorted(set(str(p) for p in faixa) - set(meses))
    if buracos:
        achados.append(
            Achado("erro", "buraco", f"{len(buracos)} competência(s) faltando: {buracos[:6]}")
        )

    # --- E4/A2: a última competência veio preenchida? ----------------------------------
    por_mes = fat.assign(positivo=fat["faturamento"].fillna(0) > 0).groupby("competencia")["positivo"].sum()
    if len(por_mes) >= 2:
        n_ultimo, n_penultimo = int(por_mes.loc[ultima]), int(por_mes.iloc[-2])
        if n_ultimo == 0:
            achados.append(Achado("erro", "ultima_vazia", f"nenhuma unidade com faturamento em {ultima}"))
        elif n_ultimo < n_penultimo * 0.9:
            achados.append(
                Achado(
                    "aviso",
                    "queda_cobertura",
                    f"{ultima} tem {n_ultimo} unidades com faturamento contra "
                    f"{n_penultimo} no mês anterior — mês possivelmente incompleto",
                )
            )

    # --- A3: unidade sem de-para -------------------------------------------------------
    sem = sorted(fat.loc[~fat["tem_depara"].astype(bool), "unidade_planilha"].unique())
    if sem:
        achados.append(
            Achado(
                "aviso",
                "sem_depara",
                f"{len(sem)} unidade(s) sem COD_UNIDADE na aba {ABA_DEPARA!r}: {sem[:8]}",
            )
        )

    # --- A1: mês fechado foi reescrito? ------------------------------------------------
    if anterior is not None and len(anterior):
        achados.extend(_achados_de_reescrita(fat, anterior))

    return achados


def _achados_de_reescrita(fat: pd.DataFrame, anterior: pd.DataFrame) -> list[Achado]:
    """Compara com o snapshot anterior e denuncia mudança em mês JÁ fechado.

    A última competência do snapshot anterior fica de fora da comparação de propósito: ela
    era o mês que ainda estava sendo preenchido, e mudar ali é o comportamento normal.
    """
    corte = sorted(anterior["competencia"].dropna().unique())[-1]
    chaves = ["unidade_planilha", "competencia"]
    a = anterior[anterior["competencia"] < corte][chaves + ["faturamento"]]
    b = fat[chaves + ["faturamento"]]
    junto = a.merge(b, on=chaves, how="inner", suffixes=("_antes", "_agora"))
    if not len(junto):
        return []
    dif = (junto["faturamento_agora"].fillna(0) - junto["faturamento_antes"].fillna(0)).abs()
    mudou = junto[dif > TOLERANCIA_CENTAVOS]
    if not len(mudou):
        return []
    return [
        Achado(
            "aviso",
            "reescrita",
            f"{len(mudou)} célula(s) de competência JÁ fechada mudaram de valor "
            f"(ex.: {mudou.iloc[0]['unidade_planilha']} em {mudou.iloc[0]['competencia']})",
        )
    ]


def carregar(parquet: Path | str) -> pd.DataFrame:
    """Lê o parquet gerado pela ingestão. DataFrame vazio se não existir."""
    caminho = Path(parquet)
    if not caminho.exists():
        return pd.DataFrame()
    fat = pd.read_parquet(caminho)
    faltando = [c for c in COLUNAS if c not in fat.columns]
    if faltando:
        raise ValueError(f"{caminho.name} sem as colunas {faltando}")
    return fat
