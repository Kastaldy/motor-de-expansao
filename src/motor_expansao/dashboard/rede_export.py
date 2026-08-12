"""Exports da Visao Executiva: CSV, XLSX e PDF - BLK-EXEC-10/11.

READ-ONLY: nada aqui escreve em disco. Todo gerador devolve **bytes** e quem chama decide o
que fazer com eles -- e' o que permite ao backend do piloto continuar passando no guardrail
AST (`test_modulos_de_rede_sao_read_only_por_ast`).

Duas escolhas que parecem detalhe e nao sao:

* **`csv.writer` sobre `StringIO`, nunca `DataFrame.to_csv`.** O guardrail AST reprova
  `to_csv` por nome; e o CSV da casa e' `sep=";"` com `utf-8-sig` (regra do `CLAUDE.md` §2),
  que e' o que o Excel em pt-BR abre sem pedir nada.
* **Numero com virgula decimal**, tambem por causa do Excel pt-BR: com ponto, a planilha
  importa faturamento como texto e o time perde a soma.

A entrada e' o MESMO payload que a tela recebe. Nao ha uma segunda fonte de verdade: se a
carteira e o CSV divergirem, e' porque alguem calculou duas vezes -- e isso e' justamente o
defeito mais caro deste projeto.
"""

from __future__ import annotations

import csv
from collections.abc import Mapping, Sequence
from io import BytesIO, StringIO
from typing import Any

from motor_expansao.dashboard.pdf_base import (
    BRANCO,
    CINZA_CLARO,
    CINZA_TEXTO,
    COR_SEVERIDADE,
    PAGINA_ALTURA,
    PAGINA_LARGURA,
    ULTRA_MAGENTA,
    ULTRA_TURQUESA,
    UltraPDF,
    ascii_seguro,
    barra_de_meta,
    barra_empilhada,
    barras,
    barras_horizontais,
    cartao,
    faixa_de_percentil,
    faixa_de_titulo,
    linha_de_tabela,
    rodape,
    rosca,
    titulo_de_grafico,
)
from motor_expansao.dashboard.pdf_base import (
    linha as linha_de_grafico,
)

CSV_SEPARADOR = ";"
CSV_ENCODING = "utf-8-sig"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Colunas do export tabular. Ordem pensada para leitura: quem e' -> como esta -> por que.
COLUNAS_CARTEIRA: tuple[tuple[str, str], ...] = (
    ("nome", "Unidade"),
    ("uf", "UF"),
    ("cidade", "Cidade"),
    ("consultor", "Consultor"),
    ("master_franquia", "Master franquia"),
    ("master", "Master (Growth)"),
    ("coorte_rotulo", "Maturidade"),
    ("meses_operacao", "Meses de operação"),
    ("severidade_rotulo", "Diagnóstico"),
    ("alertas_texto", "Alertas"),
    ("faixa_faturamento_rotulo", "Faixa de faturamento"),
)

# Metricas exportadas com o quarteto completo (valor, M-1, ranking, % vs media).
METRICAS_CARTEIRA: tuple[tuple[str, str], ...] = (
    ("faturamento", "Faturamento"),
    ("ativos", "Alunos ativos"),
    ("pagantes", "Recorrentes"),
    ("agregadores", "Agregadores"),
    ("receita_por_recorrente", "Receita por recorrente"),
    ("churn_pct", "Churn %"),
    ("conversao_pct", "Conversão %"),
    ("nps", "NPS"),
    ("saldo_operacional", "Saldo operacional"),
    ("pct_agregador_alunos", "Dependência de agregador (alunos) %"),
)

# A FICHA leva uma linha a mais: a dependencia de agregador por RECEITA. Ela nao entra no
# export da carteira porque o payload da carteira nao a serve (sao ~110 bytes por unidade e
# a carteira ja trafega 92 linhas); na ficha, que e' uma unidade so', ela cabe -- e e' a
# leitura que fala de dinheiro. As duas discordam em 14,8 p.p. na mediana da rede.
METRICAS_FICHA: tuple[tuple[str, str], ...] = (
    *METRICAS_CARTEIRA,
    ("pct_agregador_receita", "Dependência de agregador (receita) %"),
)


# ---------------------------------------------------------------------------
# Formatacao
# ---------------------------------------------------------------------------


def _br(valor: object, casas: int = 0) -> str:
    """Numero em pt-BR (milhar com ponto, decimal com virgula). Vazio se nao houver dado."""
    if valor is None:
        return ""
    try:
        numero = float(valor)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return str(valor)
    if numero != numero:  # NaN
        return ""
    texto = f"{numero:,.{casas}f}"
    return texto.replace(",", "\x00").replace(".", ",").replace("\x00", ".")


#: Caracteres com que o Excel/LibreOffice comeca a interpretar a celula como FORMULA.
_INICIO_DE_FORMULA = ('=', '+', '-', '@', chr(9), chr(13))


def _brl_da_barra(valor: float) -> str:
    """R$ 21.982.435 -- o faturamento por INTEIRO sobre a barra.

    Era compacto ("R$ 594k", "R$ 1,2M") porque o numero cheio nao cabia sobre 12 barras.
    So' que o valor arredondado le como numero ajustado pelo sistema, e o painel hoje
    promete o contrario: o faturamento vem da planilha do Financeiro, a mesma dos royalties
    (pedido do Felipe, 2026-08-12). Quem cede e' o corpo da fonte -- `pdf_base.barras`
    encolhe o rotulo ate caber na fatia, e a tela faz o mesmo em `corpoQueCabe`.
    """
    return f"R$ {_br(valor)}"


def _celula(valor: object) -> str:
    """Valor para a celula do CSV/XLSX: numero em pt-BR, texto neutralizado."""
    if isinstance(valor, bool):
        return "Sim" if valor else "Não"
    if isinstance(valor, (int, float)):
        return _br(valor, 0 if float(valor).is_integer() else 2)
    if valor is None:
        return ""
    texto = str(valor)
    # O cadastro e' EDITAVEL pela tela e semeado de uma planilha mantida a mao. Um
    # `consultor` que comece com "=" vira formula viva ao abrir o CSV -- HYPERLINK e
    # WEBSERVICE exfiltram, DDE executa. O apostrofo a frente e' o que o proprio Excel usa
    # para dizer "isto e' texto"; ele nao aparece na celula.
    if texto[:1] in _INICIO_DE_FORMULA:
        return "'" + texto
    return texto


def _linhas_carteira(payload: Mapping[str, Any]) -> tuple[list[str], list[list[str]]]:
    """Cabecalho + linhas do export tabular, a partir do payload da carteira."""
    cabecalho = [rotulo for _, rotulo in COLUNAS_CARTEIRA]
    for _, rotulo in METRICAS_CARTEIRA:
        cabecalho += [rotulo, f"{rotulo} (M-1)", f"{rotulo} (ranking)", f"{rotulo} (% vs média)"]

    linhas: list[list[str]] = []
    for unidade in payload.get("unidades", []):
        alertas = "; ".join(a.get("titulo", "") for a in unidade.get("alertas", []))
        registro = dict(unidade)
        registro["alertas_texto"] = alertas
        linha = [_celula(registro.get(chave)) for chave, _ in COLUNAS_CARTEIRA]
        metricas = unidade.get("metricas", {})
        for chave, _ in METRICAS_CARTEIRA:
            metrica = metricas.get(chave) or {}
            rank = metrica.get("rank")
            total = metrica.get("rank_total")
            linha += [
                _celula(metrica.get("atual")),
                _celula(metrica.get("m1")),
                f"{int(rank)}/{int(total)}" if rank and total else "",
                _celula(metrica.get("vs_media_pct")),
            ]
        linhas.append(linha)
    return cabecalho, linhas


# ---------------------------------------------------------------------------
# CSV / XLSX
# ---------------------------------------------------------------------------


def carteira_csv(payload: Mapping[str, Any]) -> bytes:
    """CSV da carteira, no dialeto da casa (`;` + utf-8-sig)."""
    cabecalho, linhas = _linhas_carteira(payload)
    buffer = StringIO()
    escritor = csv.writer(buffer, delimiter=CSV_SEPARADOR, lineterminator="\r\n")
    escritor.writerow(cabecalho)
    escritor.writerows(linhas)
    return buffer.getvalue().encode(CSV_ENCODING)


def carteira_xlsx(payload: Mapping[str, Any]) -> bytes:
    """XLSX da carteira, com a aba de metodo junto.

    A segunda aba nao e' enfeite: sem ela, a planilha circula pela rede sem dizer de que
    competencia e' o numero nem qual regua acendeu cada alerta.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    cabecalho, linhas = _linhas_carteira(payload)
    livro = openpyxl.Workbook()
    aba = livro.active
    aba.title = "Carteira"
    aba.append(cabecalho)
    for linha in linhas:
        aba.append(linha)

    negrito_branco = Font(bold=True, color="FFFFFF")
    fundo = PatternFill("solid", fgColor="00A79D")
    for celula in aba[1]:
        celula.font = negrito_branco
        celula.fill = fundo
        celula.alignment = Alignment(wrap_text=True, vertical="center")
    aba.freeze_panes = "B2"
    for indice, titulo in enumerate(cabecalho, start=1):
        letra = openpyxl.utils.get_column_letter(indice)
        aba.column_dimensions[letra].width = min(max(12, len(titulo) + 2), 34)

    metodo = livro.create_sheet("Método")
    for linha in _linhas_de_metodo(payload):
        metodo.append(linha)
    metodo.column_dimensions["A"].width = 30
    metodo.column_dimensions["B"].width = 96

    saida = BytesIO()
    livro.save(saida)
    return saida.getvalue()


def _linhas_de_metodo(payload: Mapping[str, Any]) -> list[list[str]]:
    linhas = [
        ["Fonte", "Growth API (growth_api_historico.parquet) - camada paralela, read-only M1"],
        ["Competência", str(payload.get("mes", ""))],
        ["Dia de referência", str(payload.get("referencia", ""))],
        ["Comparação M-1", str(payload.get("referencia_m1", ""))],
        ["Diagnóstico calculado em", str(payload.get("competencia_diagnostico", ""))],
        ["", ""],
        ["Réguas vigentes", ""],
    ]
    for chave, regua in (payload.get("reguas") or {}).items():
        limiar = regua.get("limiar")
        sentido = {"acima": "acima de", "abaixo": "abaixo de", "persistencia": "por"}.get(
            str(regua.get("sentido")), ""
        )
        meses = regua.get("meses")
        detalhe = (
            f"{sentido} {_br(limiar, 1)} {regua.get('unidade', '')}"
            if meses is None
            else f"{sentido} {int(meses)} meses fechados seguidos"
        )
        linhas.append([str(regua.get("rotulo", chave)), detalhe.strip()])
    linhas.append(["", ""])
    for nota in payload.get("notas", []):
        linhas.append(["Nota", str(nota)])
    return linhas


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

_LINHAS_POR_PAGINA = 22


def carteira_pdf(payload: Mapping[str, Any]) -> bytes:
    """PDF da carteira: capa com os KPIs da rede + tabela paginada por prioridade."""
    pdf = UltraPDF()
    unidades: Sequence[Mapping[str, Any]] = payload.get("unidades", [])
    subtitulo = f"Competência {payload.get('mes', '')} - até {payload.get('referencia', '')}"

    pdf.add_page()
    faixa_de_titulo(pdf, "Rede Ultra - carteira", subtitulo)
    _cartoes_da_rede(pdf, payload)
    _resumo_do_semaforo(pdf, payload)
    rodape(pdf, _texto_do_rodape(payload))

    # Os MESMOS graficos que a tela mostra, lendo os MESMOS numeros do payload. Quem
    # recebe o PDF numa reuniao nao tem a tela ao lado para conferir; se as duas
    # superficies desenharem series diferentes, a divergencia so aparece na frente do
    # comite.
    _pagina_de_graficos_da_rede(pdf, payload, subtitulo)

    cabecalho = [
        (30.0, "#", "L"),
        (176.0, "Unidade", "L"),
        (34.0, "UF", "L"),
        (96.0, "Consultor", "L"),
        (86.0, "Maturidade", "L"),
        (92.0, "Faturamento", "R"),
        (62.0, "Ativos", "R"),
        (56.0, "Churn", "R"),
        (48.0, "NPS", "R"),
        (206.0, "Diagnóstico", "L"),
    ]
    for inicio in range(0, max(len(unidades), 1), _LINHAS_POR_PAGINA):
        bloco = unidades[inicio : inicio + _LINHAS_POR_PAGINA]
        pdf.add_page()
        faixa_de_titulo(pdf, "Carteira por prioridade", subtitulo, rgb=ULTRA_MAGENTA)
        y = 76.0
        linha_de_tabela(pdf, 36.0, y, cabecalho, negrito=True, fundo=CINZA_CLARO)
        y += 18.0
        for posicao, unidade in enumerate(bloco, start=inicio + 1):
            metricas = unidade.get("metricas", {})
            severidade = str(unidade.get("severidade", "ok"))
            resumo = str(unidade.get("resumo", ""))
            linha_de_tabela(
                pdf,
                36.0,
                y,
                [
                    (30.0, str(posicao), "L"),
                    (176.0, str(unidade.get("nome", ""))[:30], "L"),
                    (34.0, str(unidade.get("uf", "")), "L"),
                    (96.0, str(unidade.get("consultor") or "-")[:16], "L"),
                    (86.0, str(unidade.get("coorte_rotulo", ""))[:14], "L"),
                    (92.0, _br(_valor(metricas, "faturamento")), "R"),
                    (62.0, _br(_valor(metricas, "ativos")), "R"),
                    (56.0, _br(_valor(metricas, "churn_pct"), 1), "R"),
                    (48.0, _br(_valor(metricas, "nps"), 0), "R"),
                    (206.0, resumo[:74], "L"),
                ],
                cor=COR_SEVERIDADE.get(severidade, CINZA_TEXTO),
            )
            y += 16.0
        rodape(pdf, _texto_do_rodape(payload))

    return bytes(pdf.output())


def _valor(metricas: Mapping[str, Any], chave: str) -> object:
    return (metricas.get(chave) or {}).get("atual")


def _cartoes_da_rede(pdf: UltraPDF, payload: Mapping[str, Any]) -> None:
    kpis = payload.get("kpis", {})
    cartoes = [
        ("Faturamento no período", "R$ " + _br(_valor(kpis, "faturamento")), "acumulado até o dia"),
        ("Alunos ativos", _br(_valor(kpis, "ativos")), ""),
        ("Churn", _br(_valor(kpis, "churn_pct"), 1) + "%", "média ponderada"),
        ("Receita por recorrente", "R$ " + _br(_valor(kpis, "receita_por_recorrente"), 2), ""),
        ("NPS", _br(_valor(kpis, "nps"), 1), f"meta {int(payload.get('meta_nps', 60))}"),
    ]
    largura = (PAGINA_LARGURA - 72 - 4 * 12) / 5
    for indice, (rotulo, valor, apoio) in enumerate(cartoes):
        cartao(
            pdf,
            36.0 + indice * (largura + 12),
            84.0,
            largura,
            72.0,
            rotulo=rotulo,
            valor=valor,
            apoio=apoio,
        )


def _resumo_do_semaforo(pdf: UltraPDF, payload: Mapping[str, Any]) -> None:
    semaforo = payload.get("semaforo", {})
    total = sum(int(v or 0) for v in semaforo.values()) or 1
    pdf.set_text_color(*CINZA_TEXTO)
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_xy(36, 182)
    pdf.cell(400, 16, ascii_seguro("Como está a rede"))

    y = 210.0
    for chave, rotulo in (
        ("alta", "Prioridade alta"),
        ("media", "Atenção"),
        ("ok", "Sem alerta"),
        ("sem_base", "Sem base de comparação"),
    ):
        quantidade = int(semaforo.get(chave, 0) or 0)
        pdf.set_fill_color(*COR_SEVERIDADE[chave])
        pdf.rect(36, y, 10, 10, style="F")
        pdf.set_text_color(*CINZA_TEXTO)
        pdf.set_font("Helvetica", "", 10)
        pdf.set_xy(52, y - 1)
        pdf.cell(300, 12, ascii_seguro(f"{rotulo}: {quantidade} unidades ({quantidade / total:.0%})"))
        largura_barra = 420.0 * quantidade / total
        pdf.set_fill_color(*COR_SEVERIDADE[chave])
        pdf.rect(300, y, max(largura_barra, 0.5), 10, style="F")
        y += 22.0

    notas = payload.get("notas", [])
    pdf.set_text_color(120, 120, 120)
    pdf.set_font("Helvetica", "", 8.5)
    for indice, nota in enumerate(notas[:4]):
        pdf.set_xy(36, 320 + indice * 13)
        pdf.cell(PAGINA_LARGURA - 72, 11, ascii_seguro(f"- {nota}"))


def _fonte_do_faturamento(payload: Mapping[str, Any]) -> str:
    """Uma frase dizendo de onde veio o dinheiro deste PDF.

    O rodape dizia so' "Growth API" enquanto o faturamento vinha mesmo de la'. Desde que os
    meses fechados passaram a sair da planilha do Financeiro, dizer so' Growth e' errado -- e
    e' o rodape que o franqueado le' quando o numero nao bate com o que ele esperava.
    """
    fonte = payload.get("fonte_faturamento") or {}
    origens = set((fonte.get("por_mes") or {}).values()) | {fonte.get("periodo")}
    if "financeiro" not in origens:
        return "Fonte: Growth API - camada paralela, read-only sobre o M1."
    if origens <= {"financeiro", None}:
        return (
            "Fonte: faturamento da planilha do Financeiro (base dos royalties); "
            "demais metricas da Growth API, read-only sobre o M1."
        )
    return (
        "Fonte: faturamento da planilha do Financeiro nos meses FECHADOS e da Growth API "
        "no periodo parcial; demais metricas da Growth API, read-only sobre o M1."
    )


def _texto_do_rodape(payload: Mapping[str, Any]) -> str:
    """As reguas vigentes VAO no rodape: e' impossivel a tela dizer uma e o motor aplicar
    outra."""
    reguas = payload.get("reguas") or {}
    partes = []
    for regua in reguas.values():
        limiar = regua.get("limiar")
        if regua.get("meses") is not None:
            partes.append(f"{regua.get('rotulo')}: {int(regua['meses'])} meses")
        else:
            sinal = ">" if regua.get("sentido") == "acima" else "<"
            partes.append(f"{regua.get('rotulo')}: {sinal} {_br(limiar, 1)}")
    return ascii_seguro(_fonte_do_faturamento(payload) + " Réguas: " + " | ".join(partes))


def ficha_pdf(payload: Mapping[str, Any]) -> bytes:
    """PDF da ficha de uma unidade: identidade, quarteto, coorte e recomendacoes."""
    unidade = payload.get("unidade", {})
    diagnostico = payload.get("diagnostico", {})
    pdf = UltraPDF()
    pdf.add_page()
    faixa_de_titulo(
        pdf,
        str(unidade.get("nome", "Unidade")),
        f"{unidade.get('uf', '')} - competência {payload.get('mes', '')}",
    )

    pdf.set_text_color(*CINZA_TEXTO)
    pdf.set_font("Helvetica", "", 10)
    identidade = " | ".join(
        parte
        for parte in (
            str(unidade.get("cidade") or ""),
            f"Consultor: {unidade.get('consultor') or 'a atribuir'}",
            f"Master: {unidade.get('master_franquia') or unidade.get('master') or '-'}",
            f"Maturidade: {unidade.get('coorte_rotulo') or '-'}",
            f"Inaugurada em {unidade.get('inauguracao') or '-'}",
        )
        if parte
    )
    pdf.set_xy(36, 68)
    pdf.cell(PAGINA_LARGURA - 72, 12, ascii_seguro(identidade))

    severidade = str(diagnostico.get("severidade", "ok"))
    pdf.set_fill_color(*COR_SEVERIDADE.get(severidade, CINZA_TEXTO))
    pdf.rect(36, 88, 300, 22, style="F")
    pdf.set_text_color(*BRANCO)
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_xy(44, 92)
    pdf.cell(290, 14, ascii_seguro(str(diagnostico.get("severidade_rotulo", ""))))

    metricas = payload.get("metricas", {})
    cabecalho = [
        (220.0, "Métrica", "L"),
        (110.0, "Mês", "R"),
        (110.0, "M-1", "R"),
        (110.0, "Ranking", "R"),
        (130.0, "% vs média da rede", "R"),
        (200.0, "Coorte (mediana)", "R"),
    ]
    y = 124.0
    linha_de_tabela(pdf, 36.0, y, cabecalho, negrito=True, fundo=CINZA_CLARO)
    y += 18.0
    referencias = (payload.get("coorte") or {}).get("metricas", {})
    for chave, rotulo in METRICAS_FICHA:
        metrica = metricas.get(chave) or {}
        if metrica.get("atual") is None:
            continue
        rank, total = metrica.get("rank"), metrica.get("rank_total")
        casas = 1 if chave.endswith("_pct") or chave == "nps" else 0
        referencia = (referencias.get(chave) or {}).get("p50")
        linha_de_tabela(
            pdf,
            36.0,
            y,
            [
                (220.0, rotulo, "L"),
                (110.0, _br(metrica.get("atual"), casas), "R"),
                (110.0, _br(metrica.get("m1"), casas), "R"),
                (110.0, f"{int(rank)}/{int(total)}" if rank and total else "-", "R"),
                (130.0, _br(metrica.get("vs_media_pct"), 1) + "%" if metrica.get("vs_media_pct") is not None else "-", "R"),
                (200.0, _br(referencia, casas) if referencia is not None else "-", "R"),
            ],
        )
        y += 15.0

    y += 8.0
    coorte = payload.get("coorte") or {}
    pdf.set_text_color(120, 120, 120)
    pdf.set_font("Helvetica", "", 8.5)
    pdf.set_xy(36, y)
    pdf.cell(
        PAGINA_LARGURA - 72,
        11,
        ascii_seguro(
            f"Comparada com {coorte.get('n', 0)} unidades - base: {coorte.get('base_rotulo', '-')}"
        ),
    )

    _pagina_de_graficos_da_ficha(pdf, payload)

    recomendacoes = diagnostico.get("recomendacoes") or []
    if recomendacoes:
        pdf.add_page()
        faixa_de_titulo(
            pdf,
            "O que fazer",
            str(unidade.get("nome", "")),
            rgb=ULTRA_TURQUESA,
        )
        y = 84.0
        for item in recomendacoes[:6]:
            pdf.set_text_color(*ULTRA_MAGENTA)
            pdf.set_font("Helvetica", "B", 12)
            pdf.set_xy(36, y)
            pdf.cell(PAGINA_LARGURA - 72, 14, ascii_seguro(str(item.get("titulo", ""))))
            y += 18.0
            pdf.set_text_color(*CINZA_TEXTO)
            pdf.set_font("Helvetica", "", 9.5)
            pdf.set_xy(36, y)
            pdf.multi_cell(PAGINA_LARGURA - 72, 12, ascii_seguro(str(item.get("corpo", ""))))
            y = pdf.get_y() + 12.0
            if y > PAGINA_ALTURA - 70:
                break
    rodape(pdf, _texto_do_rodape(payload))
    return bytes(pdf.output())


# ---------------------------------------------------------------------------
# Paginas de grafico
#
# Espelham os cards do dashboard, na mesma ordem e com os mesmos numeros -- que vem
# prontos do payload, nunca recalculados aqui. O PDF circula sem a tela ao lado: se as
# duas desenharem series diferentes, a divergencia aparece na frente do comite.
# ---------------------------------------------------------------------------

_CORES_FUNIL = (
    ULTRA_TURQUESA,
    (111, 164, 247),
    (217, 74, 134),
    (95, 208, 140),
)


def _pagina_de_graficos_da_rede(
    pdf: UltraPDF, payload: Mapping[str, Any], subtitulo: str
) -> None:
    """Faturamento da rede, distribuicao do semaforo e o split recorrentes x agregadores."""
    meses = list(payload.get("serie_meses") or [])
    serie = list(payload.get("serie_rede") or [])
    if not meses and not payload.get("split"):
        return

    pdf.add_page()
    faixa_de_titulo(pdf, "Como a rede chegou aqui", subtitulo)

    if meses:
        titulo_de_grafico(
            pdf,
            36,
            80,
            "Faturamento da rede no recorte",
            f"Soma dos meses FECHADOS das {payload.get('totais', {}).get('no_recorte', 0)} "
            "unidades do recorte. A competência em curso não entra.",
            largura=PAGINA_LARGURA - 72,
        )
        barras(
            pdf,
            36,
            126,
            PAGINA_LARGURA - 72,
            150,
            [_mes_curto(m) for m in meses],
            serie,
            formatar=_brl_da_barra,
        )

    semaforo = payload.get("semaforo") or {}
    titulo_de_grafico(pdf, 36, 316, "Fila de trabalho", "Severidade pelas réguas do rodapé.", largura=420)
    barra_empilhada(
        pdf,
        36,
        352,
        420,
        16,
        [
            (float(semaforo.get(chave, 0) or 0), COR_SEVERIDADE[chave])
            for chave in ("alta", "media", "ok", "sem_base")
        ],
    )
    legenda_x = 36.0
    for chave, rotulo in (
        ("alta", "Prioridade alta"),
        ("media", "Atenção"),
        ("ok", "Sem alerta"),
        ("sem_base", "Sem base"),
    ):
        quantidade = int(semaforo.get(chave, 0) or 0)
        pdf.set_fill_color(*COR_SEVERIDADE[chave])
        pdf.rect(legenda_x, 380, 8, 8, style="F")
        pdf.set_text_color(110, 110, 110)
        pdf.set_font("Helvetica", "", 8)
        pdf.set_xy(legenda_x + 12, 379)
        texto = f"{rotulo}: {quantidade}"
        pdf.cell(96, 10, ascii_seguro(texto))
        legenda_x += 108.0

    split = payload.get("split") or {}
    recorrentes = float(split.get("recorrentes") or 0)
    agregadores = float(split.get("agregadores") or 0)
    titulo_de_grafico(
        pdf,
        520,
        316,
        "Recorrentes x agregadores",
        "Agregador paga menos por aluno e pode sair em bloco por decisão do parceiro.",
        largura=PAGINA_LARGURA - 556,
    )
    barra_empilhada(
        pdf,
        520,
        352,
        PAGINA_LARGURA - 556,
        16,
        [(recorrentes, ULTRA_TURQUESA), (agregadores, ULTRA_MAGENTA)],
    )
    pdf.set_text_color(110, 110, 110)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_xy(520, 379)
    pdf.cell(
        200,
        10,
        ascii_seguro(f"Recorrentes {_br(recorrentes)} ({_br(split.get('pct_recorrentes'), 0)}%)"),
    )
    pdf.set_xy(PAGINA_LARGURA - 300, 379)
    pdf.cell(
        264,
        10,
        ascii_seguro(f"Agregadores {_br(agregadores)} ({_br(split.get('pct_agregadores'), 0)}%)"),
        align="R",
    )

    sss = payload.get("sss") or {}
    if sss.get("disponivel") and sss.get("metricas"):
        faturamento = sss["metricas"].get("faturamento") or {}
        variacao = faturamento.get("var_pct")
        pdf.set_text_color(*CINZA_TEXTO)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_xy(36, 424)
        pdf.cell(
            PAGINA_LARGURA - 72,
            12,
            # O denominador entra junto de propósito. "18 unidades, +9%" sozinho é lido
            # como o crescimento da carteira inteira; o que o SSS mede é a parte dela que
            # já existia um ano atrás, e o resto ficou de fora porque abrir loja não é
            # crescer. Desde 2026-08 o SSS acompanha os filtros da tela, então este texto
            # fala do RECORTE impresso, não da rede.
            ascii_seguro(
                f"Mesma base ano a ano (SSS): {sss.get('unidades', 0)} de "
                f"{sss.get('unidades_recorte', sss.get('unidades', 0))} unidades do recorte "
                f"estavam abertas nos dois períodos; faturamento "
                f"{'+' if (variacao or 0) >= 0 else ''}{_br(variacao, 1)}% contra "
                f"{sss.get('competencia_base', '')}."
            ),
        )

    rodape(pdf, _texto_do_rodape(payload))


def _pagina_de_graficos_da_ficha(pdf: UltraPDF, payload: Mapping[str, Any]) -> None:
    """Serie de 12 meses, base de alunos, funil, NPS contra a meta e a coorte."""
    unidade = payload.get("unidade", {})
    serie = payload.get("serie") or {}
    meses = [_mes_curto(m) for m in (serie.get("meses") or [])]
    if not meses:
        return

    pdf.add_page()
    faixa_de_titulo(
        pdf,
        "Como está a unidade",
        f"{unidade.get('nome', '')} - competência {payload.get('mes', '')}",
        rgb=ULTRA_MAGENTA,
    )

    titulo_de_grafico(pdf, 36, 76, "Faturamento nos 12 meses fechados")
    barras(pdf, 36, 118, 430, 108, meses, serie.get("faturamento") or [], formatar=_brl_da_barra)

    titulo_de_grafico(pdf, 520, 76, "Alunos ativos")
    linha_de_grafico(pdf, 520, 116, PAGINA_LARGURA - 556, 52, serie.get("ativos") or [])
    titulo_de_grafico(pdf, 520, 182, "Churn (%)")
    linha_de_grafico(
        pdf,
        520,
        212,
        PAGINA_LARGURA - 556,
        44,
        serie.get("churn_pct") or [],
        cor=COR_SEVERIDADE["alta"],
        formatar=lambda v: f"{_br(v, 1)}%",
    )

    # Faixa de baixo: funil | composicao da base | (NPS e coorte, na coluna da direita).
    # As larguras sao explicitas porque os tres blocos dividem a mesma faixa -- foi a
    # sobreposicao entre eles que os renders mostraram.
    funil = payload.get("funil") or {}
    titulo_de_grafico(
        pdf,
        36,
        272,
        "Funil comercial do período",
        str(funil.get("aviso") or f"Conversão de visita em aluno: {_br(funil.get('conversao_pct'), 1)}%"),
        largura=248,
    )
    barras_horizontais(
        pdf,
        36,
        316,
        248,
        [
            ("Visitas", funil.get("visitas"), _CORES_FUNIL[0]),
            ("Convertidos", funil.get("convertidos"), _CORES_FUNIL[1]),
            ("Vendas", funil.get("vendas"), _CORES_FUNIL[2]),
            ("Novos alunos", funil.get("novos_alunos"), _CORES_FUNIL[3]),
        ],
    )

    metricas = payload.get("metricas", {})
    recorrentes = float((metricas.get("pagantes") or {}).get("atual") or 0.0)
    agregadores = float((metricas.get("agregadores") or {}).get("atual") or 0.0)
    dependencia = (metricas.get("pct_agregador_alunos") or {}).get("atual")
    # A rosca é da BASE (cabeças), e o "paga menos" da legenda deixou de ser afirmação
    # genérica: com o faturamento vindo da planilha do Financeiro dá para dizer quanto. As
    # duas leituras discordam em 14,8 p.p. na mediana da rede, então mostrar só a de alunos
    # faz a unidade parecer mais dependente do que o caixa dela mostra.
    dep_receita = (metricas.get("pct_agregador_receita") or {}).get("atual")
    legenda = "Aluno de agregador paga menos e sai em bloco se o parceiro decidir."
    if dep_receita is not None:
        legenda = (
            f"Aluno de agregador paga menos: são {_br(dependencia, 0)}% da base e "
            f"{_br(dep_receita, 0)}% da receita. E saem em bloco se o parceiro decidir."
        )
    titulo_de_grafico(pdf, 304, 272, "Composição da base", legenda, largura=196)
    rosca(
        pdf,
        304,
        318,
        68,
        [
            ("Recorrentes", recorrentes, ULTRA_TURQUESA),
            ("Agregadores", agregadores, ULTRA_MAGENTA),
        ],
        espessura=13,
        centro_valor="-" if dependencia is None else f"{_br(dependencia, 0)}%",
        centro_rotulo="agregadores",
        legenda_abaixo=True,
    )

    meta = float(payload.get("meta_nps") or 60)
    nps = (payload.get("metricas", {}).get("nps") or {}).get("atual")
    titulo_de_grafico(
        pdf,
        520,
        272,
        "NPS contra a meta da rede",
        f"Meta oficial {meta:.0f}. O alerta só dispara bem abaixo dela: meta não é alerta.",
        largura=PAGINA_LARGURA - 556,
    )
    barra_de_meta(pdf, 520, 316, PAGINA_LARGURA - 556, valor=nps, meta=meta)
    pdf.set_text_color(*CINZA_TEXTO)
    pdf.set_font("Helvetica", "B", 15)
    pdf.set_xy(520, 332)
    pdf.cell(120, 18, ascii_seguro("-" if nps is None else _br(nps, 1)))

    coorte = payload.get("coorte") or {}
    referencias = coorte.get("metricas") or {}
    titulo_de_grafico(
        pdf,
        520,
        366,
        "Contra os pares de mesma maturidade",
        f"{coorte.get('n', 0)} unidades - {coorte.get('base_rotulo', '-')}",
        largura=PAGINA_LARGURA - 556,
    )
    y = 404.0
    # "Churn - percentil 92" lê como elogio e é o oposto: 92% dos pares têm churn MENOR.
    # A direção vai escrita ao lado do rótulo, e o percentil continua sendo o número cru
    # (inverter em silêncio seria pior que não mostrar).
    for chave, rotulo, casas in (
        ("faturamento", "Faturamento", 0),
        ("receita_por_recorrente", "Receita/recorrente", 0),
        ("churn_pct", "Churn (menor é melhor)", 1),
    ):
        referencia = referencias.get(chave) or {}
        if referencia.get("unidade") is None:
            continue
        pdf.set_text_color(110, 110, 110)
        pdf.set_font("Helvetica", "", 8)
        pdf.set_xy(520, y - 11)
        pdf.cell(200, 10, ascii_seguro(rotulo))
        pdf.set_text_color(*CINZA_TEXTO)
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_xy(PAGINA_LARGURA - 300, y - 11)
        percentil = referencia.get("percentil")
        pdf.cell(
            264,
            10,
            ascii_seguro(
                f"{_br(referencia.get('unidade'), casas)}"
                + (f" - percentil {percentil:.0f}" if percentil is not None else "")
            ),
            align="R",
        )
        faixa_de_percentil(
            pdf,
            520,
            y,
            PAGINA_LARGURA - 556,
            p25=referencia.get("p25"),
            p50=referencia.get("p50"),
            p75=referencia.get("p75"),
            unidade=referencia.get("unidade"),
        )
        y += 38.0

    rodape(pdf, _texto_do_rodape(payload))


_MESES_PT = ("jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez")


def _mes_curto(competencia: str) -> str:
    """"2026-06" -> "jun". No eixo, o ano e' redundante."""
    partes = str(competencia).split("-")
    if len(partes) < 2 or not partes[1].isdigit():
        return str(competencia)
    return _MESES_PT[int(partes[1]) - 1]
