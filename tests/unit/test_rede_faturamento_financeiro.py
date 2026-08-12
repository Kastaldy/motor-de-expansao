"""Leitura da planilha do Financeiro, portões de qualidade e sobreposição no fechamento.

Os números citados nas asserções vêm da medição feita em 2026-08-12 sobre a planilha real
(`ULTRA - cadastro de franqueados e recebíveis 2026.xlsx`, competências 2021-01..2026-07):
a aritmética interna fecha em 100% das 6.432 células e a Growth subdimensiona a rede em
~20% por ter perdido a receita de agregador em maio/2025.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from motor_expansao.dashboard import rede_faturamento_financeiro as fin
from motor_expansao.dashboard import rede_metricas as rm
from tests.unit.rede_fixtures import base, unidade_saudavel

MESES = ("2026-05", "2026-06", "2026-07")


def _planilha(
    caminho: Path,
    unidades: list[dict[str, object]],
    meses: tuple[str, ...] = MESES,
    *,
    linha_do_cabecalho: int = 4,
    depara: list[tuple[str, object, str]] | None = None,
    aba: str = fin.ABA_FATURAMENTO,
) -> Path:
    """Escreve um .xlsx com a MESMA forma da planilha do Financeiro.

    Cada unidade é um dicionário `{"rotulo": "01 - AUGUSTA", "vendas_ux": [...], ...}`;
    o TOTAL é escrito na linha-cabeçalho do bloco e as componentes logo abaixo.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = aba
    for j, mes in enumerate(meses):
        ws.cell(row=linha_do_cabecalho, column=3 + j, value=pd.Timestamp(f"{mes}-01").to_pydatetime())

    linha = linha_do_cabecalho + 3
    for i, unidade in enumerate(unidades, start=1):
        ws.cell(row=linha, column=1, value=i)
        ws.cell(row=linha, column=2, value=unidade["rotulo"])
        for j in range(len(meses)):
            ws.cell(row=linha, column=3 + j, value=unidade.get("total", [None] * len(meses))[j])
        for k, (rotulo, chave) in enumerate(
            [("VENDAS UX", "vendas_ux"), ("GYMPASS", "gympass"), ("TOTALPASS", "totalpass"), ("(-) TEM SAÚDE", "tem_saude")],
            start=1,
        ):
            ws.cell(row=linha + k, column=2, value=rotulo)
            for j in range(len(meses)):
                ws.cell(row=linha + k, column=3 + j, value=unidade.get(chave, [None] * len(meses))[j])
        linha += 6

    if depara is not None:
        aba_dp = wb.create_sheet(fin.ABA_DEPARA)
        aba_dp.append(["UNIDADE", "COD_UNIDADE", "UNIDADE_UX"])
        for nome, cod, ux in depara:
            aba_dp.append([nome, cod, ux])

    wb.save(caminho)
    return caminho


def _unidade(rotulo: str, ux: float, gym: float = 0.0, tp: float = 0.0, saude: float = 0.0) -> dict[str, object]:
    n = len(MESES)
    return {
        "rotulo": rotulo,
        "total": [ux + gym + tp - saude] * n,
        "vendas_ux": [ux] * n,
        "gympass": [gym] * n,
        "totalpass": [tp] * n,
        "tem_saude": [saude] * n,
    }


# ---------------------------------------------------------------------------
# Leitura
# ---------------------------------------------------------------------------


def test_le_formato_longo_com_componentes(tmp_path: Path) -> None:
    caminho = _planilha(
        tmp_path / "p.xlsx",
        [_unidade("01 - AUGUSTA", 100.0, gym=30.0, tp=20.0, saude=5.0)],
        depara=[("AUGUSTA", "01", "AUGUSTA")],
    )
    fat = fin.ler_planilha(caminho)
    assert len(fat) == len(MESES)
    linha = fat.iloc[0]
    assert linha["cod_unidade"] == "01"
    assert linha["unidade_planilha"] == "AUGUSTA"
    assert linha["unidade_ux"] == "AUGUSTA"
    assert bool(linha["tem_depara"]) is True
    assert linha["faturamento"] == pytest.approx(145.0)
    assert linha["vendas_ux"] == pytest.approx(100.0)


def test_traco_vira_nulo_e_nao_zero(tmp_path: Path) -> None:
    """`-` na planilha é "não existia neste mês". Virar 0 inventaria faturamento nulo."""
    caminho = _planilha(
        tmp_path / "p.xlsx",
        [{"rotulo": "01 - NOVA", "total": ["-", "-", 50.0], "vendas_ux": ["-", "-", 50.0]}],
        depara=[("NOVA", "01", "NOVA")],
    )
    fat = fin.ler_planilha(caminho).set_index("competencia")
    assert pd.isna(fat.loc["2026-05", "faturamento"])
    assert fat.loc["2026-07", "faturamento"] == pytest.approx(50.0)


def test_acha_o_cabecalho_mesmo_fora_da_linha_4(tmp_path: Path) -> None:
    """O layout já mudou uma vez DENTRO do mesmo arquivo (cabeçalho na 4 e na 6)."""
    caminho = _planilha(tmp_path / "p.xlsx", [_unidade("01 - X", 10.0)], linha_do_cabecalho=6, depara=[("X", "01", "X")])
    fat = fin.ler_planilha(caminho)
    assert sorted(fat["competencia"]) == list(MESES)


def test_prefixo_de_uf_nao_vira_codigo(tmp_path: Path) -> None:
    """O snapshot antigo rotula os blocos como "SP - AUGUSTA"; "SP" não é código."""
    caminho = _planilha(tmp_path / "p.xlsx", [_unidade("SP - AUGUSTA", 10.0)], depara=[("AUGUSTA", "01", "AUGUSTA")])
    fat = fin.ler_planilha(caminho)
    assert fat["cod_unidade"].isna().all()
    assert fat["unidade_planilha"].unique().tolist() == ["AUGUSTA"]


def test_codigo_numerico_casa_com_texto_no_depara(tmp_path: Path) -> None:
    """O Excel entrega "01" como texto numa aba e 1 como número noutra."""
    caminho = _planilha(tmp_path / "p.xlsx", [_unidade("01 - AUGUSTA", 10.0)], depara=[("AUGUSTA", 1, "AUGUSTA - SP")])
    fat = fin.ler_planilha(caminho)
    assert fat["unidade_ux"].unique().tolist() == ["AUGUSTA - SP"]
    assert bool(fat["tem_depara"].all()) is True


def test_sem_aba_de_depara_degrada_para_o_nome(tmp_path: Path) -> None:
    caminho = _planilha(tmp_path / "p.xlsx", [_unidade("01 - AUGUSTA", 10.0)], depara=None)
    fat = fin.ler_planilha(caminho)
    assert fat["unidade_ux"].unique().tolist() == ["AUGUSTA"]
    assert not fat["tem_depara"].any()


def test_aba_inexistente_falha_claro(tmp_path: Path) -> None:
    caminho = _planilha(tmp_path / "p.xlsx", [_unidade("01 - X", 1.0)])
    with pytest.raises(ValueError, match="não existe"):
        fin.ler_planilha(caminho, aba="Aba Que Nao Existe")


# ---------------------------------------------------------------------------
# Portões
# ---------------------------------------------------------------------------


def _fat_valida(tmp_path: Path) -> pd.DataFrame:
    caminho = _planilha(
        tmp_path / "ok.xlsx",
        [_unidade("01 - A", 100.0, gym=10.0), _unidade("02 - B", 200.0, tp=20.0)],
        depara=[("A", "01", "A"), ("B", "02", "B")],
    )
    return fin.ler_planilha(caminho)


def test_planilha_integra_nao_acusa_erro(tmp_path: Path) -> None:
    achados = fin.validar(_fat_valida(tmp_path), hoje=date(2026, 8, 12))
    assert [a for a in achados if a.eh_erro] == []


def test_aritmetica_que_nao_fecha_e_erro(tmp_path: Path) -> None:
    """O selo de integridade da fonte: TOTAL = UX + GYM + TP - TEM SAÚDE, em 100%."""
    caminho = _planilha(
        tmp_path / "p.xlsx",
        [{"rotulo": "01 - A", "total": [999.0] * 3, "vendas_ux": [100.0] * 3, "gympass": [10.0] * 3}],
        depara=[("A", "01", "A")],
    )
    achados = fin.validar(fin.ler_planilha(caminho), hoje=date(2026, 8, 12))
    assert [a.codigo for a in achados if a.eh_erro] == ["aritmetica"]


def test_mes_ainda_aberto_e_erro(tmp_path: Path) -> None:
    """Ler cedo demais grava uma competência pela metade como se fosse definitiva.

    O risco é medido, não hipotético: o snapshot antigo embutido na planilha real tinha 19
    unidades zeradas na sua última competência, que depois viraram valor cheio.
    """
    achados = fin.validar(_fat_valida(tmp_path), hoje=date(2026, 7, 20))
    erros = [a for a in achados if a.eh_erro]
    assert [a.codigo for a in erros] == ["mes_aberto"]
    assert "2026-07" in erros[0].mensagem


def test_planilha_atrasada_e_so_aviso(tmp_path: Path) -> None:
    achados = fin.validar(_fat_valida(tmp_path), hoje=date(2026, 10, 3))
    assert [a for a in achados if a.eh_erro] == []
    assert "planilha_atrasada" in [a.codigo for a in achados]


def test_buraco_na_serie_e_erro(tmp_path: Path) -> None:
    fat = _fat_valida(tmp_path)
    achados = fin.validar(fat[fat["competencia"] != "2026-06"], hoje=date(2026, 8, 12))
    assert "buraco" in [a.codigo for a in achados if a.eh_erro]


def test_queda_de_cobertura_avisa(tmp_path: Path) -> None:
    """Metade da rede sem faturamento no último mês = mês provavelmente incompleto."""
    fat = _fat_valida(tmp_path).copy()
    fat.loc[(fat["competencia"] == "2026-07") & (fat["unidade_planilha"] == "A"), "faturamento"] = 0.0
    achados = fin.validar(fat, hoje=date(2026, 8, 12))
    assert "queda_cobertura" in [a.codigo for a in achados]


def test_reescrita_de_mes_fechado_avisa(tmp_path: Path) -> None:
    anterior = _fat_valida(tmp_path)
    agora = anterior.copy()
    agora.loc[agora["competencia"] == "2026-05", "faturamento"] += 1_000.0
    achados = fin.validar(agora, hoje=date(2026, 8, 12), anterior=anterior)
    assert "reescrita" in [a.codigo for a in achados]


def test_ultima_competencia_do_snapshot_anterior_pode_mudar(tmp_path: Path) -> None:
    """Ela era o mês em preenchimento: mudar ali é o comportamento normal, não um alerta."""
    anterior = _fat_valida(tmp_path)
    agora = anterior.copy()
    agora.loc[agora["competencia"] == "2026-07", "faturamento"] += 1_000.0
    achados = fin.validar(agora, hoje=date(2026, 8, 12), anterior=anterior)
    assert "reescrita" not in [a.codigo for a in achados]


def test_competencia_fechada_e_sempre_o_mes_anterior() -> None:
    assert fin.competencia_fechada(date(2026, 8, 1)) == "2026-07"
    assert fin.competencia_fechada(date(2026, 8, 31)) == "2026-07"
    assert fin.competencia_fechada(date(2026, 1, 5)) == "2025-12"


def test_carregar_parquet_ausente_devolve_vazio(tmp_path: Path) -> None:
    assert not len(fin.carregar(tmp_path / "nao-existe.parquet"))


def test_carregar_recusa_esquema_errado(tmp_path: Path) -> None:
    caminho = tmp_path / "ruim.parquet"
    pd.DataFrame({"algo": [1]}).to_parquet(caminho)
    with pytest.raises(ValueError, match="sem as colunas"):
        fin.carregar(caminho)


# ---------------------------------------------------------------------------
# Sobreposição no fechamento
# ---------------------------------------------------------------------------


def _base_growth() -> pd.DataFrame:
    return rm.preparar_base(
        base(
            unidade_saudavel("AUGUSTA", 2026, 6),
            unidade_saudavel("AUGUSTA", 2026, 7),
            unidade_saudavel("MOOCA", 2026, 6),
            unidade_saudavel("MOOCA", 2026, 7),
        )
    )


def _financeiro(**por_unidade: tuple[float, float]) -> pd.DataFrame:
    """`{"AUGUSTA": (total, vendas_ux)}` -> quadro no formato do parquet do Financeiro."""
    linhas = []
    for nome, (total, ux) in por_unidade.items():
        for mes in ("2026-06", "2026-07"):
            linhas.append(
                {
                    "cod_unidade": "01",
                    "unidade_planilha": nome,
                    "unidade_ux": nome,
                    "tem_depara": True,
                    "competencia": mes,
                    "faturamento": total,
                    "vendas_ux": ux,
                    "gympass": total - ux,
                    "totalpass": 0.0,
                    "tem_saude": 0.0,
                }
            )
    return pd.DataFrame(linhas)


def test_sobreposicao_e_override_e_nunca_insert() -> None:
    """Unidade que só existe na planilha não vira linha nova na carteira.

    Ela não teria pagantes, churn nem NPS — entraria na tela pela metade e ainda mexeria no
    denominador de toda média ponderada da aba. Medido na base real: São Carlos-SP, Jardim
    das Américas-MT e Vila Izabel-PR estão na planilha e não existem na Growth.
    """
    growth = _base_growth()
    fin_df = _financeiro(AUGUSTA=(300_000.0, 250_000.0), FANTASMA=(999.0, 999.0))
    antes = rm.fechamento_mensal(growth)
    depois = rm.fechamento_mensal(growth, financeiro=fin_df)
    assert len(depois) == len(antes)
    assert "FANTASMA" in depois.attrs["financeiro_sem_par"]


def test_sobreposicao_troca_faturamento_e_recalcula_o_agregador() -> None:
    growth = _base_growth()
    fech = rm.fechamento_mensal(growth, financeiro=_financeiro(AUGUSTA=(300_000.0, 250_000.0)))
    linha = fech.set_index(["unidade_cru", "competencia"]).loc[("AUGUSTA", "2026-07")]
    assert linha["faturamento"] == pytest.approx(300_000.0)
    assert linha["faturamento_sem_agregador"] == pytest.approx(250_000.0)
    # A derivada acompanha: era 0 na Growth desde maio/2025, volta a existir aqui.
    assert linha["faturamento_agregador"] == pytest.approx(50_000.0)
    assert linha["origem_faturamento"] == rm.ORIGEM_FINANCEIRO
    # Quem não casou fica na Growth, com o rótulo dizendo isso.
    outra = fech.set_index(["unidade_cru", "competencia"]).loc[("MOOCA", "2026-07")]
    assert outra["origem_faturamento"] == rm.ORIGEM_UX
    assert outra["faturamento"] == pytest.approx(200_000.0)


def test_sobreposicao_nao_toca_no_que_nao_e_faturamento() -> None:
    growth = _base_growth()
    antes = rm.fechamento_mensal(growth).set_index(["unidade_id", "competencia"])
    depois = rm.fechamento_mensal(
        growth, financeiro=_financeiro(AUGUSTA=(300_000.0, 250_000.0))
    ).set_index(["unidade_id", "competencia"])
    for coluna in ("pagantes", "ativos", "cancelados", "churn_pct", "nps", "conversao_pct", "agregadores"):
        pd.testing.assert_series_equal(antes[coluna], depois[coluna], check_names=False)


def test_janela_parcial_nunca_recebe_a_planilha() -> None:
    """A planilha é MENSAL: ela não sabe responder por 1 a 10 de agosto."""
    growth = _base_growth()
    fech = rm.fechamento_mensal(growth, dia_corte=10, financeiro=_financeiro(AUGUSTA=(300_000.0, 250_000.0)))
    assert set(fech["origem_faturamento"]) == {rm.ORIGEM_UX}
    augusta = fech.set_index(["unidade_cru", "competencia"]).loc[("AUGUSTA", "2026-07")]
    assert augusta["faturamento"] < 300_000.0


def test_resgate_por_aperto_casa_variacao_de_espaco_e_hifen() -> None:
    """"CEILANDIA QNN 32 - DF" x "CEILANDIA QNN32 - DF " — o de-para não cobre isso."""
    growth = rm.preparar_base(
        base(
            unidade_saudavel("CEILANDIA QNN32 - DF ", 2026, 6, uf="DF"),
            unidade_saudavel("CEILANDIA QNN32 - DF ", 2026, 7, uf="DF"),
        )
    )
    fin_df = _financeiro(**{"CEILANDIA QNN 32 - DF": (300_000.0, 250_000.0)})
    fech = rm.fechamento_mensal(growth, financeiro=fin_df)
    assert set(fech["origem_faturamento"]) == {rm.ORIGEM_FINANCEIRO}
    assert fech.attrs["financeiro_sem_par"] == []


def test_resgate_nao_colapsa_as_duas_aguas_claras() -> None:
    """A academia e o studio são unidades DIFERENTES; o sufixo de UF os separa.

    `chave_unidade` funde as duas de propósito (é chave de join, não de identidade) — se o
    resgate usasse ela, o faturamento da academia iria parar no studio.
    """
    assert rm._apertar("AGUAS CLARAS") != rm._apertar("AGUAS CLARAS - DF")
    growth = rm.preparar_base(
        base(
            unidade_saudavel("AGUAS CLARAS", 2026, 6, uf="DF"),
            unidade_saudavel("AGUAS CLARAS", 2026, 7, uf="DF"),
        )
    )
    fech = rm.fechamento_mensal(
        growth, financeiro=_financeiro(**{"AGUAS CLARAS - DF": (300_000.0, 250_000.0)})
    )
    assert set(fech["origem_faturamento"]) == {rm.ORIGEM_UX}


def test_resgate_recusa_ambiguidade() -> None:
    """Duas candidatas na mesma chave apertada -> ninguém casa.

    Um palpite errado joga o faturamento de uma academia na outra; um buraco visível é
    melhor que um número silenciosamente trocado.
    """
    resgate = rm._resgate_por_aperto({"A B": "id-1", "AB": "id-2"}, ["A-B"])
    assert resgate == {}


def test_sem_planilha_a_aba_continua_de_pe() -> None:
    growth = _base_growth()
    fech = rm.fechamento_mensal(growth, financeiro=None)
    assert set(fech["origem_faturamento"]) == {rm.ORIGEM_UX}
    assert len(fech) == 4


def test_base_vazia_com_planilha_nao_quebra() -> None:
    vazio = rm.fechamento_mensal(pd.DataFrame(), financeiro=_financeiro(AUGUSTA=(1.0, 1.0)))
    assert not len(vazio)
    assert "origem_faturamento" in vazio.columns
