"""Testes do export Excel do simulador de viabilidade — BLK-DIM-22.

CA-12a..g: bytes nao vazio, xlsx valido + 4 abas, campos na aba Resumo,
60 linhas na Curva, sem escrita em disco, nome_ponto, sensibilidade.
ZERO leitura de parquets ou arquivos externos.
"""
from __future__ import annotations

from io import BytesIO

import openpyxl
import pandas as pd
import pytest

from motor_expansao.dimensionamento.excel_export import gerar_excel_viabilidade
from motor_expansao.dimensionamento.simulador import ViabilidadeResult, gerar_serie_mensal
from motor_expansao.dimensionamento.viabilidade_ponto import ViabilidadePontoResult

# ---------------------------------------------------------------------------
# Fixtures compartilhadas
# ---------------------------------------------------------------------------

_VIAB_RESULT = ViabilidadeResult(
    faturamento_mensal_steady=170_000.0,
    receita_liquida=169_150.0,
    receita_pos_impostos=155_618.0,
    ebitda_mensal=28_000.0,
    margem_ebitda_pct=0.165,
    payback_meses=24.0,
    roic_anual=0.28,
    lucro_liquido_mensal=19_500.0,
    flag_viavel=True,
)

_RESULT = ViabilidadePontoResult(
    lat=-23.5505,
    lng=-46.6333,
    m2=1500.0,
    aluguel_pedido=20_000.0,
    demanda_premissa=938.0,
    faixa_alunos_p10=600.0,
    faixa_alunos_p50=750.0,
    faixa_alunos_p90=900.0,
    n_comparaveis=5,
    flag_zona_morta=False,
    motivo_zona_morta="ok",
    pop_captacao=45_000.0,
    renda_per_capita_captacao=6_500.0,
    viabilidade=_VIAB_RESULT,
    aluguel_teto_calculado=28_000.0,
    alunos_breakeven=550.0,
    grade_sensibilidade=pd.DataFrame({
        "alunos":        [200.0, 200.0, 400.0, 400.0],
        "aluguel":       [12000.0, 20000.0, 12000.0, 20000.0],
        "fator_aluguel": [0.6, 1.0, 0.6, 1.0],
        "margem_liq":    [0.15, 0.05, 0.20, 0.12],
        "viavel":        [True, False, True, True],
        "payback":       [20.0, 38.0, 15.0, 22.0],
    }),
    alunos_balcao_premissa=647.22,
    alunos_agregadores_premissa=290.78,
    alunos_para_margem_alvo=600.0,
    demanda_fonte="premissa_explicita",
)

_SERIE = gerar_serie_mensal(
    647.22, 1500.0, 20_000.0, 137.0,
    alunos_agregadores=290.78,
)


# ---------------------------------------------------------------------------
# CA-12a: retorna bytes nao vazio
# ---------------------------------------------------------------------------


def test_retorna_bytes_nao_vazio() -> None:
    """CA-12a: gerar_excel_viabilidade retorna bytes nao vazio."""
    result = gerar_excel_viabilidade(_RESULT, _SERIE)
    assert isinstance(result, bytes)
    assert len(result) > 0


# ---------------------------------------------------------------------------
# CA-12b: xlsx valido + 4 abas
# ---------------------------------------------------------------------------


def test_xlsx_valido_abrivel() -> None:
    """CA-12b: bytes sao um .xlsx valido abrivel com openpyxl."""
    result = gerar_excel_viabilidade(_RESULT, _SERIE)
    wb = openpyxl.load_workbook(BytesIO(result))
    assert wb is not None


def test_4_abas_presentes() -> None:
    """CA-12b+: 4 abas presentes com nomes exatos."""
    result = gerar_excel_viabilidade(_RESULT, _SERIE)
    wb = openpyxl.load_workbook(BytesIO(result))
    assert set(wb.sheetnames) == {"Resumo", "DRE", "Sensibilidade", "Curva"}


# ---------------------------------------------------------------------------
# CA-12c: campos na aba Resumo
# ---------------------------------------------------------------------------


def test_aba_resumo_contem_campos_chave() -> None:
    """CA-12c: aba Resumo contem campos-chave de ViabilidadeResult."""
    result_bytes = gerar_excel_viabilidade(_RESULT, _SERIE, nome_ponto="Teste SP")
    wb = openpyxl.load_workbook(BytesIO(result_bytes))
    ws = wb["Resumo"]
    # Coletar todos os valores da coluna A (indicadores)
    indicadores = {ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)}
    # Verificar presença de campos minimos
    for campo in ["Faturamento bruto/mês", "Margem EBITDA", "Payback (meses)", "Viável?"]:
        assert campo in indicadores, f"Campo '{campo}' ausente na aba Resumo"


# ---------------------------------------------------------------------------
# CA-12d: 60 linhas na Curva
# ---------------------------------------------------------------------------


def test_aba_curva_tem_60_linhas() -> None:
    """CA-12d: aba Curva tem exatamente 60 linhas de dados (excluindo cabecalhos)."""
    result_bytes = gerar_excel_viabilidade(_RESULT, _SERIE)
    wb = openpyxl.load_workbook(BytesIO(result_bytes))
    ws = wb["Curva"]
    # linha 1 = cabeçalho Ultra, linha 2 = cabeçalho de colunas, linhas 3..62 = dados
    data_rows = [
        r for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value not in (None, "", "Mês", "Projeção Financeira — 60 meses")
        and isinstance(ws.cell(row=r, column=1).value, (int, float))
    ]
    assert len(data_rows) == 60


# ---------------------------------------------------------------------------
# CA-12e: sem escrita em disco
# ---------------------------------------------------------------------------


def test_sem_escrita_em_disco(tmp_path: object, monkeypatch: pytest.MonkeyPatch) -> None:
    """CA-12e: gerar_excel_viabilidade nao escreve em disco em caminhos de usuario.

    Verifica que a funcao retorna bytes via BytesIO, nao persiste o arquivo em
    nenhum caminho controlado pelo usuario (fora de /tmp — openpyxl usa /tmp
    internamente durante a compressao zip, o que e comportamento esperado).
    """
    import builtins
    import os
    import tempfile

    original_open = builtins.open

    # Caminhos de usuario nao-esperados: excluindo /tmp (uso interno do openpyxl)
    user_written: list[str] = []

    def _guarded_open(file: object, mode: str = "r", *args: object, **kwargs: object) -> object:
        if isinstance(file, str) and "w" in str(mode):
            # Temporarios sao permitidos: o openpyxl usa um arquivo temporario para
            # montar o zip do .xlsx. A allowlist era so' POSIX (`/tmp`, `/var/tmp`) e
            # por isso o teste FALHAVA em qualquer Windows, onde o temporario cai em
            # `%TEMP%` -- passava no CI apenas porque os 9 jobs sao `ubuntu-latest`.
            # `tempfile.gettempdir()` resolve o diretorio da plataforma em que se roda.
            normalized = os.path.normpath(file)
            permitidos = (
                os.path.normpath(tempfile.gettempdir()),
                os.path.normpath("/tmp"),
                os.path.normpath("/var/tmp"),
            )
            if not any(normalized.startswith(base) for base in permitidos):
                user_written.append(str(file))
        return original_open(file, mode, *args, **kwargs)  # type: ignore[call-overload]

    monkeypatch.setattr(builtins, "open", _guarded_open)

    result = gerar_excel_viabilidade(_RESULT, _SERIE)

    # Verifica que o retorno e bytes validos (BytesIO foi usado)
    assert isinstance(result, bytes) and len(result) > 0
    # Verifica que nenhum caminho de usuario foi escrito
    assert user_written == [], f"Escrita em caminho de usuario detectada: {user_written}"


# ---------------------------------------------------------------------------
# CA-12f: nome_ponto aparece no cabecalho da aba Resumo
# ---------------------------------------------------------------------------


def test_nome_ponto_no_resumo() -> None:
    """CA-12f: aba Resumo contem o nome_ponto informado no cabecalho."""
    result_bytes = gerar_excel_viabilidade(_RESULT, _SERIE, nome_ponto="Jardins SP")
    wb = openpyxl.load_workbook(BytesIO(result_bytes))
    ws = wb["Resumo"]
    todos_valores = [
        ws.cell(row=r, column=c).value
        for r in range(1, 5)
        for c in range(1, 4)
    ]
    assert any("Jardins SP" in str(v) for v in todos_valores if v)


# ---------------------------------------------------------------------------
# CA-12g: sensibilidade contem valores da grade
# ---------------------------------------------------------------------------


def test_sensibilidade_valores_corretos() -> None:
    """CA-12g: aba Sensibilidade contem valores de margem_liq da grade."""
    result_bytes = gerar_excel_viabilidade(_RESULT, _SERIE)
    wb = openpyxl.load_workbook(BytesIO(result_bytes))
    ws = wb["Sensibilidade"]
    # Coletar todos os valores numericos da aba
    numericos = [
        ws.cell(row=r, column=c).value
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
        if isinstance(ws.cell(row=r, column=c).value, (int, float))
    ]
    # Os valores de margem_liq da grade sao [0.15, 0.05, 0.20, 0.12]
    for valor in [0.15, 0.05, 0.20, 0.12]:
        assert any(abs(v - valor) < 1e-6 for v in numericos), \
            f"Valor {valor} nao encontrado na aba Sensibilidade"


# ---------------------------------------------------------------------------
# Pentest Onda B #12: nome_ponto iniciado em metacaractere de formula e neutralizado
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("prefixo", ["=", "+", "-", "@", "\t", "\r"])
def test_nome_ponto_iniciado_em_formula_e_neutralizado(prefixo: str) -> None:
    """Um `nome_ponto` como `=HYPERLINK(...)` NAO pode virar formula viva no XLSX."""
    payload = f"{prefixo}HYPERLINK(\"http://evil.example\",\"x\")"
    result_bytes = gerar_excel_viabilidade(_RESULT, _SERIE, nome_ponto=payload)
    wb = openpyxl.load_workbook(BytesIO(result_bytes))
    c = wb["Resumo"].cell(row=2, column=1)
    assert c.data_type != "f", "a celula virou formula viva"
    # Propriedade de seguranca: o valor comeca por `'` (neutralizado). Nao checamos o
    # char seguinte porque o openpyxl normaliza \r -> \n no XML da celula.
    assert str(c.value).startswith("'")


def test_nome_ponto_benigno_nao_ganha_apostrofo() -> None:
    """Nome normal passa intacto (sem o apostrofo de neutralizacao)."""
    result_bytes = gerar_excel_viabilidade(_RESULT, _SERIE, nome_ponto="Jardins SP")
    c = openpyxl.load_workbook(BytesIO(result_bytes))["Resumo"].cell(row=2, column=1)
    assert c.value == "Jardins SP"
