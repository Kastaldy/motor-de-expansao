"""Testes do parser do simulador (fixture .xlsx sintetica; sem o arquivo real)."""

from __future__ import annotations

import openpyxl
import pytest

from motor_expansao.dimensionamento.simulador_parser import parse_simulador


@pytest.fixture
def xlsx_sintetico(tmp_path):
    """Gera um .xlsx minimo com as mesmas celulas-chave (sem dado real)."""
    wb = openpyxl.Workbook()
    sim = wb.active
    sim.title = "Simulador"
    sim["E9"] = 500
    sim["E10"] = 938
    sim["E11"] = 651
    sim["E12"] = 0.06
    sim["E13"] = 8
    sim["J9"] = "=IF(N12=0,137,IF(N12=1,147,0))"  # FORMULA
    sim["J10"] = 99
    sim["J11"] = 0
    sim["J12"] = 12
    sim["N9"] = 20000
    sim["N10"] = 0
    sim["N11"] = 0.08
    sim["N12"] = 0
    sim["R9"] = "=IF(N12=0,SUM(FC!C11:C16)*-1,0)"  # FORMULA
    sim["R10"] = 140000
    sim["R11"] = 1.5
    sim["R12"] = "presumido"
    sim["G7"] = 4
    sim["I7"] = 1

    dre = wb.create_sheet("DRE")
    dre["F30"] = 0.005
    dre["F61"] = "=Simulador!N11"
    dre["F63"] = 0.02
    dre["F67"] = 0.02
    dre["F79"] = 0.0105

    trib = wb.create_sheet("Tributos")
    trib["E38"] = 0.0065
    trib["E40"] = 0.03
    trib["E42"] = 0.03
    trib["E44"] = "=32%*25%"
    trib["E46"] = "=32%*9%"

    path = tmp_path / "simulador.xlsx"
    wb.save(path)
    return path


def test_parse_drivers_valores(xlsx_sintetico):
    est = parse_simulador(xlsx_sintetico)
    d = est["drivers"]
    assert d["alunos_inicial"]["valor_default"] == 500
    assert d["alunos_inicial"]["celula"] == "E9"
    assert d["churn"]["valor_default"] == 0.06
    assert d["maturacao_meses"]["valor_default"] == 8
    assert d["royalties_pct"]["valor_default"] == 0.08
    assert d["regime_tributario"]["valor_default"] == "presumido"
    assert d["taxa_franquia"]["valor_default"] == 140000


def test_parse_celulas_formula_registram_string(xlsx_sintetico):
    d = parse_simulador(xlsx_sintetico)["drivers"]
    # J9 e R9 sao formulas -> valor_default None + formula string
    assert d["mensalidade"]["valor_default"] is None
    assert d["mensalidade"]["formula"].startswith("=IF(N12=0,137")
    assert d["capex_total"]["valor_default"] is None
    assert d["capex_total"]["formula"].startswith("=IF(N12=0,SUM(FC!")


def test_parse_ratios_dre(xlsx_sintetico):
    r = parse_simulador(xlsx_sintetico)["ratios_dre"]
    assert r["devolucoes_pct_receita"]["valor_default"] == 0.005
    assert r["marketing_pct_receita"]["valor_default"] == 0.02
    assert r["manutencao_pct_receita"]["valor_default"] == 0.02
    assert r["cartoes_pct_receita"]["valor_default"] == 0.0105
    # royalties no DRE e formula =Simulador!N11
    assert r["royalties_pct_receita"]["valor_default"] is None
    assert r["royalties_pct_receita"]["formula"] == "=Simulador!N11"


def test_parse_impostos_presumido(xlsx_sintetico):
    imp = parse_simulador(xlsx_sintetico)["impostos_presumido"]
    assert imp["pis"]["valor_default"] == 0.0065
    assert imp["cofins"]["valor_default"] == 0.03
    assert imp["iss"]["valor_default"] == 0.03
    # IR efetivo = 32%*25% = 8% (registrado como formula)
    assert imp["ir_aliquota"]["formula"] == "=32%*25%"


def test_estrutura_top_level(xlsx_sintetico):
    est = parse_simulador(xlsx_sintetico)
    assert est["regime_tributario"] == "presumido"
    assert set(est.keys()) >= {"fonte", "drivers", "ratios_dre", "impostos_presumido"}
    assert len(est["drivers"]) == 19
