"""Testes do simulador financeiro XLSX com formulas vivas (FIN-VIAB-01).

DOIS NIVEIS DE TESTE, de proposito:

1) ESTRUTURA (openpyxl, sempre roda). openpyxl NAO avalia formula, entao aqui se
   testa a estrutura do arquivo (8 abas, 64 colunas de mes, sublinhas de custo),
   que as celulas de DRE/fluxo/resumo contem FORMULA (e nao valor cravado) e que
   os VALORES DO MOTOR gravados na aba Afericao conferem com `simular()` ao vivo —
   e a afericao que defende o arquivo na frente do investidor, e um parametro de
   comparacao errado la tornaria a aba inutil (ou pior, tranquilizadora a toa).

2) RECALCULO REAL (pacote `formulas`, SKIP quando nao instalado). O nivel 1 nao
   pega formula que compila e devolve o numero errado — a rodada anterior achou 4
   defeitos exatamente assim. Os testes de recalculo avaliam as 4.293 formulas do
   arquivo e comparam DRE e fluxo nos 64 meses, mais as duas colunas da Afericao,
   contra `simular()`. Instalar com `python -m pip install formulas` (nao e
   dependencia do projeto; no CI estes testes ficam SKIPPED).

Cobertura especifica das duas mudancas de produto de 2026-07-24 (folha FIXA desde o
mes 1 e taxa de franquia PARCELADA 4x): as duas sao INVISIVEIS no mes de steady, por
isso os testes olham os meses de RAMPA (1/4/8), o degrau de reajuste (M13) e os meses
de pre-abertura (M-4..M-1) — nao so a coluna de regime pleno.
"""

from __future__ import annotations

import re
import unicodedata
import warnings
import zipfile
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import get_column_letter

from motor_expansao.dimensionamento.config import (
    SIM_OUTROS_FIXOS_MES,
    SIM_PARCELAS_FRANQUIA_DEFAULT,
)
from motor_expansao.dimensionamento.simulador import Premissas, simular
from motor_expansao.dimensionamento.simulador_xlsx import (
    _DRE_ROW,
    _FLX_ROW,
    _FOLHA_MODO_ROW,
    _INVEST_ROW,
    _MES_COL_INI,
    _MODOS_FOLHA,
    _OUTROS_FIXOS_DECOMP,
    _RESUMO_ALTERNA,
    _RESUMO_ROW_INI,
    _RESUMO_SAIDA,
    _VERMELHO_FONTE,
    ABA_AFERICAO,
    ABA_DRE,
    ABA_FLUXO,
    ABA_FOLHA,
    ABA_INVESTIMENTO,
    ABA_PREMISSAS,
    ABA_RESUMO,
    ABAS_ESPERADAS,
    gerar_simulador_xlsx,
)

# Caso golden (Boulevard Londrina) — os mesmos inputs do resto do ciclo.
_DEMANDA = 2304.0
_INVEST = {
    "obra": 600_000.0,
    "parcelas_obra": 4,
    "equipamentos": 1_400_000.0,
    "prazo_equipamentos": 60,
    "juros_equipamentos_am": 0.018,
    "taxa_franquia": 160_000.0,
}

# Indice da coluna de cada mes na linha do tempo (M-4 = 0 ... M60 = 63).
_COL_DE_MES = {m: i for i, m in enumerate([-4, -3, -2, -1, *range(1, 61)])}


def _col(mes: int) -> int:
    return _MES_COL_INI + _COL_DE_MES[mes]


def _premissas() -> Premissas:
    return Premissas(ticket_cheio=147.0, aluguel_mes=30_000.0, maturacao_meses=8)


@pytest.fixture(scope="module")
def blob() -> bytes:
    return gerar_simulador_xlsx(
        _DEMANDA, _premissas(), nome_ponto="Boulevard Londrina", **_INVEST
    )


@pytest.fixture(scope="module")
def wb(blob: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(BytesIO(blob), data_only=False)


@pytest.fixture(scope="module")
def resultado():
    return simular(_DEMANDA, _premissas(), **_INVEST)


def _norm(texto: str) -> str:
    """Compara rotulos sem depender de acento nem de travessao vs hifen."""
    sem = unicodedata.normalize("NFKD", str(texto))
    sem = "".join(c for c in sem if not unicodedata.combining(c))
    for traco in ("—", "–", "−"):
        sem = sem.replace(traco, "-")
    return re.sub(r"\s+", " ", sem).strip().lower()


# ---------------------------------------------------------------------------
# Arquivo e estrutura
# ---------------------------------------------------------------------------


def test_arquivo_e_xlsx_valido(blob: bytes) -> None:
    assert blob[:2] == b"PK", "xlsx e um container ZIP"
    with zipfile.ZipFile(BytesIO(blob)) as z:
        assert z.testzip() is None
        nomes = z.namelist()
    assert "xl/workbook.xml" in nomes
    assert any(n.startswith("xl/worksheets/sheet") for n in nomes)


def test_reabre_sem_warning_de_corrupcao(blob: bytes) -> None:
    with warnings.catch_warnings(record=True) as capturados:
        warnings.simplefilter("always")
        openpyxl.load_workbook(BytesIO(blob), data_only=False)
    assert [str(w.message) for w in capturados] == []


def test_oito_abas_na_ordem_esperada(wb: openpyxl.Workbook) -> None:
    assert wb.sheetnames == list(ABAS_ESPERADAS)
    assert len(ABAS_ESPERADAS) == 8


def test_recalculo_ao_abrir_esta_ligado(wb: openpyxl.Workbook) -> None:
    # Sem valor em cache, o Excel precisa recalcular na abertura.
    assert wb.calculation.fullCalcOnLoad is True


# ---------------------------------------------------------------------------
# Linha do tempo
# ---------------------------------------------------------------------------


def test_dre_tem_64_colunas_de_mes_comecando_em_m_menos_4(wb: openpyxl.Workbook) -> None:
    ws = wb[ABA_DRE]
    rotulos: list[str] = []
    col = _MES_COL_INI
    while True:
        v = ws.cell(row=4, column=col).value
        if not isinstance(v, str) or not re.fullmatch(r"M-?\d+", v):
            break
        rotulos.append(v)
        col += 1
    assert len(rotulos) == 64
    assert rotulos[0] == "M-4"
    assert rotulos[:5] == ["M-4", "M-3", "M-2", "M-1", "M1"]
    assert rotulos[-1] == "M60"
    # Nao existe "M0": a linha do tempo pula de M-1 para M1, como o motor.
    assert "M0" not in rotulos


def test_linha_do_tempo_da_dre_bate_com_a_serie_do_motor(
    wb: openpyxl.Workbook, resultado
) -> None:
    ws = wb[ABA_DRE]
    meses_planilha = [
        ws.cell(row=_DRE_ROW["mes"], column=_MES_COL_INI + j).value for j in range(64)
    ]
    assert meses_planilha == [int(r["mes"]) for r in resultado.serie_mensal]


def test_fluxo_usa_a_mesma_linha_do_tempo_da_dre(wb: openpyxl.Workbook) -> None:
    ws = wb[ABA_FLUXO]
    rotulos = [ws.cell(row=4, column=_MES_COL_INI + j).value for j in range(64)]
    assert rotulos[0] == "M-4"
    assert rotulos[-1] == "M60"


def test_dre_tem_coluna_de_steady_apos_os_meses(wb: openpyxl.Workbook) -> None:
    ws = wb[ABA_DRE]
    cel = ws.cell(row=4, column=_MES_COL_INI + 64)
    assert "Steady" in str(cel.value)
    # A coluna de steady e FORMULA (INDEX/MATCH pelo mes de referencia), nao um
    # numero cravado — se a maturacao mudar, ela acompanha.
    formula = ws.cell(row=_DRE_ROW["faturamento"], column=_MES_COL_INI + 64).value
    assert str(formula).startswith("=INDEX(")
    assert "MATCH(" in str(formula)


# ---------------------------------------------------------------------------
# Sublinhas exigidas
# ---------------------------------------------------------------------------


def test_dre_tem_as_4_sublinhas_de_custo_variavel(wb: openpyxl.Workbook) -> None:
    ws = wb[ABA_DRE]
    esperado = {
        "cvar_royalties": "royalties",
        "cvar_marketing": "marketing",
        "cvar_manutencao": "manutencao",
        "cvar_cartoes": "cartao",
    }
    assert len(esperado) == 4
    for key, trecho in esperado.items():
        rotulo = _norm(ws.cell(row=_DRE_ROW[key], column=1).value)
        assert trecho in rotulo, f"{key}: {rotulo!r}"


def test_dre_detalha_outros_fixos_em_linhas_redondas(wb: openpyxl.Workbook) -> None:
    """As componentes fecham EXATAMENTE no total do motor, sem rateio.

    O comentario do config.py listava SETE componentes somando R$ 40.150, contra os
    R$ 38.150 da constante; as SEIS reais fecham no centavo e o "Outros (2.000)" era
    espurio. Importa porque ratear as sete para fechar no total dava "IPTU R$ 1.900,37"
    e "Telefone R$ 475,09" — valores que ninguem defende linha a linha na frente de um
    investidor, que e o proposito desta planilha.
    """
    ws = wb[ABA_DRE]
    assert len(_OUTROS_FIXOS_DECOMP) == 6
    soma = sum(v for _k, _l, v in _OUTROS_FIXOS_DECOMP)
    assert soma == pytest.approx(SIM_OUTROS_FIXOS_MES, abs=0.01), (
        f"a decomposicao soma {soma} e a constante vale {SIM_OUTROS_FIXOS_MES}: "
        "se divergirem, a planilha rateia e os valores saem quebrados"
    )
    for key, label, valor in _OUTROS_FIXOS_DECOMP:
        rotulo = _norm(ws.cell(row=_DRE_ROW[key], column=1).value)
        assert _norm(label) in rotulo, f"{key}: {rotulo!r}"
        # Valores REDONDOS: nada de centavo vindo de rateio.
        assert valor == round(valor, 2) and valor % 50 == 0, f"{label} = {valor}"


def test_fluxo_detalha_o_financiamento_e_o_investimento(wb: openpyxl.Workbook) -> None:
    ws = wb[ABA_FLUXO]
    for key in ("saldo_inicial", "juros", "pmt", "amortizacao", "saldo_final",
                "inv_obra", "inv_franquia", "inv_equip_vista", "payback"):
        assert ws.cell(row=_FLX_ROW[key], column=1).value, f"linha {key} sem rótulo"


# ---------------------------------------------------------------------------
# Sao FORMULAS, nao valores cravados
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "aba,linhas",
    [
        (ABA_DRE, ("faturamento", "receita_liquida", "impostos", "cvar_total", "folha",
                   "outros_total", "aluguel", "custos_op", "ebitda", "margem_ebitda",
                   "ir_total", "juros", "resultado_desalav")),
        (ABA_FLUXO, ("ebitda", "ir_csll", "pmt", "amortizacao", "saldo_final",
                     "investimento", "fcf", "fcf_acumulado")),
    ],
)
def test_linhas_da_linha_do_tempo_sao_formulas(
    wb: openpyxl.Workbook, aba: str, linhas: tuple[str, ...]
) -> None:
    ws = wb[aba]
    mapa = _DRE_ROW if aba == ABA_DRE else _FLX_ROW
    for key in linhas:
        for j in range(64):
            v = ws.cell(row=mapa[key], column=_MES_COL_INI + j).value
            assert isinstance(v, str) and v.startswith("="), (
                f"{aba}/{key} mês {j}: esperava fórmula, veio {v!r}"
            )


def test_resumo_nao_tem_numero_digitado(wb: openpyxl.Workbook) -> None:
    ws = wb[ABA_RESUMO]
    formulas = 0
    for row in range(_RESUMO_ROW_INI, ws.max_row + 1):
        v = ws.cell(row=row, column=2).value
        if v is None:
            continue
        assert isinstance(v, str) and v.startswith("="), (
            f"Resumo linha {row}: KPI cravado como valor ({v!r})"
        )
        formulas += 1
    assert formulas >= 25


def test_folha_calcula_custo_por_cargo_com_formula(wb: openpyxl.Workbook) -> None:
    ws = wb[ABA_FOLHA]
    # Qtd x Salario x (1 + Encargos) por cargo, e TOTAL somando.
    achou = 0
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=5).value
        if isinstance(v, str) and re.fullmatch(r"=B\d+\*C\d+\*\(1\+D\d+\)", v):
            achou += 1
    assert achou >= 5, "quadro de pessoal deveria calcular o custo por fórmula"
    assert any(
        isinstance(ws.cell(row=r, column=5).value, str)
        and str(ws.cell(row=r, column=5).value).startswith("=SUM(E")
        for r in range(1, ws.max_row + 1)
    )


def _premissas_por_rotulo(wb: openpyxl.Workbook) -> dict[str, object]:
    """Rotulo normalizado -> valor BRUTO da celula B (formula como str, ou o numero)."""
    ws = wb[ABA_PREMISSAS]
    return {
        _norm(ws.cell(row=r, column=1).value or ""): ws.cell(row=r, column=2).value
        for r in range(4, ws.max_row + 1)
    }


def _ref_premissa(wb: openpyxl.Workbook, rotulo_norm: str) -> str:
    ws = wb[ABA_PREMISSAS]
    for r in range(4, ws.max_row + 1):
        if _norm(ws.cell(row=r, column=1).value or "").startswith(rotulo_norm):
            return f"{ABA_PREMISSAS}!$B${r}"
    raise AssertionError(f"premissa {rotulo_norm!r} não existe na aba {ABA_PREMISSAS}")


def _usa(formula: object, ref: str) -> bool:
    """A formula referencia EXATAMENTE esta celula?

    Substring crua daria falso positivo: `Premissas!$B$6` casa dentro de
    `Premissas!$B$66`. O lookahead fecha a referencia no fim do numero da linha.
    """
    return re.search(re.escape(ref) + r"(?!\d)", str(formula)) is not None


def test_dre_le_a_folha_de_uma_unica_celula_de_premissas(wb: openpyxl.Workbook) -> None:
    """A folha da DRE e UMA celula de Premissas x reajuste, igual em todos os meses.

    O interruptor de modo continua existindo, mas mudou de lugar: quem alterna entre
    `folha_pct x faturamento MADURO` e o TOTAL do quadro e a celula "Folha mensal FIXA"
    da aba Premissas. A DRE nao decide mais nada por coluna — se ela voltasse a decidir,
    voltaria tambem o defeito de dimensionar a folha pelo faturamento DAQUELE mes.
    """
    ref_folha_fixa = _ref_premissa(wb, "folha mensal fixa")
    ws = wb[ABA_DRE]
    for mes in (1, 8, 12, 13, 60):
        letra = get_column_letter(_col(mes))
        formula = str(ws.cell(row=_DRE_ROW["folha"], column=_col(mes)).value)
        assert _usa(formula, ref_folha_fixa), f"M{mes}: {formula!r}"
        # So o reajuste anual de custos entra; o faturamento do mes NAO.
        assert f"{letra}{_DRE_ROW['f_custos']}" in formula, f"M{mes}: {formula!r}"
        assert f"{letra}{_DRE_ROW['faturamento']}" not in formula, (
            f"M{mes}: a folha voltou a depender do faturamento do mês: {formula!r}"
        )

    # A ponte Premissas -> Folha: uma celula aponta para o interruptor, outra para o
    # TOTAL do quadro, e a "Folha mensal FIXA" combina as duas com o % do maduro.
    pontes = _premissas_por_rotulo(wb)
    assert pontes["modo da folha (interruptor)"] == f"={ABA_FOLHA}!$B${_FOLHA_MODO_ROW}"
    assert str(pontes["total do quadro de pessoal (aba folha)"]).startswith(
        f"={ABA_FOLHA}!$E$"
    )
    folha_fixa = str(pontes["folha mensal fixa (vale desde o mes 1)"])
    assert "quadro de pessoal" in folha_fixa
    assert _usa(folha_fixa, _ref_premissa(wb, "faturamento maduro"))
    assert _usa(folha_fixa, _ref_premissa(wb, "folha como % do faturamento maduro"))


def test_folha_fixa_e_a_mesma_formula_nos_64_meses(wb: openpyxl.Workbook) -> None:
    """Um valor unico repetido: a folha nao pode variar de mes para mes por si.

    A unica diferenca legitima entre as colunas e o fator de reajuste, que e uma
    celula DA PROPRIA coluna — entao trocando o nome da coluna as formulas coincidem.
    """
    ws = wb[ABA_DRE]
    normalizadas = set()
    for mes in [-4, -3, -2, -1, *range(1, 61)]:
        col = _col(mes)
        letra = get_column_letter(col)
        formula = str(ws.cell(row=_DRE_ROW["folha"], column=col).value)
        # Troca SO as referencias relativas da propria coluna (`F11` -> `@11`); um
        # `replace` cru trocaria tambem o "F" de "IF" e o "B" de "$B$51".
        normalizadas.add(re.sub(rf"(?<![A-Z$]){letra}(\d+)", r"@\1", formula))
    assert len(normalizadas) == 1, f"a folha tem {len(normalizadas)} fórmulas diferentes"


def test_faturamento_maduro_e_a_base_da_folha_a_precos_do_ano_1(
    wb: openpyxl.Workbook,
) -> None:
    """A base e demanda x receita por aluno + personal, sem fator de reajuste."""
    formula = str(
        _premissas_por_rotulo(wb)["faturamento maduro (base de dimensionamento da folha)"]
    )
    assert _usa(formula, _ref_premissa(wb, "demanda total"))
    assert _usa(formula, _ref_premissa(wb, "receita por aluno total"))
    assert _usa(formula, _ref_premissa(wb, "receita fixa de personal"))
    # Nao pode olhar nenhuma coluna de mes: se olhasse, o reajuste entraria duas vezes.
    assert ABA_DRE not in formula and ABA_FLUXO not in formula


# ---------------------------------------------------------------------------
# Interruptor de modo da folha
# ---------------------------------------------------------------------------


def test_interruptor_modo_da_folha_existe_com_default_percentual(
    wb: openpyxl.Workbook,
) -> None:
    ws = wb[ABA_FOLHA]
    assert _norm(ws.cell(row=_FOLHA_MODO_ROW, column=1).value) == "modo da folha"
    assert ws.cell(row=_FOLHA_MODO_ROW, column=2).value == "percentual do faturamento maduro"


def test_rotulo_do_modo_percentual_nao_mente_sobre_a_base(wb: openpyxl.Workbook) -> None:
    """O rotulo era "percentual da receita" e passou a mentir.

    Com a folha FIXA, o percentual e do faturamento MADURO e o valor nao acompanha a
    receita do mes — chamar o modo de "percentual da receita" faria quem abre a planilha
    esperar uma folha que encolhe na rampa.
    """
    assert _MODOS_FOLHA[0] == "percentual do faturamento maduro"
    assert "maduro" in _norm(_MODOS_FOLHA[0])
    ws = wb[ABA_FOLHA]
    texto = _norm(
        " ".join(str(ws.cell(row=r, column=1).value or "") for r in range(1, _FOLHA_MODO_ROW + 3))
    )
    assert "maduro" in texto, "a aba tem de dizer, em texto visível, que a base é o maduro"
    assert "nao escala com a rampa" in texto or "nao acompanha a rampa" in texto


def test_interruptor_modo_da_folha_tem_validacao_de_lista(wb: openpyxl.Workbook) -> None:
    ws = wb[ABA_FOLHA]
    alvo = f"B{_FOLHA_MODO_ROW}"
    candidatos = [
        dv for dv in ws.data_validations.dataValidation
        if dv.type == "list" and alvo in str(dv.sqref)
    ]
    assert candidatos, "interruptor sem validação de lista"
    dv = candidatos[0]
    assert "percentual do faturamento maduro" in dv.formula1
    assert "quadro de pessoal" in dv.formula1


def test_folha_avisa_que_os_salarios_sao_estimativa(wb: openpyxl.Workbook) -> None:
    ws = wb[ABA_FOLHA]
    texto = _norm(
        " ".join(
            str(ws.cell(row=r, column=1).value or "") for r in range(1, _FOLHA_MODO_ROW)
        )
    )
    assert "estimativa" in texto, "a aba tem de avisar, em texto visível, que é estimativa"


def test_quadro_default_fica_proximo_de_17pct_da_bruta_do_golden(resultado) -> None:
    from motor_expansao.dimensionamento.simulador_xlsx import (
        _ENCARGOS_DEFAULT,
        _QUADRO_PESSOAL_DEFAULT,
    )

    total = sum(q * s for _c, q, s in _QUADRO_PESSOAL_DEFAULT) * (1 + _ENCARGOS_DEFAULT)
    # A folha do caso golden e 17% da bruta; o quadro estimado nao pode estar longe,
    # senao abrir o detalhe na frente do investidor mostraria outra empresa.
    assert abs(total - resultado.folha_mensal) < 500.0, (
        f"quadro estimado R$ {total:,.2f} vs folha do motor R$ {resultado.folha_mensal:,.2f}"
    )


# ---------------------------------------------------------------------------
# Folha virou custo FIXO: o "k" e o custo fixo do break-even
# ---------------------------------------------------------------------------


def test_k_nao_subtrai_mais_a_folha(wb: openpyxl.Workbook) -> None:
    """`k = (1-deducoes)*(1-impostos-cvar)`, SEM o termo da folha.

    Enquanto a folha era percentual da receita ela era subtraida aqui. Deixar o termo
    depois de a folha virar custo fixo a contaria DUAS vezes (no k e no custo fixo) e
    devolveria k = 0,628985 no lugar de 0,798985.
    """
    formula = str(_premissas_por_rotulo(wb)["fator receita -> ebitda (k)"])
    for chave in ("devolucoes", "impostos sobre receita (total)", "custo variavel total"):
        assert _usa(formula, _ref_premissa(wb, chave)), f"{chave} ausente de {formula!r}"
    for chave in ("folha como % do faturamento maduro", "folha mensal fixa"):
        assert not _usa(formula, _ref_premissa(wb, chave)), (
            f"a folha voltou para dentro do k: {formula!r}"
        )


def test_custo_fixo_do_break_even_inclui_a_folha(wb: openpyxl.Workbook) -> None:
    """O custo fixo (sem aluguel) = outros fixos + folha FIXA, nos dois modos."""
    formula = str(
        _premissas_por_rotulo(wb)["custo fixo total, sem aluguel (outros fixos + folha)"]
    )
    assert _usa(formula, _ref_premissa(wb, "outros fixos (total)"))
    assert _usa(formula, _ref_premissa(wb, "folha mensal fixa"))


def test_break_even_do_resumo_usa_o_custo_fixo_com_folha(wb: openpyxl.Workbook) -> None:
    ws = wb[ABA_RESUMO]
    ref_cf = _ref_premissa(wb, "custo fixo total")
    ref_k = _ref_premissa(wb, "fator receita -> ebitda")
    achados = 0
    for r in range(_RESUMO_ROW_INI, ws.max_row + 1):
        if not _norm(ws.cell(row=r, column=1).value or "").startswith("break-even"):
            continue
        formula = ws.cell(row=r, column=2).value
        if not (isinstance(formula, str) and formula.startswith("=")):
            continue  # faixa de titulo do bloco "Break-even (...)", sem valor
        assert _usa(formula, ref_cf), f"break-even sem a folha no custo fixo: {formula!r}"
        assert _usa(formula, ref_k)
        achados += 1
    assert achados == 2, "esperava break-even de EBITDA e de caixa"


# ---------------------------------------------------------------------------
# Taxa de franquia PARCELADA (4x sem juros)
# ---------------------------------------------------------------------------


def test_premissas_expoe_parcelas_e_valor_da_parcela_da_franquia(
    wb: openpyxl.Workbook,
) -> None:
    prem = _premissas_por_rotulo(wb)
    assert prem["parcelas da taxa de franquia"] == SIM_PARCELAS_FRANQUIA_DEFAULT == 4
    parcela = str(prem["valor da parcela da franquia"])
    assert _usa(parcela, _ref_premissa(wb, "taxa de franquia"))
    assert _usa(parcela, _ref_premissa(wb, "parcelas da taxa de franquia"))


def test_fluxo_lanca_a_parcela_da_franquia_e_nao_a_taxa_inteira(
    wb: openpyxl.Workbook,
) -> None:
    """A linha do fluxo tem de olhar as PARCELAS, nao um "mes_contrato = 1"."""
    ws = wb[ABA_FLUXO]
    ref_parcelas = _ref_premissa(wb, "parcelas da taxa de franquia")
    ref_parcela = _ref_premissa(wb, "valor da parcela da franquia")
    for mes in (-4, -2, -1, 1):
        formula = str(ws.cell(row=_FLX_ROW["inv_franquia"], column=_col(mes)).value)
        assert _usa(formula, ref_parcelas), f"M{mes}: {formula!r}"
        assert _usa(formula, ref_parcela), f"M{mes}: {formula!r}"
        assert f"{_FLX_ROW['mes_contrato']}" in formula


def test_investimento_tem_parcelas_e_valor_da_parcela_da_franquia(
    wb: openpyxl.Workbook,
) -> None:
    ws = wb[ABA_INVESTIMENTO]
    rotulos = {
        k: _norm(ws.cell(row=_INVEST_ROW[k], column=1).value or "")
        for k in ("parcelas_franquia", "franquia_parcela")
    }
    assert rotulos["parcelas_franquia"] == "parcelas da franquia"
    assert rotulos["franquia_parcela"] == "valor da parcela da franquia"
    for k in rotulos:
        v = ws.cell(row=_INVEST_ROW[k], column=2).value
        assert isinstance(v, str) and v.startswith("="), f"{k} cravado: {v!r}"


# ---------------------------------------------------------------------------
# Resumo: saida de dinheiro em VERMELHO (pedido de Felipe, 2026-07-24)
# ---------------------------------------------------------------------------


class _RefsQualquer(dict):
    """Refs de premissa que resolvem qualquer chave — só a ORDEM das linhas importa."""

    def __missing__(self, chave: str) -> str:
        return f"{ABA_PREMISSAS}!$B$999"


def _resumo_por_key(wb: openpyxl.Workbook) -> dict[str, int]:
    """Mapa `key -> linha` do Resumo, reconstruido pela MESMA ordem que o writer usa.

    Reconstruir e melhor que casar rotulo por rotulo: se alguem inserir uma linha no
    meio, o mapa acompanha e os testes de cor continuam apontando para a linha certa.
    """
    from motor_expansao.dimensionamento.simulador_xlsx import _linhas_resumo

    meses = [-4, -3, -2, -1, *range(1, 61)]
    linhas = _linhas_resumo(meses, _RefsQualquer(), "BM", "BN")
    out: dict[str, int] = {}
    r = _RESUMO_ROW_INI
    for key, *_resto in linhas:
        if key:
            out[key] = r
        r += 1
    # Consistencia: a linha reconstruida tem de ter, na aba, uma formula na coluna B.
    ws = wb[ABA_RESUMO]
    for key, row in out.items():
        v = ws.cell(row=row, column=2).value
        assert isinstance(v, str) and v.startswith("="), f"{key}: linha {row} = {v!r}"
    return out


def _cf_por_faixa(ws) -> dict[str, list]:
    return {str(cf.sqref): list(cf.rules) for cf in ws.conditional_formatting}


def test_resumo_pinta_saida_de_dinheiro_em_vermelho(wb: openpyxl.Workbook) -> None:
    ws = wb[ABA_RESUMO]
    linhas = _resumo_por_key(wb)
    assert _RESUMO_SAIDA <= set(linhas), "key de saída sem linha no Resumo"
    for key in sorted(_RESUMO_SAIDA):
        cel = ws.cell(row=linhas[key], column=2)
        assert cel.font.color is not None and cel.font.color.rgb == _VERMELHO_FONTE, (
            f"{key} (linha {linhas[key]}) deveria estar em vermelho: {cel.font.color}"
        )


def test_resumo_nao_pinta_entrada_nem_resultado_de_vermelho_fixo(
    wb: openpyxl.Workbook,
) -> None:
    """Faturamento, receitas, EBITDA, VPL e acumulado NAO levam vermelho fixo."""
    ws = wb[ABA_RESUMO]
    linhas = _resumo_por_key(wb)
    neutras = (
        "faturamento", "receita_anuidade", "receita_liquida", "receita_pos_impostos",
        "ebitda", "vpl", "acumulado_m60", "ticket_blended", "receita_por_aluno",
        "break_even_ebitda", "teto_teto",
    )
    for key in neutras:
        cel = ws.cell(row=linhas[key], column=2)
        cor = cel.font.color.rgb if cel.font.color is not None else None
        assert cor != _VERMELHO_FONTE, f"{key} não é saída de dinheiro, mas está vermelho"


def test_valor_que_alterna_de_sinal_usa_formatacao_condicional(
    wb: openpyxl.Workbook,
) -> None:
    """EBITDA/FCF/VPL/TIR: vermelho SO quando negativo, senao a planilha mentiria."""
    ws = wb[ABA_RESUMO]
    linhas = _resumo_por_key(wb)
    faixas = _cf_por_faixa(ws)
    alvo = {f"B{linhas[k]}" for k in _RESUMO_ALTERNA}
    cobertas = {ref for ref in faixas if ref in alvo}
    assert cobertas == alvo, f"sem formatação condicional: {sorted(alvo - cobertas)}"
    for ref, regras in faixas.items():
        if ref not in alvo:
            continue
        assert any(
            r.operator == "lessThan" and r.dxf is not None
            and r.dxf.font is not None
            and r.dxf.font.color is not None
            and r.dxf.font.color.rgb == _VERMELHO_FONTE
            for r in regras
        ), f"{ref}: regra de negativo em vermelho ausente"


def test_resumo_explica_a_convencao_de_cor(wb: openpyxl.Workbook) -> None:
    ws = wb[ABA_RESUMO]
    texto = _norm(" ".join(str(ws.cell(row=r, column=1).value or "") for r in range(1, 5)))
    assert "vermelho" in texto and "saida de dinheiro" in texto


# ---------------------------------------------------------------------------
# Aba Afericao — o parametro de comparacao dentro do arquivo
# ---------------------------------------------------------------------------


def _afericao(wb: openpyxl.Workbook) -> dict[str, tuple[object, object]]:
    ws = wb[ABA_AFERICAO]
    out: dict[str, tuple[object, object]] = {}
    for row in range(1, ws.max_row + 1):
        rotulo = ws.cell(row=row, column=1).value
        motor = ws.cell(row=row, column=2).value
        formula = ws.cell(row=row, column=3).value
        if rotulo and isinstance(formula, str) and formula.startswith("="):
            out[_norm(rotulo)] = (motor, formula)
    return out


def test_afericao_cobre_ao_menos_20_kpis(wb: openpyxl.Workbook) -> None:
    assert len(_afericao(wb)) >= 20


def test_afericao_explica_para_que_serve(wb: openpyxl.Workbook) -> None:
    ws = wb[ABA_AFERICAO]
    cabecalho = _norm(
        " ".join(str(ws.cell(row=r, column=1).value or "") for r in range(1, 5))
    )
    assert "motor" in cabecalho and "formula" in cabecalho


def test_afericao_tem_delta_e_status_como_formula(wb: openpyxl.Workbook) -> None:
    ws = wb[ABA_AFERICAO]
    vistos = 0
    for row in range(1, ws.max_row + 1):
        if not ws.cell(row=row, column=1).value:
            continue
        delta = ws.cell(row=row, column=4).value
        status = ws.cell(row=row, column=5).value
        if not (isinstance(delta, str) and delta.startswith("=")):
            continue
        assert "C" in delta and "B" in delta, delta
        assert isinstance(status, str) and "DIVERGENTE" in status and "OK" in status
        assert "0.01" in status, "a tolerância declarada é R$ 0,01"
        vistos += 1
    assert vistos >= 20


def test_afericao_grava_os_valores_do_motor_corretos(wb: openpyxl.Workbook, resultado) -> None:
    """O parametro de comparacao dentro do arquivo tem de ser o numero do motor."""
    afer = _afericao(wb)
    serie = {int(r["mes"]): r for r in resultado.serie_mensal}
    steady = serie[int(resultado.mes_referencia_steady)]

    esperado = {
        "mes de referencia do steady": float(resultado.mes_referencia_steady),
        "alunos totais no mes de referencia": steady["alunos_total"],
        "faturamento bruto (steady)": resultado.faturamento_mensal_steady,
        "dos quais receita de anuidade": resultado.receita_anuidade_mensal,
        "deducoes (devolucoes)": steady["deducoes"],
        "receita liquida": resultado.receita_liquida,
        "impostos sobre receita": steady["impostos"],
        "receita pos-impostos": resultado.receita_pos_impostos,
        "custo variavel": resultado.custos_variaveis_mensal,
        "folha": resultado.folha_mensal,
        # Folha FIXA: a base madura, o R$ dimensionado e o valor DO MES 1 (que era o
        # defeito — antes o mes 1 pagava menos folha porque a rampa nao tinha chegado).
        "faturamento maduro (base de dimensionamento da folha)": _premissas().faturamento_maduro(
            _DEMANDA
        ),
        "folha fixa dimensionada pelo faturamento maduro": _premissas().folha_fixa_mes(_DEMANDA),
        "folha no mes 1 (a folha nao acompanha a rampa)": serie[1]["folha"],
        "custo fixo total sem aluguel (outros fixos + folha)": _premissas().custo_fixo_total_mes(
            _DEMANDA
        ),
        "fator receita -> ebitda (k), sem a folha": _premissas().fator_receita_para_ebitda,
        "outros custos fixos": steady["outros_fixos"],
        "aluguel": steady["aluguel"],
        "custos operacionais totais": resultado.custos_op_mensal,
        "ebitda": resultado.ebitda_mensal,
        "margem ebitda": resultado.margem_ebitda_pct,
        "ir/csll": resultado.ir_csll_mensal,
        "despesa financeira (juros do mes de referencia)": resultado.despesa_financeira_mensal,
        "resultado desalavancado apos ir": resultado.resultado_apos_ir_mensal,
        "pmt do financiamento": resultado.pmt_mensal,
        "juros totais do financiamento": resultado.juros_totais,
        "capex total": resultado.capex_total,
        "investimento total": resultado.investimento_total,
        "break-even de ebitda (alunos totais)": resultado.alunos_break_even_total,
        "break-even de caixa (alunos totais)": resultado.alunos_break_even_caixa_total,
        "payback (meses)": resultado.payback_meses,
        "tir mensal": resultado.tir_mensal,
        "tir anual": resultado.tir_anual,
        "vpl na taxa de desconto": resultado.vpl,
        "fcf acumulado no fim do horizonte": resultado.acumulado_mes_final,
        "retorno anual desalavancado": resultado.retorno_anual_desalavancado,
        "retorno anual do equity": resultado.retorno_anual_equity,
        "ticket blended": resultado.ticket_blended,
        "aluguel-teto - faixa ideal": resultado.aluguel_teto["ideal"],
        "aluguel-teto - teto (canonico)": resultado.aluguel_teto["teto"],
        "aluguel-teto - excecao": resultado.aluguel_teto["excecao"],
        "ebitda do mes 1 (negativo por construcao)": serie[1]["ebitda_mensal"],
        "ebitda do mes 4 (meio da rampa)": serie[4]["ebitda_mensal"],
        "ebitda do mes 8 (fim da rampa)": serie[8]["ebitda_mensal"],
        # Franquia PARCELADA: o M-4 e o M-1 desembolsam o MESMO valor (obra + parcela),
        # e nada da franquia vaza para o mes 1.
        "investimento do m-4 (obra + parcela da franquia)": serie[-4]["investimento"],
        "investimento do m-1 (4a parcela da franquia ainda dentro da obra)":
            serie[-1]["investimento"],
        "investimento no mes 1 (as parcelas nao vazam da pre-abertura)":
            serie[1]["investimento"],
    }
    assert len(esperado) >= 20

    for rotulo, valor_motor in esperado.items():
        assert rotulo in afer, f"rótulo {rotulo!r} ausente da aba Aferição"
        gravado = afer[rotulo][0]
        assert isinstance(gravado, (int, float)), f"{rotulo}: {gravado!r} não é número"
        assert abs(float(gravado) - float(valor_motor)) < 0.01, (
            f"{rotulo}: planilha grava {gravado!r}, motor calcula {valor_motor!r}"
        )


def test_afericao_referencia_celulas_e_nao_repete_o_valor(wb: openpyxl.Workbook) -> None:
    for rotulo, (_motor, formula) in _afericao(wb).items():
        assert re.search(r"=[A-Za-z'].*![$]?[A-Z]{1,3}[$]?\d+", str(formula)), (
            f"{rotulo}: coluna da fórmula deveria referenciar outra aba, veio {formula!r}"
        )


# ---------------------------------------------------------------------------
# Guardrails
# ---------------------------------------------------------------------------


def test_nenhum_valor_estatico_e_inf_ou_nan(wb: openpyxl.Workbook) -> None:
    for aba in wb.sheetnames:
        for row in wb[aba].iter_rows():
            for c in row:
                if isinstance(c.value, float):
                    assert c.value == c.value, f"{aba}!{c.coordinate} é NaN"
                    assert abs(c.value) != float("inf"), f"{aba}!{c.coordinate} é inf"


def test_formulas_usam_virgula_como_separador_de_argumentos(wb: openpyxl.Workbook) -> None:
    """openpyxl grava en-US; ponto e virgula aqui quebraria a abertura no Excel."""
    for aba in wb.sheetnames:
        for row in wb[aba].iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    assert ";" not in v, f"{aba}!{c.coordinate}: {v!r}"


def test_ranges_entre_abas_sao_sintaticamente_validos(wb: openpyxl.Workbook) -> None:
    """Regressao: `Aba!$B$5:Aba!$B$9` e sintaxe INVALIDA (a aba vem uma vez so)."""
    for aba in wb.sheetnames:
        for row in wb[aba].iter_rows():
            for c in row:
                v = c.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                for trecho in re.findall(r"[$]?[A-Z]{1,3}[$]?\d+:[^,)\s]+", v):
                    depois = trecho.split(":", 1)[1]
                    assert "!" not in depois, f"{aba}!{c.coordinate}: {v!r}"


def test_premissas_marca_editavel_derivada_e_fonte(wb: openpyxl.Workbook) -> None:
    ws = wb[ABA_PREMISSAS]
    editaveis = derivadas = 0
    for row in range(4, ws.max_row + 1):
        rotulo = ws.cell(row=row, column=1).value
        valor = ws.cell(row=row, column=2)
        if not rotulo or valor.value is None:
            continue
        quem = ws.cell(row=row, column=5).value
        fonte = ws.cell(row=row, column=4).value
        assert quem, f"linha {row} sem 'Quem pode alterar'"
        assert fonte, f"linha {row} sem 'Fonte'"
        if isinstance(valor.value, str) and valor.value.startswith("="):
            derivadas += 1
        else:
            editaveis += 1
    assert editaveis >= 40
    assert derivadas >= 15


def test_premissas_puxa_os_defaults_do_objeto_recebido() -> None:
    """Nenhuma constante e redigitada: mudar a premissa muda a celula."""
    p = Premissas(ticket_cheio=199.0, aluguel_mes=41_000.0, maturacao_meses=10)
    wb2 = openpyxl.load_workbook(
        BytesIO(gerar_simulador_xlsx(1_500.0, p, **_INVEST)), data_only=False
    )
    ws = wb2[ABA_PREMISSAS]
    achados: dict[str, object] = {}
    for row in range(4, ws.max_row + 1):
        rotulo = _norm(ws.cell(row=row, column=1).value or "")
        achados[rotulo] = ws.cell(row=row, column=2).value
    assert achados["ticket cheio (balcao)"] == 199.0
    assert achados["aluguel"] == 41_000.0
    assert achados["maturacao da rampa"] == 10
    assert achados["demanda total (alunos na maturidade)"] == 1_500.0


def test_folha_absoluta_abre_no_modo_quadro_e_fecha_no_centavo() -> None:
    """Com `pessoal_mes_override`, a folha e ABSOLUTA e o quadro tem de somar exato.

    Nesse caminho o interruptor ja abre em "quadro de pessoal" (a DRE le o TOTAL da
    aba), entao um residuo de arredondamento no rateio dos salarios faria a planilha
    divergir do motor por centavos.
    """
    from motor_expansao.dimensionamento.simulador_xlsx import (
        _ENCARGOS_DEFAULT,
        _FOLHA_CARGO_ROW_INI,
        _quadro_pessoal,
    )

    override = 50_128.16
    p = Premissas(ticket_cheio=147.0, aluguel_mes=30_000.0, pessoal_mes_override=override)
    quadro = _quadro_pessoal(p)
    total = sum(q * s for _c, q, s in quadro) * (1 + _ENCARGOS_DEFAULT)
    assert abs(total - override) < 0.005, f"quadro soma {total!r}, override é {override!r}"

    wb2 = openpyxl.load_workbook(
        BytesIO(gerar_simulador_xlsx(_DEMANDA, p, dict(_INVEST))), data_only=False
    )
    ws = wb2[ABA_FOLHA]
    assert ws.cell(row=_FOLHA_MODO_ROW, column=2).value == "quadro de pessoal"
    assert ws.cell(row=_FOLHA_CARGO_ROW_INI, column=3).value is not None
    afer = _afericao(wb2)
    assert abs(float(afer["folha"][0]) - override) < 0.01


def test_aceita_investimento_como_dict_posicional(resultado) -> None:
    """Contrato com a rota do piloto web: `gerar(demanda, premissas, investimento)`.

    A rota passa o investimento POSICIONAL, como dict, e `rotulo`/`m2` como kwargs
    opcionais. Um nome diferente aqui viraria TypeError em runtime (HTTP 500 no
    lugar do arquivo).
    """
    blob2 = gerar_simulador_xlsx(
        _DEMANDA, _premissas(), dict(_INVEST), rotulo="Ponto A", m2=1800.0
    )
    wb2 = openpyxl.load_workbook(BytesIO(blob2), data_only=False)
    assert wb2.sheetnames == list(ABAS_ESPERADAS)
    # O dict posicional tem de ter REALMENTE alimentado o investimento.
    afer = _afericao(wb2)
    assert abs(float(afer["investimento total"][0]) - resultado.investimento_total) < 0.01
    assert abs(float(afer["pmt do financiamento"][0]) - resultado.pmt_mensal) < 0.01
    # rotulo e m2 aparecem no cabeçalho (enfeite, não entram em conta nenhuma).
    titulo = str(wb2[ABA_PREMISSAS].cell(row=1, column=1).value)
    assert "Ponto A" in titulo and "m²" in titulo


def test_dict_de_investimento_ignora_chave_desconhecida() -> None:
    blob2 = gerar_simulador_xlsx(
        _DEMANDA, _premissas(), {**_INVEST, "chave_que_nao_existe": 1}
    )
    assert blob2[:2] == b"PK"


def test_nomes_definidos_sao_ascii(wb: openpyxl.Workbook) -> None:
    nomes = list(wb.defined_names)
    assert nomes, "nenhum nome definido registrado"
    for n in nomes:
        assert n.isascii(), f"identificador acentuado: {n}"
        assert n.startswith("prem_")


# ---------------------------------------------------------------------------
# RECALCULO REAL das formulas (pacote `formulas`; SKIP quando nao instalado)
#
# Este bloco e o unico que pega formula que COMPILA e devolve numero ERRADO —
# openpyxl nunca calcula nada. Instalar com `python -m pip install formulas`.
# ---------------------------------------------------------------------------

# linha da planilha -> campo da serie do motor.
_DRE_VS_MOTOR = {
    "alunos_total": "alunos_total",
    "alunos_balcao": "alunos_balcao",
    "alunos_agregadores": "alunos_agregadores",
    "faturamento": "faturamento_mensal",
    "rec_anuidade": "receita_anuidade",
    "deducoes": "deducoes",
    "receita_liquida": "receita_liquida",
    "impostos": "impostos",
    "receita_pos_impostos": "receita_pos_impostos",
    "cvar_total": "custos_variaveis",
    "folha": "folha",
    "outros_total": "outros_fixos",
    "aluguel": "aluguel",
    "custo_pre_op": "custo_pre_operacional",
    "custos_op": "custos_op",
    "ebitda": "ebitda_mensal",
    "ir_total": "ir_csll",
    "juros": "juros",
}
_FLX_VS_MOTOR = {
    "ebitda": "ebitda_mensal",
    "ir_csll": "ir_csll",
    "juros": "juros",
    "pmt": "pmt",
    "amortizacao": "amortizacao",
    "investimento": "investimento",
    "fcf": "fcf_mensal",
    "fcf_acumulado": "fcf_acumulado",
}


@pytest.fixture(scope="module")
def ler_celula(blob: bytes, tmp_path_factory: pytest.TempPathFactory):
    """Recalcula o workbook e devolve `ler(aba, "B7") -> valor`.

    SKIP quando o pacote `formulas` nao esta instalado (nao e dependencia do projeto).
    """
    formulas = pytest.importorskip(
        "formulas", reason="instale com `python -m pip install formulas` para recalcular"
    )
    caminho: Path = tmp_path_factory.mktemp("simulador_xlsx") / "simulador.xlsx"
    caminho.write_bytes(blob)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        sol = formulas.ExcelModel().loads(str(caminho)).finish().calculate()
    base = f"'[{caminho.name}]"

    def ler(aba: str, coord: str) -> float:
        chave = f"{base}{aba.upper()}'!{coord}"
        assert chave in sol, f"célula {aba}!{coord} não existe no modelo recalculado"
        valor = sol[chave]
        try:
            valor = valor.value[0, 0]
        except (AttributeError, TypeError, IndexError):
            pass
        return valor

    return ler


def _num(valor: object) -> float:
    assert not isinstance(valor, str), f"a fórmula devolveu texto/erro: {valor!r}"
    return float(valor)  # type: ignore[arg-type]


def test_recalculo_reproduz_a_dre_e_o_fluxo_nos_64_meses(ler_celula, resultado) -> None:
    """As 4.293 fórmulas avaliadas, mês a mês, contra `simular()`."""
    divergencias: list[str] = []
    for j, linha in enumerate(resultado.serie_mensal):
        letra = get_column_letter(_MES_COL_INI + j)
        mes = int(linha["mes"])
        for aba, mapa, rows in (
            (ABA_DRE, _DRE_VS_MOTOR, _DRE_ROW),
            (ABA_FLUXO, _FLX_VS_MOTOR, _FLX_ROW),
        ):
            for key, campo in mapa.items():
                obtido = _num(ler_celula(aba, f"{letra}{rows[key]}"))
                esperado = float(linha[campo])
                if abs(obtido - esperado) > 0.01:
                    divergencias.append(
                        f"{aba}/{key} M{mes}: planilha {obtido:.4f} vs motor {esperado:.4f}"
                    )
    assert not divergencias, "\n".join(divergencias[:30])


def test_recalculo_da_afericao_nao_tem_nenhum_divergente(
    ler_celula, wb: openpyxl.Workbook
) -> None:
    """Coluna C (fórmula recalculada) == coluna B (motor) em TODAS as linhas."""
    ws = wb[ABA_AFERICAO]
    conferidas = 0
    for row in range(1, ws.max_row + 1):
        rotulo = ws.cell(row=row, column=1).value
        motor = ws.cell(row=row, column=2).value
        formula = ws.cell(row=row, column=3).value
        if not (rotulo and isinstance(formula, str) and formula.startswith("=")):
            continue
        assert isinstance(motor, (int, float)), f"{rotulo}: motor não numérico ({motor!r})"
        obtido = _num(ler_celula(ABA_AFERICAO, f"C{row}"))
        assert abs(obtido - float(motor)) < 0.01, (
            f"{rotulo}: fórmula {obtido!r} vs motor {motor!r}"
        )
        conferidas += 1
    assert conferidas >= 40


def test_recalculo_prova_que_a_folha_e_fixa_e_so_reajusta_no_mes_13(
    ler_celula, resultado
) -> None:
    """O que o Felipe reportou: a folha não pode escalar com a unidade."""
    folha_m1 = _num(ler_celula(ABA_DRE, f"{get_column_letter(_col(1))}{_DRE_ROW['folha']}"))
    assert folha_m1 == pytest.approx(resultado.folha_mensal, abs=0.01)
    for mes in range(1, 13):
        atual = _num(ler_celula(ABA_DRE, f"{get_column_letter(_col(mes))}{_DRE_ROW['folha']}"))
        assert atual == pytest.approx(folha_m1, abs=0.01), f"a folha mudou no M{mes}"
    # Único movimento legítimo: o degrau anual de custos a partir do mês 13.
    folha_m13 = _num(ler_celula(ABA_DRE, f"{get_column_letter(_col(13))}{_DRE_ROW['folha']}"))
    assert folha_m13 == pytest.approx(folha_m1 * 1.04, abs=0.01)
    # O faturamento do mês 1 é uma fração do maduro, mas a folha é integral: é isso
    # que faz o EBITDA do mês 1 ser bem mais negativo do que no modelo antigo.
    fat_m1 = _num(ler_celula(ABA_DRE, f"{get_column_letter(_col(1))}{_DRE_ROW['faturamento']}"))
    assert fat_m1 < 0.5 * resultado.faturamento_mensal_steady
    assert folha_m1 > 0.4 * fat_m1


def test_recalculo_prova_a_franquia_parcelada_em_4x(ler_celula) -> None:
    linha = _FLX_ROW["inv_franquia"]
    for mes in (-4, -3, -2, -1):
        parcela = _num(ler_celula(ABA_FLUXO, f"{get_column_letter(_col(mes))}{linha}"))
        assert parcela == pytest.approx(40_000.0, abs=0.01), f"M{mes}"
        total = _num(
            ler_celula(ABA_FLUXO, f"{get_column_letter(_col(mes))}{_FLX_ROW['investimento']}")
        )
        assert total == pytest.approx(190_000.0, abs=0.01), f"M{mes}"
    for mes in (1, 2, 12):
        assert _num(ler_celula(ABA_FLUXO, f"{get_column_letter(_col(mes))}{linha}")) == 0
    # As 4 parcelas somam a taxa CHEIA: parcelar é timing, não desconto nem invenção
    # de dinheiro. A coluna "Total do horizonte" vem logo depois dos 64 meses.
    total_horizonte = get_column_letter(_MES_COL_INI + 64)
    assert _num(ler_celula(ABA_FLUXO, f"{total_horizonte}{linha}")) == pytest.approx(
        _INVEST["taxa_franquia"], abs=0.01
    )


def test_recalculo_do_resumo_bate_com_os_kpis_do_motor(
    ler_celula, wb: openpyxl.Workbook, resultado
) -> None:
    linhas = _resumo_por_key(wb)
    esperado = {
        "faturamento": resultado.faturamento_mensal_steady,
        "folha": resultado.folha_mensal,
        "custos_op": resultado.custos_op_mensal,
        "ebitda": resultado.ebitda_mensal,
        "margem": resultado.margem_ebitda_pct,
        "break_even_ebitda": resultado.alunos_break_even_total,
        "break_even_caixa": resultado.alunos_break_even_caixa_total,
        "payback": resultado.payback_meses,
        "tir_anual": resultado.tir_anual,
        "vpl": resultado.vpl,
        "acumulado_m60": resultado.acumulado_mes_final,
        "ebitda_m1": resultado.serie_mensal[4]["ebitda_mensal"],
        "teto_teto": resultado.aluguel_teto["teto"],
    }
    for key, valor in esperado.items():
        obtido = _num(ler_celula(ABA_RESUMO, f"B{linhas[key]}"))
        tol = 0.0001 if abs(float(valor)) < 1 else 0.01
        assert obtido == pytest.approx(float(valor), abs=tol), key


def test_recalculo_confirma_o_k_e_o_custo_fixo_com_folha(ler_celula, wb) -> None:
    """Os dois lados da mudança: k sem folha, custo fixo COM folha."""
    p = _premissas()
    ref_k = _ref_premissa(wb, "fator receita -> ebitda").split("!")[1].replace("$", "")
    ref_cf = _ref_premissa(wb, "custo fixo total").split("!")[1].replace("$", "")
    ref_folha = _ref_premissa(wb, "folha mensal fixa").split("!")[1].replace("$", "")
    ref_fat = _ref_premissa(wb, "faturamento maduro").split("!")[1].replace("$", "")
    assert _num(ler_celula(ABA_PREMISSAS, ref_k)) == pytest.approx(
        p.fator_receita_para_ebitda, abs=1e-9
    )
    assert _num(ler_celula(ABA_PREMISSAS, ref_k)) == pytest.approx(0.798985, abs=1e-9)
    assert _num(ler_celula(ABA_PREMISSAS, ref_cf)) == pytest.approx(
        p.custo_fixo_total_mes(_DEMANDA), abs=0.01
    )
    assert _num(ler_celula(ABA_PREMISSAS, ref_folha)) == pytest.approx(
        p.folha_fixa_mes(_DEMANDA), abs=0.01
    )
    assert _num(ler_celula(ABA_PREMISSAS, ref_fat)) == pytest.approx(
        p.faturamento_maduro(_DEMANDA), abs=0.01
    )
