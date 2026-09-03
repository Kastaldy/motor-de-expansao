"""Bloco C+ — o aviso do perfil VIAJA no XLSX do simulador (decisao 0.7).

O que este arquivo prova, e por que na planilha e nao no caller:

- **Com `aviso_nota`, o texto aparece numa celula com fundo de ALERTA** (o
  `_nota(..., alerta=True)` que a decisao 0.7 manda reusar — fundo amarelo, negrito),
  na linha 3 da aba Premissas: a primeira aba do arquivo, acima do primeiro bloco. O
  arquivo vai ao locador e ao comite; nota discreta em aba do meio nao e aviso.
- **Sem `aviso_nota`, o texto nao existe em celula nenhuma** e a saida e IGUAL a de
  outra chamada sem aviso (determinismo do gerador): e o que sustenta o criterio de
  aceite "o par de artefatos do Brasil nao muda um byte" — o caller brasileiro passa
  `None` porque `avisos` do perfil BR e `{}`.

O gerador nao conhece pais nenhum (DEC-047): quem decide o texto e o caller, lendo o
perfil da instancia (`web/server/app.py`, `_texto_do_aviso_de_viabilidade`).
"""

from __future__ import annotations

import zipfile
from io import BytesIO

import openpyxl
import pytest

from motor_expansao.dimensionamento.excel_export import _AMARELO_CLR
from motor_expansao.dimensionamento.simulador import Premissas
from motor_expansao.dimensionamento.simulador_xlsx import ABA_PREMISSAS, gerar_simulador_xlsx

# O texto REAL do aviso argentino tem o formato de `texto_curto`; aqui basta um marcador
# inequivoco — o conteudo vem do perfil e e travado por tests/contracts.
_AVISO = "ATENCAO: simulacao com tributacao provisoria (teste do carimbo)."

_DEMANDA = 1200.0


def _premissas() -> Premissas:
    return Premissas(ticket_cheio=147.0, aluguel_mes=30_000.0, maturacao_meses=8)


@pytest.fixture(scope="module", autouse=True)
def _sem_cache_de_valores():
    """Desliga a injecao de valores em cache (lib `formulas`) nestes testes.

    O cache recalcula o arquivo inteiro (~dezenas de segundos) e nao participa do que
    se prova aqui — o carimbo e escrito pelo openpyxl, antes do cache. Sem ele a
    comparacao de determinismo tambem fica sobre o artefato do proprio gerador.
    """
    mp = pytest.MonkeyPatch()
    mp.setenv("MOTOR_SIMULADOR_XLSX_SEM_CACHE", "1")
    yield
    mp.undo()


@pytest.fixture(scope="module")
def blob_sem_aviso() -> bytes:
    return gerar_simulador_xlsx(_DEMANDA, _premissas(), nome_ponto="Ponto Teste")


@pytest.fixture(scope="module")
def blob_com_aviso() -> bytes:
    return gerar_simulador_xlsx(
        _DEMANDA, _premissas(), nome_ponto="Ponto Teste", aviso_nota=_AVISO
    )


def _celulas_com_texto(blob: bytes, texto: str) -> list[tuple[str, str]]:
    """[(aba, coordenada)] de toda celula cujo valor contem `texto`."""
    wb = openpyxl.load_workbook(BytesIO(blob), data_only=False)
    achadas: list[tuple[str, str]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and texto in cell.value:
                    achadas.append((ws.title, cell.coordinate))
    return achadas


def _conteudo_das_planilhas(blob: bytes) -> dict[str, bytes]:
    """{nome do membro do zip: bytes}, SEM `docProps/core.xml`.

    `docProps/core.xml` carrega created/modified (datetime da geracao) e os cabecalhos
    do zip carregam timestamp — sao as duas unicas fontes legitimas de diferenca entre
    duas chamadas iguais. Todo o resto (abas, estilos, nomes) tem de bater byte a byte.
    """
    with zipfile.ZipFile(BytesIO(blob)) as z:
        return {n: z.read(n) for n in sorted(z.namelist()) if n != "docProps/core.xml"}


def test_com_aviso_o_texto_esta_na_linha_3_da_aba_premissas(blob_com_aviso: bytes) -> None:
    achadas = _celulas_com_texto(blob_com_aviso, _AVISO)
    assert (ABA_PREMISSAS, "A3") in achadas, achadas


def test_a_celula_do_aviso_tem_fundo_de_alerta_e_negrito(blob_com_aviso: bytes) -> None:
    """`alerta=True` do `_nota`: fundo amarelo e negrito. Sem isso o aviso vira nota de
    rodape visual — presente no arquivo, invisivel na reuniao."""
    wb = openpyxl.load_workbook(BytesIO(blob_com_aviso), data_only=False)
    cel = wb[ABA_PREMISSAS]["A3"]
    assert cel.fill.start_color.rgb == _AMARELO_CLR
    assert cel.font.bold is True


def test_sem_aviso_o_texto_nao_existe_em_celula_nenhuma(blob_sem_aviso: bytes) -> None:
    assert _celulas_com_texto(blob_sem_aviso, _AVISO) == []
    # E a linha 3 da aba Premissas segue vazia, como sempre foi.
    wb = openpyxl.load_workbook(BytesIO(blob_sem_aviso), data_only=False)
    assert wb[ABA_PREMISSAS]["A3"].value is None


def test_sem_aviso_a_saida_e_deterministica(blob_sem_aviso: bytes) -> None:
    """Duas chamadas sem aviso produzem o MESMO conteudo de planilha. E o alicerce do
    criterio "o artefato do Brasil nao muda um byte": se o gerador nao for
    deterministico, a comparacao com/sem aviso nao prova nada."""
    de_novo = gerar_simulador_xlsx(_DEMANDA, _premissas(), nome_ponto="Ponto Teste")
    assert _conteudo_das_planilhas(de_novo) == _conteudo_das_planilhas(blob_sem_aviso)


def test_o_aviso_so_muda_a_celula_da_nota(blob_sem_aviso: bytes, blob_com_aviso: bytes) -> None:
    """Diff cirurgico ao nivel de CELULA: com aviso, a unica diferenca de conteudo do
    arquivo inteiro e `Premissas!A3`. DRE, Fluxo, Resumo e as demais abas ficam
    identicas — o carimbo nao pode mexer em numero nenhum.

    (A comparacao nao e sobre o XML cru de proposito: uma string nova no arquivo
    desloca os indices do `sharedStrings.xml` de TODAS as abas — ruido de serializacao,
    nao de conteudo.)
    """
    wb_sem = openpyxl.load_workbook(BytesIO(blob_sem_aviso), data_only=False)
    wb_com = openpyxl.load_workbook(BytesIO(blob_com_aviso), data_only=False)
    assert wb_sem.sheetnames == wb_com.sheetnames

    diferentes: list[str] = []
    for nome in wb_sem.sheetnames:
        ws_sem, ws_com = wb_sem[nome], wb_com[nome]
        assert (ws_com.max_row, ws_com.max_column) >= (ws_sem.max_row, ws_sem.max_column)
        for row in ws_com.iter_rows():
            for cell in row:
                valor_sem = ws_sem.cell(row=cell.row, column=cell.column).value
                if valor_sem != cell.value:
                    diferentes.append(f"{nome}!{cell.coordinate}")
    assert diferentes == [f"{ABA_PREMISSAS}!A3"], diferentes
