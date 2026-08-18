"""Valores em cache do simulador XLSX (fix de 2026-08-18).

Defeito de producao reportado por Felipe: a planilha "vinha sem informacoes em
varias abas (DRE, Fluxo de caixa...) mesmo com as premissas preenchidas". Causa:
openpyxl grava formula SEM resultado calculado, e visualizador que nao recalcula
(Excel em Modo de Exibicao Protegida, previews, LibreOffice default) mostra as
abas 100% formula em branco — so a Premissas (valores literais) aparecia.

Estes testes cobrem a etapa `_com_valores_em_cache`: TODA formula ganha `<v>`,
os valores conferem com `simular()`, o kill-switch por env funciona e falha do
calculo degrada para o arquivo original (download nunca quebra).

SKIP sem o pacote `formulas` (mesma regra do nivel 2 de test_simulador_xlsx.py).
"""

from __future__ import annotations

import os
from io import BytesIO

import openpyxl
import pytest

pytest.importorskip("formulas")

from motor_expansao.dimensionamento import simulador_xlsx  # noqa: E402
from motor_expansao.dimensionamento.simulador import Premissas, simular  # noqa: E402
from motor_expansao.dimensionamento.simulador_xlsx import (  # noqa: E402
    _DRE_ROW,
    _FLX_ROW,
    _MES_COL_INI,
    ABA_DRE,
    ABA_FLUXO,
    ABAS_ESPERADAS,
    gerar_simulador_xlsx,
)

# Mesmo caso golden do nivel 2 (Boulevard Londrina).
_DEMANDA = 2304.0
_INVEST = {
    "obra": 600_000.0,
    "parcelas_obra": 4,
    "equipamentos": 1_000_000.0,
    "prazo_equipamentos": 60,
    "juros_equipamentos_am": 0.018,
}
_PREMISSAS = Premissas(ticket_cheio=119.9, aluguel_mes=55_000.0)


@pytest.fixture(scope="module")
def xlsx_com_cache() -> bytes:
    """Gera UMA vez (o recalculo custa ~9s) com a etapa de cache LIGADA.

    O conftest da suite desliga a etapa por env para todo teste; aqui a env e'
    controlada na mao (fixture de modulo nao pode usar monkeypatch de funcao).
    """
    anterior = os.environ.get(simulador_xlsx._ENV_SEM_CACHE)
    os.environ[simulador_xlsx._ENV_SEM_CACHE] = "0"
    try:
        return gerar_simulador_xlsx(_DEMANDA, _PREMISSAS, _INVEST, nome_ponto="Golden")
    finally:
        if anterior is None:
            os.environ.pop(simulador_xlsx._ENV_SEM_CACHE, None)
        else:
            os.environ[simulador_xlsx._ENV_SEM_CACHE] = anterior


def _wb(conteudo: bytes, *, data_only: bool) -> openpyxl.Workbook:
    return openpyxl.load_workbook(BytesIO(conteudo), data_only=data_only)


def test_toda_formula_tem_valor_em_cache(xlsx_com_cache: bytes) -> None:
    """A regressao central: nenhuma celula de formula pode ficar sem valor gravado."""
    wb_formulas = _wb(xlsx_com_cache, data_only=False)
    wb_valores = _wb(xlsx_com_cache, data_only=True)
    sem_valor: list[str] = []
    total = 0
    for aba in ABAS_ESPERADAS:
        ws_f, ws_v = wb_formulas[aba], wb_valores[aba]
        for row in ws_f.iter_rows():
            for celula in row:
                if not (isinstance(celula.value, str) and celula.value.startswith("=")):
                    continue
                total += 1
                if ws_v.cell(row=celula.row, column=celula.column).value is None:
                    sem_valor.append(f"{aba}!{celula.coordinate}")
    assert total > 4000, f"esperava as ~4.600 formulas do arquivo; achei {total}"
    # UNICA excecao permitida: a linha do marcador de payback do Fluxo
    # (`=IF(acum>=0,"PAYBACK","")`) — quando o resultado e' texto VAZIO, em branco
    # E' a exibicao correta, e nao ha valor a gravar.
    linha_payback = _FLX_ROW["payback"]
    inesperadas = [
        ref
        for ref in sem_valor
        if not (
            ref.startswith(f"{ABA_FLUXO}!")
            and int("".join(c for c in ref.split("!")[1] if c.isdigit())) == linha_payback
        )
    ]
    assert not inesperadas, (
        f"{len(inesperadas)} de {total} formulas sem valor em cache fora da excecao "
        f"do marcador de payback (primeiras: {inesperadas[:10]}) — essas celulas "
        "abrem EM BRANCO em preview"
    )


def test_valores_em_cache_conferem_com_o_motor(xlsx_com_cache: bytes) -> None:
    """O cache mostra o numero do MOTOR: DRE (faturamento/EBITDA) e Fluxo (acumulado)
    conferem mes a mes com `simular()` — mesma verdade da aba Afericao."""
    r = simular(_DEMANDA, _PREMISSAS, **_INVEST)
    wb_valores = _wb(xlsx_com_cache, data_only=True)
    dre, flx = wb_valores[ABA_DRE], wb_valores[ABA_FLUXO]
    conferencias = (
        (dre, _DRE_ROW["faturamento"], "faturamento_mensal"),
        (dre, _DRE_ROW["ebitda"], "ebitda_mensal"),
        (flx, _FLX_ROW["fcf_acumulado"], "fcf_acumulado"),
    )
    for ws, linha, chave in conferencias:
        for i, mes_motor in enumerate(r.serie_mensal):
            valor = ws.cell(row=linha, column=_MES_COL_INI + i).value
            assert valor == pytest.approx(mes_motor[chave], rel=1e-6, abs=0.01), (
                f"{ws.title} linha {linha} ({chave}), mes indice {i}: "
                f"cache={valor!r} vs motor={mes_motor[chave]!r}"
            )


def test_exports_concorrentes_ambos_saem_com_cache(monkeypatch: pytest.MonkeyPatch) -> None:
    """Regressao do race achado na revisao adversarial de 2026-08-18: dois exports
    concorrentes no threadpool — o `_LOCK_RECALCULO` serializa o recalculo e os DOIS
    arquivos tem de sair com cache (antes, o perdedor caia no fallback sem cache)."""
    import threading

    monkeypatch.setenv(simulador_xlsx._ENV_SEM_CACHE, "0")
    resultados: list[bytes] = []
    erros: list[BaseException] = []

    def _gera() -> None:
        try:
            resultados.append(gerar_simulador_xlsx(_DEMANDA, _PREMISSAS, _INVEST))
        except BaseException as exc:  # noqa: BLE001 — thread nao propaga sozinha
            erros.append(exc)

    threads = [threading.Thread(target=_gera) for _ in range(2)]
    for t in threads:
        t.start()
    for t in threads:
        t.join(timeout=300)
    assert not erros, f"geracao concorrente levantou: {erros}"
    assert len(resultados) == 2
    for conteudo in resultados:
        dre = _wb(conteudo, data_only=True)[ABA_DRE]
        assert dre.cell(row=_DRE_ROW["faturamento"], column=_MES_COL_INI).value is not None, (
            "export concorrente saiu SEM cache — fallback silencioso do race voltou"
        )


def test_env_desliga_a_etapa(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv(simulador_xlsx._ENV_SEM_CACHE, "1")
    conteudo = gerar_simulador_xlsx(_DEMANDA, _PREMISSAS, _INVEST)
    dre = _wb(conteudo, data_only=True)[ABA_DRE]
    assert dre.cell(row=_DRE_ROW["faturamento"], column=_MES_COL_INI).value is None, (
        "com a env ligada nao pode haver cache (comportamento antigo preservado)"
    )


def test_falha_do_calculo_degrada_para_o_arquivo_original(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """O download NUNCA quebra por causa do cache: erro -> arquivo so com formulas."""
    monkeypatch.delenv(simulador_xlsx._ENV_SEM_CACHE, raising=False)

    def _explode(_caminho: str) -> dict:
        raise RuntimeError("cenario de falha do teste")

    monkeypatch.setattr(simulador_xlsx, "_valores_por_aba", _explode)
    conteudo = gerar_simulador_xlsx(_DEMANDA, _PREMISSAS, _INVEST)

    wb = _wb(conteudo, data_only=False)
    assert list(wb.sheetnames) == list(ABAS_ESPERADAS)
    dre = wb[ABA_DRE]
    celula = dre.cell(row=_DRE_ROW["faturamento"], column=_MES_COL_INI).value
    assert isinstance(celula, str) and celula.startswith("="), "formulas seguem intactas"
