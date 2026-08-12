"""O CLI da ingestao mensal do faturamento do Financeiro.

As funcoes de LEITURA e os portoes vivem em `motor_expansao.dashboard.
rede_faturamento_financeiro` e sao testados em `test_rede_faturamento_financeiro.py`.
Aqui esta o que so' existe no script: achar a planilha na pasta, o digest de procedencia,
e o fluxo do comando -- em especial os dois portoes que decidem se algo e' gravado.
"""

from __future__ import annotations

import hashlib
import importlib.util
import json
import sys
from pathlib import Path

import pandas as pd
import pytest
from click.testing import CliRunner

_REPO = Path(__file__).resolve().parents[2]


def _modulo():
    """Importa o script por caminho: `scripts/` nao e' pacote."""
    caminho = _REPO / "scripts" / "ingerir_faturamento_financeiro.py"
    spec = importlib.util.spec_from_file_location("ingerir_faturamento_financeiro", caminho)
    assert spec and spec.loader
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


CLI = _modulo()


# ---------------------------------------------------------------------------
# Procedencia
# ---------------------------------------------------------------------------


def test_manifesto_acompanha_o_parquet(tmp_path: Path) -> None:
    """Era um caminho FIXO: `--saida` noutro lugar sobrescrevia o manifesto de producao.

    Manifesto dessincronizado e' pior que manifesto nenhum -- ele existe justamente para
    dizer de qual planilha aquele parquet saiu.
    """
    assert CLI.manifesto_de(Path("data/staging/faturamento_financeiro.parquet")).name == (
        "faturamento_financeiro.json"
    )
    fora = tmp_path / "outro" / "teste.parquet"
    assert CLI.manifesto_de(fora) == tmp_path / "outro" / "teste.json"


def test_sha256_do_arquivo(tmp_path: Path) -> None:
    """O digest e' o que liga o parquet a UMA planilha; le em blocos (o arquivo tem 4 MB)."""
    arquivo = tmp_path / "p.xlsx"
    conteudo = b"planilha" * 1000
    arquivo.write_bytes(conteudo)
    assert CLI.sha256_de(arquivo) == hashlib.sha256(conteudo).hexdigest()


# ---------------------------------------------------------------------------
# Achar a planilha na pasta
# ---------------------------------------------------------------------------


def test_pega_o_xlsx_mais_recente(tmp_path: Path) -> None:
    import os
    import time

    antigo, novo = tmp_path / "antigo.xlsx", tmp_path / "novo.xlsx"
    antigo.write_bytes(b"a")
    novo.write_bytes(b"b")
    agora = time.time()
    os.utime(antigo, (agora - 10_000, agora - 10_000))
    os.utime(novo, (agora, agora))
    assert CLI.achar_planilha(tmp_path) == novo


def test_ignora_o_temporario_do_excel(tmp_path: Path) -> None:
    """`~$...xlsx` e' o lock que o Excel deixa com a planilha ABERTA. Nao e' a planilha."""
    (tmp_path / "~$ULTRA.xlsx").write_bytes(b"lock")
    (tmp_path / "ULTRA.xlsx").write_bytes(b"real")
    assert CLI.achar_planilha(tmp_path).name == "ULTRA.xlsx"


@pytest.mark.parametrize("cenario", ["pasta ausente", "pasta vazia"])
def test_sem_planilha_diz_o_que_fazer(tmp_path: Path, cenario: str) -> None:
    pasta = tmp_path / "nao-existe" if cenario == "pasta ausente" else tmp_path
    with pytest.raises(SystemExit) as saida:
        CLI.achar_planilha(pasta)
    assert "planilha" in str(saida.value).lower()


# ---------------------------------------------------------------------------
# O comando
# ---------------------------------------------------------------------------


def _planilha(caminho: Path, competencias: tuple[str, ...]) -> Path:
    """Planilha minima no layout do Financeiro (ver `_planilha` do teste do parser)."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Faturamento & Alunos"
    for j, mes in enumerate(competencias):
        ws.cell(row=4, column=3 + j, value=pd.Timestamp(f"{mes}-01").to_pydatetime())
    ws.cell(row=7, column=1, value=1)
    ws.cell(row=7, column=2, value="01 - AUGUSTA")
    for j in range(len(competencias)):
        ws.cell(row=7, column=3 + j, value=100.0)
        ws.cell(row=8, column=3 + j, value=100.0)
    ws.cell(row=8, column=2, value="VENDAS UX")
    dp = wb.create_sheet("Unidades_UX")
    dp.append(["UNIDADE", "COD_UNIDADE", "UNIDADE_UX"])
    dp.append(["AUGUSTA", "01", "AUGUSTA"])
    wb.save(caminho)
    return caminho


def test_simular_nao_escreve_nada(tmp_path: Path) -> None:
    origem = _planilha(tmp_path / "p.xlsx", ("2026-06", "2026-07"))
    saida = tmp_path / "fat.parquet"

    r = CliRunner().invoke(
        CLI.main,
        ["--planilha", str(origem), "--saida", str(saida), "--hoje", "2026-08-12", "--simular"],
    )

    assert r.exit_code == 0, r.output
    assert not saida.exists()
    assert not CLI.manifesto_de(saida).exists()


def test_mes_aberto_aborta_sem_escrever(tmp_path: Path) -> None:
    """O portao que mais importa: ler cedo grava competencia pela metade como definitiva."""
    origem = _planilha(tmp_path / "p.xlsx", ("2026-06", "2026-07"))
    saida = tmp_path / "fat.parquet"

    r = CliRunner().invoke(
        CLI.main, ["--planilha", str(origem), "--saida", str(saida), "--hoje", "2026-07-20"]
    )

    assert r.exit_code == 1
    assert "ABORTADO" in r.output
    assert not saida.exists(), "nada pode ser gravado quando um portao reprova"


def test_permitir_mes_aberto_e_a_valvula_explicita(tmp_path: Path) -> None:
    origem = _planilha(tmp_path / "p.xlsx", ("2026-06", "2026-07"))
    saida = tmp_path / "fat.parquet"

    r = CliRunner().invoke(
        CLI.main,
        ["--planilha", str(origem), "--saida", str(saida), "--hoje", "2026-07-20",
         "--permitir-mes-aberto"],
    )

    assert r.exit_code == 0, r.output
    assert saida.exists()


def test_grava_parquet_e_manifesto_com_a_procedencia(tmp_path: Path) -> None:
    origem = _planilha(tmp_path / "p.xlsx", ("2026-06", "2026-07"))
    saida = tmp_path / "sub" / "fat.parquet"

    r = CliRunner().invoke(
        CLI.main, ["--planilha", str(origem), "--saida", str(saida), "--hoje", "2026-08-12"]
    )

    assert r.exit_code == 0, r.output
    assert len(pd.read_parquet(saida)) == 2
    manifesto = json.loads(CLI.manifesto_de(saida).read_text(encoding="utf-8"))
    assert manifesto["planilha"] == "p.xlsx"
    assert manifesto["sha256"] == CLI.sha256_de(origem)
    assert manifesto["competencia_fim"] == "2026-07"
    assert manifesto["unidades"] == 1
